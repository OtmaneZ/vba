#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyse détaillée du fichier PLANNING.xlsm
pour comprendre la structure actuelle avant corrections
"""

import openpyxl
from openpyxl import load_workbook
import sys

def analyser_planning():
    try:
        print("=" * 80)
        print("ANALYSE DU FICHIER PLANNING.xlsm")
        print("=" * 80)

        wb = load_workbook('/Users/otmaneboulahia/Documents/Excel-Auto/PLANNING.xlsm',
                          keep_vba=True, data_only=False)

        print(f"\n📋 Feuilles disponibles: {wb.sheetnames}\n")

        # ===== ANALYSE FEUILLE DISPONIBILITES =====
        if "Disponibilites" in wb.sheetnames or "DISPONIBILITES" in wb.sheetnames:
            ws_name = "Disponibilites" if "Disponibilites" in wb.sheetnames else "DISPONIBILITES"
            ws = wb[ws_name]
            print("=" * 80)
            print(f"📅 FEUILLE: {ws_name}")
            print("=" * 80)

            # En-têtes
            print("\n🔍 EN-TÊTES (ligne 1):")
            for col in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(1, col).value
                print(f"  Colonne {col}: '{val}'")

            # Données exemples
            print(f"\n📊 DONNÉES (lignes 2-6 / Total: {ws.max_row} lignes):")
            for row in range(2, min(7, ws.max_row + 1)):
                print(f"\n  Ligne {row}:")
                for col in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    cell_format = ws.cell(row, col).number_format
                    print(f"    Col {col}: '{val}' (format: {cell_format})")

        # ===== ANALYSE FEUILLE VISITES =====
        if "Visites" in wb.sheetnames or "VISITES" in wb.sheetnames:
            ws_name = "Visites" if "Visites" in wb.sheetnames else "VISITES"
            ws = wb[ws_name]
            print("\n" + "=" * 80)
            print(f"🎭 FEUILLE: {ws_name}")
            print("=" * 80)

            # En-têtes
            print("\n🔍 EN-TÊTES (ligne 1):")
            for col in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(1, col).value
                print(f"  Colonne {col}: '{val}'")

            # Données exemples
            print(f"\n📊 DONNÉES (lignes 2-4 / Total: {ws.max_row} lignes):")
            for row in range(2, min(5, ws.max_row + 1)):
                print(f"\n  Ligne {row}:")
                for col in range(1, min(15, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    cell_format = ws.cell(row, col).number_format
                    if col == 3:  # Colonne heure
                        print(f"    Col {col} (HEURE): '{val}' (format: {cell_format}) [Type: {type(val).__name__}]")
                    else:
                        print(f"    Col {col}: '{val}'")

        # ===== ANALYSE FEUILLE SPECIALISATIONS =====
        if "Specialisations" in wb.sheetnames or "SPECIALISATIONS" in wb.sheetnames or "Spécialisations" in wb.sheetnames:
            for possible_name in ["Specialisations", "SPECIALISATIONS", "Spécialisations"]:
                if possible_name in wb.sheetnames:
                    ws_name = possible_name
                    break

            ws = wb[ws_name]
            print("\n" + "=" * 80)
            print(f"⭐ FEUILLE: {ws_name}")
            print("=" * 80)

            # Vérifier visibilité
            sheet = wb[ws_name]
            print(f"\n👁️ VISIBILITÉ: {sheet.sheet_state}")

            # En-têtes
            print("\n🔍 EN-TÊTES:")
            print("  Ligne 1:")
            for col in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(1, col).value
                print(f"    Colonne {col}: '{val}'")

            # Données exemples
            print(f"\n📊 DONNÉES (lignes 2-10 / Total: {ws.max_row} lignes):")
            for row in range(2, min(11, ws.max_row + 1)):
                print(f"\n  Ligne {row}:")
                for col in range(1, min(6, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    print(f"    Col {col}: '{val}'")

        # ===== ANALYSE FEUILLE PLANNING =====
        if "Planning" in wb.sheetnames or "PLANNING" in wb.sheetnames:
            ws_name = "Planning" if "Planning" in wb.sheetnames else "PLANNING"
            ws = wb[ws_name]
            print("\n" + "=" * 80)
            print(f"📋 FEUILLE: {ws_name} (RÉSULTAT)")
            print("=" * 80)

            # En-têtes
            print("\n🔍 EN-TÊTES (ligne 1):")
            for col in range(1, min(12, ws.max_column + 1)):
                val = ws.cell(1, col).value
                print(f"  Colonne {col}: '{val}'")

            # Données exemples
            print(f"\n📊 DONNÉES (lignes 2-5 / Total: {ws.max_row} lignes):")
            for row in range(2, min(6, ws.max_row + 1)):
                print(f"\n  Ligne {row}:")
                for col in range(1, min(12, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    cell_format = ws.cell(row, col).number_format
                    if "HEURE" in str(ws.cell(1, col).value).upper():
                        print(f"    Col {col} (HEURE): '{val}' (format: {cell_format}) [Type: {type(val).__name__}]")
                    elif "GUIDE" in str(ws.cell(1, col).value).upper():
                        print(f"    Col {col} (GUIDES): '{val}'")
                    else:
                        print(f"    Col {col}: '{val}'")

        print("\n" + "=" * 80)
        print("✅ ANALYSE TERMINÉE")
        print("=" * 80)

        wb.close()

    except Exception as e:
        print(f"❌ ERREUR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyser_planning()

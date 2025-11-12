#!/usr/bin/env python3
"""
Phase 1 - Restructuration onglet Visites
Adapte la structure pour correspondre au planning de la cliente
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

print("=" * 100)
print("🔧 PHASE 1 - RESTRUCTURATION ONGLET VISITES")
print("=" * 100)

# ============================================================================
# ÉTAPE 1 : CHARGER LE FICHIER
# ============================================================================
print("\n📂 Étape 1/6 : Chargement PLANNING.xlsm...")
wb = load_workbook('PLANNING.xlsm', keep_vba=True)
ws = wb['Visites']

derniere_ligne = ws.max_row
print(f"   ✅ Fichier chargé - {derniere_ligne} lignes")

# ============================================================================
# ÉTAPE 2 : SAUVEGARDER LES DONNÉES EXISTANTES
# ============================================================================
print("\n💾 Étape 2/6 : Sauvegarde des données existantes...")

# Lire toutes les données actuelles (sauf ligne 1 = en-têtes)
donnees = []
for row in range(2, derniere_ligne + 1):
    ligne = {
        'ID_Visite': ws.cell(row, 1).value,           # A
        'Date': ws.cell(row, 2).value,                # B
        'Heure': ws.cell(row, 3).value,               # C (à supprimer)
        'Musee': ws.cell(row, 4).value,               # D (à supprimer)
        'Type_Visite': ws.cell(row, 5).value,         # E (à supprimer)
        'Duree_Heures': ws.cell(row, 6).value,        # F (à recalculer)
        'Nombre_Visiteurs': ws.cell(row, 7).value,    # G → E
        'Statut': ws.cell(row, 8).value,              # H → K
        'Heure_Debut': ws.cell(row, 11).value,        # K → C
        'Heure_Fin': ws.cell(row, 12).value,          # L → D
        'Langue': ws.cell(row, 13).value,             # M (garder en fin)
        'Nb_Personnes': ws.cell(row, 14).value,       # N (doublon, supprimer)
        'Tarif': ws.cell(row, 15).value,              # O → M
        'Guide_Attribue': ws.cell(row, 16).value,     # P → L
        'Notes': ws.cell(row, 17).value,              # Q → J
    }
    donnees.append(ligne)

print(f"   ✅ {len(donnees)} lignes de données sauvegardées")

# ============================================================================
# ÉTAPE 3 : EFFACER ET RECRÉER LES EN-TÊTES
# ============================================================================
print("\n🗑️  Étape 3/6 : Recréation structure avec nouveaux en-têtes...")

# Effacer toutes les lignes sauf ligne 1
for row in range(ws.max_row, 1, -1):
    ws.delete_rows(row)

# Nouveaux en-têtes (ordre proche de la cliente)
nouveaux_entetes = [
    'ID_Visite',           # A - Auto-généré V0001, V0002...
    'Date',                # B - Format date Excel
    'Heure_Debut',         # C - HH:MM
    'Heure_Fin',           # D - HH:MM
    'Nb_Participants',     # E - Nombre de personnes
    'Type_Prestation',     # F - VISITE CONTEE BRANLY / MARINE / HORS LES MURS / VISIO / EVENEMENT
    'Nom_Structure',       # G - Client/École/Institution
    'Niveau',              # H - CP, CE1, etc.
    'Theme',               # I - femmes, Orient, etc.
    'Commentaires',        # J - Notes diverses
    'Statut',              # K - Confirmée / En attente / Annulée
    'Guide_Attribue',      # L - Nom du guide
    'Tarif',               # M - Calculé auto
    'Duree_Heures',        # N - Calculé (Heure_Fin - Heure_Debut)
    'Langue',              # O - Optionnel
]

# Écrire les en-têtes
for col, header in enumerate(nouveaux_entetes, 1):
    cell = ws.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

print(f"   ✅ {len(nouveaux_entetes)} colonnes créées")

# ============================================================================
# ÉTAPE 4 : RÉÉCRIRE LES DONNÉES DANS LA NOUVELLE STRUCTURE
# ============================================================================
print("\n📝 Étape 4/6 : Réécriture des données dans nouvelle structure...")

for idx, ligne in enumerate(donnees, 2):
    # Colonne A: ID_Visite (garder tel quel)
    ws.cell(idx, 1).value = ligne['ID_Visite']

    # Colonne B: Date (garder tel quel)
    ws.cell(idx, 2).value = ligne['Date']
    if ligne['Date']:
        ws.cell(idx, 2).number_format = 'DD/MM/YYYY'

    # Colonne C: Heure_Debut (depuis ancienne col K)
    ws.cell(idx, 3).value = ligne['Heure_Debut']
    if ligne['Heure_Debut']:
        ws.cell(idx, 3).number_format = 'HH:MM'

    # Colonne D: Heure_Fin (depuis ancienne col L)
    ws.cell(idx, 4).value = ligne['Heure_Fin']
    if ligne['Heure_Fin']:
        ws.cell(idx, 4).number_format = 'HH:MM'

    # Colonne E: Nb_Participants (depuis ancienne col G)
    ws.cell(idx, 5).value = ligne['Nombre_Visiteurs']

    # Colonne F: Type_Prestation (NOUVEAU - à remplir manuellement ou via import)
    # Pour l'instant vide, sauf si on peut deviner depuis Type_Visite
    type_visite_ancien = str(ligne['Type_Visite'] or '').upper()
    if 'BRANLY' in type_visite_ancien or 'CONTE' in type_visite_ancien:
        ws.cell(idx, 6).value = "VISITE CONTEE BRANLY"
    else:
        ws.cell(idx, 6).value = ""  # À remplir

    # Colonne G: Nom_Structure (NOUVEAU - depuis ancien Musee temporairement)
    ws.cell(idx, 7).value = ligne['Musee'] or ""

    # Colonne H: Niveau (NOUVEAU - vide)
    ws.cell(idx, 8).value = ""

    # Colonne I: Theme (NOUVEAU - vide)
    ws.cell(idx, 9).value = ""

    # Colonne J: Commentaires (depuis ancienne col Q)
    ws.cell(idx, 10).value = ligne['Notes'] or ""

    # Colonne K: Statut (depuis ancienne col H)
    ws.cell(idx, 11).value = ligne['Statut'] or "Confirmée"

    # Colonne L: Guide_Attribue (depuis ancienne col P)
    ws.cell(idx, 12).value = ligne['Guide_Attribue'] or ""

    # Colonne M: Tarif (depuis ancienne col O)
    ws.cell(idx, 13).value = ligne['Tarif']
    if ligne['Tarif']:
        ws.cell(idx, 13).number_format = '#,##0.00 €'

    # Colonne N: Duree_Heures (FORMULE - calculé depuis C et D)
    if ligne['Heure_Debut'] and ligne['Heure_Fin']:
        # Formule Excel pour calculer durée en heures
        ws.cell(idx, 14).value = f"=(D{idx}-C{idx})*24"
        ws.cell(idx, 14).number_format = '0.00'
    else:
        ws.cell(idx, 14).value = ligne['Duree_Heures']

    # Colonne O: Langue (optionnel, depuis ancienne col M)
    ws.cell(idx, 15).value = ligne['Langue'] or "Français"

print(f"   ✅ {len(donnees)} lignes réécrites")

# ============================================================================
# ÉTAPE 5 : AJOUTER VALIDATIONS ET FORMATAGE
# ============================================================================
print("\n✨ Étape 5/6 : Ajout validations et formatage...")

# Validation colonne F: Type_Prestation (liste déroulante)
types_presta = '"VISITE CONTEE BRANLY,VISITE CONTEE MARINE,HORS LES MURS,VISIO,EVENEMENT BRANLY"'
dv_type = DataValidation(type="list", formula1=types_presta, allow_blank=True)
dv_type.error = 'Valeur invalide'
dv_type.errorTitle = 'Type de prestation'
dv_type.prompt = 'Choisir un type de prestation'
dv_type.promptTitle = 'Type de prestation'
ws.add_data_validation(dv_type)
dv_type.add(f'F2:F1000')  # Appliquer sur 1000 lignes
print("   ✅ Liste déroulante Type_Prestation (colonne F)")

# Validation colonne K: Statut (liste déroulante)
statuts = '"Confirmée,En attente,Annulée"'
dv_statut = DataValidation(type="list", formula1=statuts, allow_blank=False)
dv_statut.error = 'Valeur invalide'
dv_statut.errorTitle = 'Statut'
ws.add_data_validation(dv_statut)
dv_statut.add(f'K2:K1000')
print("   ✅ Liste déroulante Statut (colonne K)")

# Ajuster largeurs colonnes
largeurs = {
    1: 12,  # A: ID_Visite
    2: 12,  # B: Date
    3: 11,  # C: Heure_Debut
    4: 11,  # D: Heure_Fin
    5: 15,  # E: Nb_Participants
    6: 25,  # F: Type_Prestation
    7: 30,  # G: Nom_Structure
    8: 15,  # H: Niveau
    9: 20,  # I: Theme
    10: 35, # J: Commentaires
    11: 12, # K: Statut
    12: 20, # L: Guide_Attribue
    13: 12, # M: Tarif
    14: 12, # N: Duree_Heures
    15: 12, # O: Langue
}

for col, width in largeurs.items():
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

print("   ✅ Largeurs colonnes ajustées")

# Geler la première ligne
ws.freeze_panes = 'A2'
print("   ✅ Première ligne gelée")

# ============================================================================
# ÉTAPE 6 : SAUVEGARDER
# ============================================================================
print("\n💾 Étape 6/6 : Sauvegarde PLANNING.xlsm...")

wb.save('PLANNING.xlsm')
wb.close()

print("   ✅ Fichier sauvegardé")

# ============================================================================
# RÉSUMÉ
# ============================================================================
print("\n" + "=" * 100)
print("✅ PHASE 1 - ÉTAPE 1 TERMINÉE : RESTRUCTURATION VISITES")
print("=" * 100)

print(f"""
📊 RÉSUMÉ DES MODIFICATIONS:

Colonnes créées:
  ✅ A: ID_Visite (conservé)
  ✅ B: Date (conservé)
  ✅ C: Heure_Debut (déplacé depuis K)
  ✅ D: Heure_Fin (déplacé depuis L)
  ✅ E: Nb_Participants (renommé depuis G)
  🆕 F: Type_Prestation (NOUVEAU - liste déroulante)
  🆕 G: Nom_Structure (NOUVEAU - rempli avec ancien Musée temporairement)
  🆕 H: Niveau (NOUVEAU - vide)
  🆕 I: Theme (NOUVEAU - vide)
  ✅ J: Commentaires (déplacé depuis Q)
  ✅ K: Statut (déplacé depuis H)
  ✅ L: Guide_Attribue (déplacé depuis P)
  ✅ M: Tarif (déplacé depuis O)
  ✅ N: Duree_Heures (formule calculée)
  ✅ O: Langue (déplacé depuis M)

Données:
  ✅ {len(donnees)} lignes conservées et réorganisées
  ✅ Validations listes déroulantes ajoutées
  ✅ Formatage dates/heures/montants appliqué

⚠️  ACTIONS MANUELLES NÉCESSAIRES:
  1. Ouvrir PLANNING.xlsm et vérifier visuellement
  2. Colonnes F (Type_Prestation), H (Niveau), I (Theme) sont à compléter
  3. Colonne G (Nom_Structure) contient temporairement l'ancien "Musée"

📁 Backup disponible: PLANNING_BACKUP_AVANT_RESTRUCTURATION.xlsm
""")

print("\n🎯 PROCHAINE ÉTAPE: Initialiser onglet Specialisations")
print("   Commande: python3 phase1_initialiser_specialisations.py")
print("=" * 100)

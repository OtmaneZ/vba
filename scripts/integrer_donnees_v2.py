"""
Script pour intégrer les données V2 du client :
1. Importer les 79 types de visites dans PLANNING_MUSEE_FINAL.xlsm
2. Pré-catégoriser automatiquement (Individuel/Groupe/Événement/Hors-les-murs/Marine)
3. Ajouter colonne Barème (Standard/Événement BRANLY/Hors-les-murs BRANLY)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

print("🚀 INTÉGRATION DONNÉES V2 - TARIFS ET NOUVEAUX TYPES VISITES")
print("=" * 80)

# Charger les deux fichiers
wb_client = openpyxl.load_workbook("FORMULAIRE_CLIENT_PRO V2.xlsx")
wb_planning = openpyxl.load_workbook("PLANNING_MUSEE_FINAL.xlsm", keep_vba=True)

ws_types_client = wb_client.worksheets[3]  # Types Visites du client
ws_visites_planning = wb_planning["Visites"]  # Feuille Visites du planning

print("\n📥 Import des types de visites depuis V2...")

# Fonction de catégorisation automatique
def detecter_categorie(type_visite, notes=""):
    type_upper = type_visite.upper()
    notes_upper = notes.upper() if notes else ""
    
    # Marine : majuscules ou mots-clés
    if any(word in type_upper for word in ["BULLE", "ZOO", "ABORDAGE", "MARINE", "JOYEUX MERCREDI", "JOURNEES DU PATRIMOINE", "NUIT DE LA LECTURE"]):
        return "Marine"
    
    # Hors-les-murs : dans le nom
    if "HORS LES MURS" in type_upper or "HORS-LES-MURS" in type_upper:
        return "Hors-les-murs"
    
    # Événements : mots-clés
    if any(word in type_upper for word in ["DIMANCHE EN FAMILLE", "TOUS AU MUSEE", "NUIT DES MUSEES", "UN AUTRE NOEL", 
                                             "JARDIN DES CONTES", "WEEK-END DE L'ETHNOLOGIE", "EVENEMENT"]):
        return "Événement"
    
    # Groupe : si mention dans notes
    if "GROUPE" in notes_upper or "ÉTABLISSEMENT" in notes_upper:
        return "Groupe"
    
    # Par défaut : Individuel
    return "Individuel"

# Fonction de détection du barème
def detecter_bareme(type_visite, notes=""):
    type_upper = type_visite.upper()
    notes_upper = notes.upper() if notes else ""
    
    # Événement BRANLY
    if "BRANLY" in type_upper and "EVENEMENT" in type_upper:
        return "Événement BRANLY"
    
    # Hors-les-murs BRANLY
    if "BRANLY" in notes_upper and "HORS-LES-MURS" in notes_upper:
        return "Hors-les-murs BRANLY"
    
    # AUTRE = cas par cas
    if "AUTRE" in type_upper:
        return "Cas par cas"
    
    # Par défaut : Standard
    return "Standard"

# Compter les types dans le fichier client
nb_types_client = 0
for i in range(4, ws_types_client.max_row + 1):
    if ws_types_client.cell(i, 1).value:
        nb_types_client += 1

print(f"   → {nb_types_client} types de visites détectés dans V2")

# Vérifier si colonne Barème existe déjà
derniere_col_visites = ws_visites_planning.max_column
col_bareme = None

# Chercher colonne Barème
for col in range(1, derniere_col_visites + 1):
    if ws_visites_planning.cell(4, col).value and "Barème" in str(ws_visites_planning.cell(4, col).value):
        col_bareme = col
        break

# Si pas trouvée, créer
if not col_bareme:
    col_bareme = derniere_col_visites + 1
    cell = ws_visites_planning.cell(4, col_bareme)
    cell.value = "Barème"
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_visites_planning.column_dimensions[openpyxl.utils.get_column_letter(col_bareme)].width = 20
    print("   → Colonne 'Barème' créée")
else:
    print(f"   → Colonne 'Barème' existante (colonne {openpyxl.utils.get_column_letter(col_bareme)})")

# Trouver colonne Catégorie (créée précédemment)
col_categorie = None
for col in range(1, derniere_col_visites + 2):
    if ws_visites_planning.cell(4, col).value and "Catégorie" in str(ws_visites_planning.cell(4, col).value):
        col_categorie = col
        break

print(f"   → Colonne 'Catégorie' trouvée (colonne {openpyxl.utils.get_column_letter(col_categorie)})")

# Effacer anciennes données visites (garder en-têtes)
derniere_ligne_visites = ws_visites_planning.max_row
if derniere_ligne_visites > 4:
    ws_visites_planning.delete_rows(5, derniere_ligne_visites - 4)
    print(f"   → Anciennes données effacées ({derniere_ligne_visites - 4} lignes)")

# Importer les types depuis V2
print("\n📝 Import et catégorisation automatique...")

ligne_dest = 5
compteurs = {"Individuel": 0, "Groupe": 0, "Événement": 0, "Hors-les-murs": 0, "Marine": 0}
baremes = {"Standard": 0, "Événement BRANLY": 0, "Hors-les-murs BRANLY": 0, "Cas par cas": 0}

for i in range(4, ws_types_client.max_row + 1):
    type_visite = ws_types_client.cell(i, 1).value
    duree = ws_types_client.cell(i, 2).value
    notes = ws_types_client.cell(i, 4).value if ws_types_client.cell(i, 4).value else ""
    
    if type_visite and type_visite != "Type de visite":
        # Détection automatique
        categorie = detecter_categorie(type_visite, notes)
        bareme = detecter_bareme(type_visite, notes)
        
        # Écrire dans planning
        # Colonnes : ID | Date | Heure Début | Heure Fin | Musée | Type | Guide | Nom Guide | Catégorie | Barème
        ws_visites_planning.cell(ligne_dest, 6).value = type_visite  # Type
        ws_visites_planning.cell(ligne_dest, col_categorie).value = categorie
        ws_visites_planning.cell(ligne_dest, col_bareme).value = bareme
        
        # Appliquer code couleur selon catégorie
        couleur = None
        if categorie == "Individuel":
            couleur = "B4C7E7"  # Bleu clair
        elif categorie == "Groupe":
            couleur = "D9E1F2"  # Bleu très clair
        elif categorie == "Événement":
            couleur = "FCE4D6"  # Rose/orange clair
        elif categorie == "Hors-les-murs":
            couleur = "F8CBAD"  # Rouge/orange clair
        elif categorie == "Marine":
            couleur = "002060"  # Bleu foncé
        
        if couleur:
            ws_visites_planning.cell(ligne_dest, 6).fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
            if categorie == "Marine":
                ws_visites_planning.cell(ligne_dest, 6).font = Font(color="FFFFFF", bold=True)
        
        compteurs[categorie] += 1
        baremes[bareme] += 1
        ligne_dest += 1

print(f"\n✅ {ligne_dest - 5} types de visites importés et catégorisés")

print("\n📊 RÉPARTITION CATÉGORIES :")
for cat, nb in compteurs.items():
    print(f"   • {cat:<18} : {nb:>2} visites")

print("\n💰 RÉPARTITION BARÈMES :")
for bar, nb in baremes.items():
    print(f"   • {bar:<25} : {nb:>2} visites")

# Ajouter validation Barème
dv_bareme = DataValidation(type="list", formula1='"Standard,Événement BRANLY,Hors-les-murs BRANLY,Cas par cas"', allow_blank=False)
dv_bareme.error = "Veuillez choisir un barème valide"
dv_bareme.errorTitle = "Barème invalide"
dv_bareme.add(f"{openpyxl.utils.get_column_letter(col_bareme)}5:{openpyxl.utils.get_column_letter(col_bareme)}500")
ws_visites_planning.add_data_validation(dv_bareme)

# Sauvegarder
print("\n💾 Sauvegarde...")
wb_planning.save("PLANNING_MUSEE_FINAL.xlsm")
wb_planning.close()
wb_client.close()

print("\n" + "=" * 80)
print("✅ INTÉGRATION TERMINÉE !")
print("\nCe qui a été fait :")
print("  1️⃣  79 types de visites importés depuis V2")
print("  2️⃣  Catégories détectées automatiquement (Individuel/Groupe/Événement/Hors-les-murs/Marine)")
print("  3️⃣  Barèmes détectés automatiquement (Standard/Événement BRANLY/Hors-les-murs BRANLY/Cas par cas)")
print("  4️⃣  Code couleur visuel appliqué pour faciliter vérification")
print("\nÀ FAIRE MANUELLEMENT :")
print("  ⚠️  Vérifier les catégorisations automatiques (ouvrir Excel et valider)")
print("  ⚠️  Ajuster les barèmes si nécessaire (colonne Barème)")
print("  ⚠️  Remplir colonnes Date, Heure, Musée quand nécessaire")
print("=" * 80)

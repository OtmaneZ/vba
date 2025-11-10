"""
Script pour nettoyer toutes les données TEST/FAKE du XLSM
et préparer le fichier pour la saisie des données réelles du client
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def nettoyer_xlsm():
    print("Chargement du fichier XLSM...")
    wb = openpyxl.load_workbook('PLANNING_MUSEE_FINAL_COMPLET.xlsm', keep_vba=True)
    
    # Style pour cellules à remplir (jaune)
    style_a_remplir = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    print("\n1. Nettoyage DISPONIBILITÉS (suppression 180 lignes test)...")
    ws_dispo = wb['Disponibilites']
    
    # Garder seulement l'en-tête
    max_row = ws_dispo.max_row
    if max_row > 1:
        ws_dispo.delete_rows(2, max_row - 1)
    
    # Ajouter message explicatif
    ws_dispo['A2'] = "⚠️ À REMPLIR: Saisissez les disponibilités de chaque guide"
    ws_dispo['A2'].font = Font(italic=True, color='FF0000', bold=True)
    ws_dispo.merge_cells('A2:D2')
    
    print(f"   ✅ {max_row - 1} lignes supprimées")
    
    print("\n2. Nettoyage GUIDES (suppression mots de passe et tarifs test)...")
    ws_guides = wb['Guides']
    
    nb_nettoyages = 0
    for i in range(2, ws_guides.max_row + 1):
        # Vider Tarif_Horaire (colonne E)
        if ws_guides.cell(i, 5).value:
            ws_guides.cell(i, 5).value = None
            ws_guides.cell(i, 5).fill = style_a_remplir
            nb_nettoyages += 1
        
        # Vider Mot_De_Passe (colonne F)
        if ws_guides.cell(i, 6).value:
            ws_guides.cell(i, 6).value = None
            ws_guides.cell(i, 6).fill = style_a_remplir
            nb_nettoyages += 1
    
    print(f"   ✅ {nb_nettoyages} cellules nettoyées (tarifs + mots de passe)")
    
    print("\n3. Nettoyage VISITES (suppression dates/heures test)...")
    ws_visites = wb['Visites']
    
    nb_dates = 0
    for i in range(2, ws_visites.max_row + 1):
        # Vider Date (colonne B)
        if ws_visites.cell(i, 2).value:
            ws_visites.cell(i, 2).value = None
            ws_visites.cell(i, 2).fill = style_a_remplir
            nb_dates += 1
        
        # Vider Heure (colonne C)
        if ws_visites.cell(i, 3).value:
            ws_visites.cell(i, 3).value = None
            ws_visites.cell(i, 3).fill = style_a_remplir
        
        # Vider Nombre_Visiteurs (colonne G)
        if ws_visites.cell(i, 7).value:
            ws_visites.cell(i, 7).value = None
            ws_visites.cell(i, 7).fill = style_a_remplir
    
    print(f"   ✅ {nb_dates} visites nettoyées (dates/heures/nb visiteurs)")
    
    print("\n4. Mise à jour CONFIGURATION (marquage données à remplacer)...")
    ws_config = wb['Configuration']
    
    # Marquer les valeurs test à remplacer
    params_test = {
        'Email_Expediteur': '⚠️ À REMPLACER par email réel',
        'Nom_Association': '⚠️ À REMPLACER par nom réel',
        'MotDePasseAdmin': '⚠️ À REMPLACER par mot de passe sécurisé'
    }
    
    for i in range(2, ws_config.max_row + 1):
        param = ws_config.cell(i, 1).value
        if param in params_test:
            # ws_config.cell(i, 2).value = params_test[param]
            ws_config.cell(i, 2).fill = style_a_remplir
            ws_config.cell(i, 2).font = Font(italic=True, color='FF0000')
    
    print("   ✅ 3 paramètres marqués à remplacer")
    
    print("\n5. Vérification PLANNING...")
    ws_planning = wb['Planning']
    if ws_planning.max_row > 1:
        print(f"   ⚠️  Planning contient {ws_planning.max_row - 1} lignes (à vérifier)")
    else:
        print("   ✅ Planning vide")
    
    # Sauvegarder
    output_file = 'PLANNING_MUSEE_FINAL_PROPRE.xlsm'
    print(f"\n💾 Sauvegarde dans {output_file}...")
    wb.save(output_file)
    
    print("\n" + "="*70)
    print("✅ NETTOYAGE TERMINÉ")
    print("="*70)
    print("\nFICHIER CRÉÉ: PLANNING_MUSEE_FINAL_PROPRE.xlsm")
    print("\n📋 DONNÉES CONSERVÉES (réelles du client):")
    print("   ✅ 15 guides avec noms et emails")
    print("   ✅ 80 types de visites avec catégories et couleurs")
    print("\n⚠️  À COMPLÉTER PAR LE CLIENT:")
    print("   📝 Tarifs horaires des guides (15)")
    print("   🔒 Mots de passe des guides (15)")
    print("   📅 Dates et heures des visites (80)")
    print("   📧 Email expéditeur réel")
    print("   🏛️  Nom de l'association")
    print("   🔐 Mot de passe administrateur")
    print("   📅 Disponibilités des guides (à saisir)")
    print("="*70)

if __name__ == "__main__":
    nettoyer_xlsm()

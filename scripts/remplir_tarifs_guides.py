"""
Script pour remplir les tarifs guides dans PLANNING_MUSEE_FINAL_PROPRE.xlsm
à partir des informations du formulaire client
"""

import openpyxl
from openpyxl.styles import PatternFill

# Tarifs standards extraits du formulaire client
TARIFS_STANDARDS = {
    "1_visite": 80,
    "2_visites": 110,
    "3_visites": 140
}

TARIFS_BRANLY = {
    "2h": 120,
    "3h": 150,
    "4h": 180
}

TARIFS_HORSLEMURS = {
    "1_visite": 100,
    "2_visites": 130,
    "3_visites": 160
}

def remplir_tarifs_guides():
    """Remplit les tarifs dans la feuille Config du XLSM"""
    
    print("📊 Remplissage des tarifs guides...")
    print("="*60)
    
    # Ouvrir le fichier XLSM
    wb = openpyxl.load_workbook(
        r'C:\Users\otman\Documents\ACTIFS\Vba\PLANNING_MUSEE_FINAL_PROPRE.xlsm',
        keep_vba=True
    )
    
    # Accéder à la feuille Configuration
    if 'Configuration' not in wb.sheetnames:
        print("❌ Feuille 'Configuration' non trouvée")
        return
    
    ws_config = wb['Configuration']
    
    # Supprimer le fond jaune (données complétées)
    white_fill = PatternFill(fill_type=None)
    
    # Remplir les tarifs standards
    print("\n✅ Tarifs standards (visites normales):")
    configs_tarifs = {
        "TARIF_1_VISITE": TARIFS_STANDARDS["1_visite"],
        "TARIF_2_VISITES": TARIFS_STANDARDS["2_visites"],
        "TARIF_3_VISITES": TARIFS_STANDARDS["3_visites"],
        "TARIF_BRANLY_2H": TARIFS_BRANLY["2h"],
        "TARIF_BRANLY_3H": TARIFS_BRANLY["3h"],
        "TARIF_BRANLY_4H": TARIFS_BRANLY["4h"],
        "TARIF_HORSLEMURS_1": TARIFS_HORSLEMURS["1_visite"],
        "TARIF_HORSLEMURS_2": TARIFS_HORSLEMURS["2_visites"],
        "TARIF_HORSLEMURS_3": TARIFS_HORSLEMURS["3_visites"]
    }
    
    tarifs_ajoutes = 0
    
    # Parcourir la feuille Config pour trouver les paramètres
    for row in range(1, ws_config.max_row + 1):
        param_name = ws_config.cell(row, 1).value
        
        if param_name and param_name in configs_tarifs:
            ws_config.cell(row, 2).value = configs_tarifs[param_name]
            ws_config.cell(row, 2).fill = white_fill
            print(f"   • {param_name}: {configs_tarifs[param_name]}€")
            tarifs_ajoutes += 1
    
    # Si les paramètres n'existent pas, les ajouter
    if tarifs_ajoutes == 0:
        print("\n⚠️  Ajout des paramètres de tarifs dans Config...")
        next_row = ws_config.max_row + 2
        
        ws_config.cell(next_row, 1).value = "=== TARIFS GUIDES ==="
        next_row += 1
        
        for param_name, tarif_value in configs_tarifs.items():
            ws_config.cell(next_row, 1).value = param_name
            ws_config.cell(next_row, 2).value = tarif_value
            ws_config.cell(next_row, 3).value = "Tarif journalier guide"
            print(f"   • {param_name}: {tarif_value}€")
            next_row += 1
            tarifs_ajoutes += 1
    
    # Sauvegarder
    wb.save(r'C:\Users\otman\Documents\ACTIFS\Vba\PLANNING_MUSEE_FINAL_PROPRE.xlsm')
    print(f"\n✅ {tarifs_ajoutes} tarifs configurés")
    print("="*60)
    print("✅ Fichier sauvegardé: PLANNING_MUSEE_FINAL_PROPRE.xlsm")
    
    # Afficher le résumé
    print("\n📋 RÉSUMÉ DES TARIFS:")
    print("\n   Tarifs standards (45min):")
    print(f"   • 1 visite/jour: 80€")
    print(f"   • 2 visites/jour: 110€")
    print(f"   • 3 visites/jour: 140€")
    print("\n   Tarifs BRANLY (événements):")
    print(f"   • 2 heures: 120€")
    print(f"   • 3 heures: 150€")
    print(f"   • 4 heures: 180€")
    print("\n   Tarifs hors-les-murs:")
    print(f"   • 1 visite/jour: 100€")
    print(f"   • 2 visites/jour: 130€")
    print(f"   • 3 visites/jour: 160€")

if __name__ == "__main__":
    remplir_tarifs_guides()

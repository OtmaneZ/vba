"""
Script pour ajouter :
1. Feuille "Spécialisations" pour mapper guides ↔ types de visites autorisées
2. Colonne "Catégorie" dans feuille Visites pour le code couleur
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

print("🚀 Ajout Spécialisations + Code Couleur au fichier PLANNING_MUSEE_FINAL.xlsm")
print("=" * 80)

# Charger le fichier
wb = openpyxl.load_workbook("PLANNING_MUSEE_FINAL.xlsm", keep_vba=True)

# ============================================
# 1. CRÉER FEUILLE SPÉCIALISATIONS
# ============================================
print("\n📋 Création feuille 'Spécialisations'...")

# Supprimer si existe déjà
if "Spécialisations" in wb.sheetnames:
    del wb["Spécialisations"]

# Créer la feuille
ws_spec = wb.create_sheet("Spécialisations")

# En-têtes
ws_spec["A1"] = "SPÉCIALISATIONS DES GUIDES"
ws_spec["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_spec["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_spec.merge_cells("A1:C1")

ws_spec["A3"] = "Guide"
ws_spec["B3"] = "Type de visite autorisée"
ws_spec["C3"] = "Notes"

for cell in ["A3", "B3", "C3"]:
    ws_spec[cell].font = Font(bold=True)
    ws_spec[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Largeurs colonnes
ws_spec.column_dimensions["A"].width = 25
ws_spec.column_dimensions["B"].width = 35
ws_spec.column_dimensions["C"].width = 40

# Instructions
ws_spec["A2"] = "⚠️ Listez ici pour chaque guide les types de visites qu'il peut effectuer"
ws_spec["A2"].font = Font(italic=True, color="FF0000")
ws_spec.merge_cells("A2:C2")

# Exemples basés sur les données client
exemples = [
    ("Peggy GENESTIE", "Tous sauf:", "Ne fait PAS: Maman Serpent, Petit Ours, BULLE"),
    ("Hanako DANJO", "Maman Serpent", "Fait UNIQUEMENT visites 3.5 ans + Couleurs + Autour du Monde"),
    ("Hanako DANJO", "Petit Ours", ""),
    ("Hanako DANJO", "Couleurs", ""),
    ("Hanako DANJO", "Autour du Monde", ""),
    ("Hanako DANJO", "Afrique 3.5 ans", ""),
    ("Hanako DANJO", "Amérique 3.5 ans", ""),
    ("Hanako DANJO", "Océanie 3.5 ans", ""),
    ("Hanako DANJO", "Asie 3.5 ans", ""),
    ("Silvia MASSEGUR", "Maman Serpent", "Fait UNIQUEMENT visites 3.5 ans + Orient"),
    ("Silvia MASSEGUR", "Petit Ours", ""),
    ("Silvia MASSEGUR", "Couleurs", ""),
    ("Silvia MASSEGUR", "Autour du Monde", ""),
    ("Silvia MASSEGUR", "Orient", ""),
    ("Marianne (nom à vérifier)", "BULLE", "Fait SEULEMENT visites MARINE"),
    ("Marianne (nom à vérifier)", "ZOO", ""),
    ("Marianne (nom à vérifier)", "A L'ABORDAGE", ""),
    ("Marianne (nom à vérifier)", "Événements MARINE", ""),
    ("Solène ARBEL", "BULLE", "Fait SEULEMENT visites MARINE + autres à définir"),
    ("Solène ARBEL", "ZOO", ""),
    ("Solène ARBEL", "A L'ABORDAGE", ""),
    ("Solène ARBEL", "Événements MARINE", ""),
    ("Shady NAFAR", "À préciser", "Contraintes à définir avec la cliente"),
]

row = 4
for guide, visite, note in exemples:
    ws_spec[f"A{row}"] = guide
    ws_spec[f"B{row}"] = visite
    ws_spec[f"C{row}"] = note
    row += 1

print("✅ Feuille 'Spécialisations' créée avec exemples")

# ============================================
# 2. AJOUTER COLONNE CATÉGORIE DANS VISITES
# ============================================
print("\n🎨 Ajout colonne 'Catégorie' dans feuille 'Visites'...")

ws_visites = wb["Visites"]

# Trouver la dernière colonne utilisée dans l'en-tête (ligne 4)
last_col = ws_visites.max_column
new_col_letter = openpyxl.utils.get_column_letter(last_col + 1)

# Ajouter en-tête "Catégorie"
cell_header = ws_visites[f"{new_col_letter}4"]
cell_header.value = "Catégorie"
cell_header.font = Font(bold=True, color="FFFFFF")
cell_header.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
cell_header.alignment = Alignment(horizontal="center", vertical="center")

# Largeur colonne
ws_visites.column_dimensions[new_col_letter].width = 18

# Créer liste déroulante pour les catégories
categories = '"Individuel,Groupe,Événement,Hors-les-murs,Marine"'
dv = DataValidation(type="list", formula1=categories, allow_blank=False)
dv.error = "Veuillez choisir une catégorie valide"
dv.errorTitle = "Catégorie invalide"
dv.prompt = "Choisissez: Individuel / Groupe / Événement / Hors-les-murs / Marine"
dv.promptTitle = "Sélection catégorie"

# Appliquer la validation sur 500 lignes (largement suffisant)
dv.add(f"{new_col_letter}5:{new_col_letter}504")
ws_visites.add_data_validation(dv)

print(f"✅ Colonne 'Catégorie' ajoutée en colonne {new_col_letter}")
print(f"   → Liste déroulante avec 5 catégories")

# ============================================
# 3. CRÉER FEUILLE INSTRUCTIONS CODE COULEUR
# ============================================
print("\n📘 Création feuille 'Instructions_Couleurs'...")

if "Instructions_Couleurs" in wb.sheetnames:
    del wb["Instructions_Couleurs"]

ws_instr = wb.create_sheet("Instructions_Couleurs")

# Titre
ws_instr["A1"] = "GUIDE D'UTILISATION - CODE COULEUR PLANNING"
ws_instr["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_instr["A1"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ws_instr.merge_cells("A1:D1")

# Instructions
instructions = [
    ("", "", "", ""),
    ("Catégorie", "Couleur", "Formatage", "Utilisation"),
    ("Individuel", "Bleu", "Standard", "Visites pour individuels"),
    ("Groupe", "Bleu clair", "Standard", "Visites pour groupes (avec établissement + niveau scolaire)"),
    ("Événement", "Rose", "Standard", "Événements: Dimanche en famille, Nuit des Musées, etc."),
    ("Hors-les-murs", "Rouge", "Standard", "Prison, hôpital, médiathèque, centre culturel, etc."),
    ("Marine", "Bleu foncé", "GRAS + MAJUSCULES", "Visites MARINE (BULLE, ZOO, A L'ABORDAGE) + événements spéciaux"),
]

for idx, (cat, couleur, fmt, usage) in enumerate(instructions, start=3):
    ws_instr[f"A{idx}"] = cat
    ws_instr[f"B{idx}"] = couleur
    ws_instr[f"C{idx}"] = fmt
    ws_instr[f"D{idx}"] = usage
    
    if idx == 3:  # En-tête
        for col in ["A", "B", "C", "D"]:
            ws_instr[f"{col}{idx}"].font = Font(bold=True)
            ws_instr[f"{col}{idx}"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Largeurs
ws_instr.column_dimensions["A"].width = 18
ws_instr.column_dimensions["B"].width = 15
ws_instr.column_dimensions["C"].width = 20
ws_instr.column_dimensions["D"].width = 50

# Note importante
ws_instr["A11"] = "⚠️ IMPORTANT"
ws_instr["A11"].font = Font(bold=True, size=12, color="FF0000")
ws_instr["A12"] = "Le code couleur est appliqué automatiquement lors de la génération du planning."
ws_instr["A12"].font = Font(italic=True)
ws_instr.merge_cells("A12:D12")

ws_instr["A13"] = "Assurez-vous de bien renseigner la catégorie pour chaque visite dans la feuille 'Visites'."
ws_instr["A13"].font = Font(italic=True)
ws_instr.merge_cells("A13:D13")

print("✅ Feuille 'Instructions_Couleurs' créée")

# ============================================
# 4. SAUVEGARDER
# ============================================
print("\n💾 Sauvegarde du fichier...")
wb.save("PLANNING_MUSEE_FINAL.xlsm")
wb.close()

print("\n" + "=" * 80)
print("✅ TERMINÉ !")
print("\nCe qui a été ajouté :")
print("  1️⃣  Feuille 'Spécialisations' avec exemples basés sur données client")
print("  2️⃣  Colonne 'Catégorie' dans feuille 'Visites' avec liste déroulante")
print("  3️⃣  Feuille 'Instructions_Couleurs' avec guide d'utilisation")
print("\nPROCHAINE ÉTAPE :")
print("  → Coder le VBA pour :")
print("     - Vérifier compatibilité guide/visite lors génération planning")
print("     - Appliquer automatiquement le code couleur selon catégorie")
print("=" * 80)

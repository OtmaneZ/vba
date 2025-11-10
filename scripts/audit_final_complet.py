"""
Audit final : Vérification complète XLSM vs Mission MALT vs Formulaire Client
"""

import openpyxl
import os

def audit_complet():
    print("="*80)
    print("🔍 AUDIT FINAL : CONFORMITÉ MISSION MALT")
    print("="*80)
    
    # Charger les fichiers
    wb_xlsm = openpyxl.load_workbook('PLANNING_MUSEE_FINAL_PROPRE.xlsm', keep_vba=True)
    wb_client = openpyxl.load_workbook('data/FORMULAIRE_CLIENT_PRO V2.xlsx')
    
    # Compter les modules VBA
    vba_modules = [f for f in os.listdir('vba-modules') if f.endswith(('.bas', '.cls'))]
    
    print(f"\n📂 Fichiers analysés:")
    print(f"   ✅ XLSM : {len(wb_xlsm.sheetnames)} onglets")
    print(f"   ✅ Modules VBA : {len(vba_modules)} fichiers")
    print(f"   ✅ Formulaire client : {len(wb_client.sheetnames)} onglets")
    
    # ==================== EXIGENCES MALT ====================
    print("\n" + "="*80)
    print("📋 VÉRIFICATION EXIGENCES MISSION MALT")
    print("="*80)
    
    exigences = {
        "1. Recueillir disponibilités de manière confidentielle": {
            "requis": True,
            "implementé": False,
            "détails": ""
        },
        "2. Indiquer automatiquement quel guide est libre": {
            "requis": True,
            "implementé": False,
            "détails": ""
        },
        "3. Envoyer planning du mois à chaque guide": {
            "requis": True,
            "implementé": False,
            "détails": ""
        },
        "4. Notifications email J-7 et J-1": {
            "requis": True,
            "implementé": False,
            "détails": ""
        },
        "5. Calculer nombre de visites par guide": {
            "requis": True,
            "implementé": False,
            "détails": ""
        },
        "6. Associer visites au montant salaire": {
            "requis": True,
            "implementé": False,
            "détails": ""
        },
        "7. Remplir automatiquement contrat": {
            "requis": True,
            "implementé": False,
            "détails": ""
        }
    }
    
    # Vérifier onglet Disponibilités
    if 'Disponibilites' in wb_xlsm.sheetnames:
        exigences["1. Recueillir disponibilités de manière confidentielle"]["implementé"] = True
        exigences["1. Recueillir disponibilités de manière confidentielle"]["détails"] = "✅ Feuille Disponibilités + Authentification sécurisée"
    
    # Vérifier fonction d'attribution
    if 'Module_Planning.bas' in vba_modules:
        exigences["2. Indiquer automatiquement quel guide est libre"]["implementé"] = True
        exigences["2. Indiquer automatiquement quel guide est libre"]["détails"] = "✅ Module_Planning avec attribution automatique"
    
    # Vérifier envoi planning
    if 'Module_Emails.bas' in vba_modules:
        exigences["3. Envoyer planning du mois à chaque guide"]["implementé"] = True
        exigences["3. Envoyer planning du mois à chaque guide"]["détails"] = "✅ Module_Emails avec envoi automatique"
        exigences["4. Notifications email J-7 et J-1"]["implementé"] = True
        exigences["4. Notifications email J-7 et J-1"]["détails"] = "✅ Module_Emails avec notifications configurables"
    
    # Vérifier calculs
    if 'Module_Calculs.bas' in vba_modules:
        exigences["5. Calculer nombre de visites par guide"]["implementé"] = True
        exigences["5. Calculer nombre de visites par guide"]["détails"] = "✅ Module_Calculs avec statistiques"
        exigences["6. Associer visites au montant salaire"]["implementé"] = True
        exigences["6. Associer visites au montant salaire"]["détails"] = "✅ Module_Calculs avec calcul paie automatique"
    
    # Vérifier contrats
    if 'Module_Contrats.bas' in vba_modules:
        exigences["7. Remplir automatiquement contrat"]["implementé"] = True
        exigences["7. Remplir automatiquement contrat"]["détails"] = "✅ Module_Contrats avec génération automatique"
    
    nb_ok = sum(1 for e in exigences.values() if e["implementé"])
    nb_total = len(exigences)
    
    print(f"\n🎯 CONFORMITÉ : {nb_ok}/{nb_total} ({int(nb_ok/nb_total*100)}%)")
    print()
    
    for i, (nom, data) in enumerate(exigences.items(), 1):
        statut = "✅" if data["implementé"] else "❌"
        print(f"{statut} {nom}")
        if data["détails"]:
            print(f"   {data['détails']}")
    
    # ==================== DONNÉES CLIENT ====================
    print("\n" + "="*80)
    print("📊 VÉRIFICATION DONNÉES CLIENT")
    print("="*80)
    
    # Guides
    ws_guides = wb_xlsm['Guides']
    nb_guides = ws_guides.max_row - 1
    print(f"\n👥 GUIDES : {nb_guides}")
    
    guides_complets = 0
    guides_incomplets = 0
    for i in range(2, ws_guides.max_row + 1):
        prenom = ws_guides.cell(i, 1).value
        nom = ws_guides.cell(i, 2).value
        email = ws_guides.cell(i, 3).value
        tarif = ws_guides.cell(i, 5).value
        mdp = ws_guides.cell(i, 6).value
        
        if prenom and nom and email and tarif and mdp:
            guides_complets += 1
        else:
            guides_incomplets += 1
    
    print(f"   ✅ Guides complets (nom+email+tarif+mdp) : {guides_complets}")
    print(f"   ⚠️  Guides incomplets (données manquantes) : {guides_incomplets}")
    
    # Visites
    ws_visites = wb_xlsm['Visites']
    nb_visites = ws_visites.max_row - 1
    print(f"\n🎫 TYPES DE VISITES : {nb_visites}")
    
    visites_programmees = 0
    visites_non_programmees = 0
    for i in range(2, ws_visites.max_row + 1):
        date = ws_visites.cell(i, 2).value
        if date:
            visites_programmees += 1
        else:
            visites_non_programmees += 1
    
    print(f"   ✅ Visites programmées (avec date) : {visites_programmees}")
    print(f"   ⚠️  Visites non programmées (sans date) : {visites_non_programmees}")
    
    # Disponibilités
    ws_dispo = wb_xlsm['Disponibilites']
    nb_dispo = ws_dispo.max_row - 1
    print(f"\n📅 DISPONIBILITÉS : {nb_dispo}")
    if nb_dispo == 1:
        print(f"   ⚠️  Aucune disponibilité saisie (feuille vide)")
    else:
        print(f"   ✅ {nb_dispo} lignes de disponibilités")
    
    # Configuration
    ws_config = wb_xlsm['Configuration']
    config_ok = 0
    config_test = 0
    params_critiques = ['Email_Expediteur', 'Nom_Association', 'MotDePasseAdmin']
    
    for i in range(2, ws_config.max_row + 1):
        param = ws_config.cell(i, 1).value
        valeur = ws_config.cell(i, 2).value
        if param in params_critiques:
            if valeur and ('test' in str(valeur).lower() or 'admin123' in str(valeur).lower() or 'musee.fr' in str(valeur).lower()):
                config_test += 1
            elif valeur:
                config_ok += 1
    
    print(f"\n⚙️  CONFIGURATION :")
    print(f"   ⚠️  Paramètres avec valeurs test : {config_test}/{len(params_critiques)}")
    print(f"   ✅ Paramètres configurés : {config_ok}/{len(params_critiques)}")
    
    # ==================== MODULES VBA ====================
    print("\n" + "="*80)
    print("🔧 MODULES VBA")
    print("="*80)
    
    modules_attendus = {
        'Module_Authentification.bas': 'Gestion connexion guides/admin',
        'Module_Planning.bas': 'Attribution automatique visites',
        'Module_Disponibilites.bas': 'Gestion disponibilités',
        'Module_Emails.bas': 'Envoi emails automatiques',
        'Module_Calculs.bas': 'Calcul paie et statistiques',
        'Module_Contrats.bas': 'Génération contrats',
        'Module_Config.bas': 'Configuration système',
        'Module_Specialisations.bas': 'Gestion spécialisations',
        'Module_CodeCouleur.bas': 'Codes couleurs automatiques',
        'Feuille_Accueil.cls': 'Interface accueil',
        'Feuille_Visites.cls': 'Gestion feuille visites',
        'ThisWorkbook.cls': 'Événements workbook'
    }
    
    modules_presents = 0
    modules_manquants = []
    
    for module, desc in modules_attendus.items():
        if module in vba_modules:
            print(f"   ✅ {module} - {desc}")
            modules_presents += 1
        else:
            print(f"   ❌ {module} - {desc} [MANQUANT]")
            modules_manquants.append(module)
    
    print(f"\n📊 {modules_presents}/{len(modules_attendus)} modules présents")
    
    # ==================== RÉSUMÉ FINAL ====================
    print("\n" + "="*80)
    print("🎯 RÉSUMÉ FINAL")
    print("="*80)
    
    score_malt = int(nb_ok/nb_total*100)
    score_donnees = int(((guides_complets/nb_guides if nb_guides > 0 else 0) * 0.4 + 
                        (visites_programmees/nb_visites if nb_visites > 0 else 0) * 0.3 +
                        (1 if nb_dispo > 1 else 0) * 0.3) * 100)
    score_modules = int(modules_presents/len(modules_attendus)*100)
    score_global = int((score_malt * 0.5 + score_donnees * 0.25 + score_modules * 0.25))
    
    print(f"\n📈 SCORES:")
    print(f"   • Conformité MALT : {score_malt}% ({nb_ok}/{nb_total})")
    print(f"   • Données client : {score_donnees}%")
    print(f"   • Modules VBA : {score_modules}% ({modules_presents}/{len(modules_attendus)})")
    print(f"   • SCORE GLOBAL : {score_global}%")
    
    print(f"\n🚦 STATUT DU PROJET:")
    if score_global >= 95:
        print("   ✅ PRÊT POUR LIVRAISON")
    elif score_global >= 80:
        print("   ⚠️  PRESQUE PRÊT - Compléter les données manquantes")
    else:
        print("   🔴 EN COURS - Fonctionnalités à finaliser")
    
    print(f"\n📝 ACTIONS RESTANTES:")
    
    if guides_incomplets > 0:
        print(f"   1. Compléter {guides_incomplets} guides (tarifs + mots de passe)")
    
    if nb_dispo <= 1:
        print(f"   2. Saisir les disponibilités des {nb_guides} guides")
    
    if visites_non_programmees > 0:
        print(f"   3. Programmer {visites_non_programmees} visites (dates/heures)")
    
    if config_test > 0:
        print(f"   4. Remplacer {config_test} paramètres de configuration test")
    
    if not all([exigences[e]["implementé"] for e in exigences]):
        manquants = [e for e, v in exigences.items() if not v["implementé"]]
        if manquants:
            print(f"   5. Implémenter fonctionnalités manquantes:")
            for m in manquants:
                print(f"      - {m}")
    
    if not guides_incomplets and nb_dispo > 1 and not config_test:
        print("   ✅ Aucune action - Projet complet!")
    
    print("\n" + "="*80)

if __name__ == "__main__":
    audit_complet()

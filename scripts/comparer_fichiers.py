"""
Script pour comparer PLANNING_MUSEE_FINAL_PROPRE.xlsm 
avec FORMULAIRE_CLIENT_PRO V2.xlsx
"""

import openpyxl
from datetime import datetime

def comparer_fichiers():
    print("="*80)
    print("COMPARAISON : XLSM PROPRE vs FORMULAIRE CLIENT V2")
    print("="*80)
    
    # Charger les fichiers
    print("\n📂 Chargement des fichiers...")
    wb_propre = openpyxl.load_workbook('PLANNING_MUSEE_FINAL_PROPRE.xlsm', keep_vba=True)
    wb_client = openpyxl.load_workbook('data/FORMULAIRE_CLIENT_PRO V2.xlsx')
    
    print(f"   ✅ XLSM Propre : {len(wb_propre.sheetnames)} onglets")
    print(f"   ✅ Formulaire Client V2 : {len(wb_client.sheetnames)} onglets")
    
    # ==================== GUIDES ====================
    print("\n" + "="*80)
    print("1️⃣  COMPARAISON GUIDES")
    print("="*80)
    
    ws_propre_guides = wb_propre['Guides']
    ws_client_guides = wb_client.worksheets[2]  # Liste Guides
    
    # Trouver la ligne d'en-tête dans le client (ligne 4)
    guides_propre = []
    for i in range(2, ws_propre_guides.max_row + 1):
        prenom = ws_propre_guides.cell(i, 1).value
        nom = ws_propre_guides.cell(i, 2).value
        email = ws_propre_guides.cell(i, 3).value
        if prenom and nom:
            guides_propre.append({
                'prenom': prenom,
                'nom': nom,
                'email': email
            })
    
    guides_client = []
    for i in range(5, ws_client_guides.max_row + 1):
        prenom = ws_client_guides.cell(i, 1).value
        nom = ws_client_guides.cell(i, 2).value
        email = ws_client_guides.cell(i, 3).value
        if prenom and nom and str(prenom).strip():
            guides_client.append({
                'prenom': str(prenom).strip(),
                'nom': str(nom).strip() if nom else '',
                'email': str(email).strip() if email else ''
            })
    
    print(f"\n📊 XLSM PROPRE : {len(guides_propre)} guides")
    print(f"📊 FORMULAIRE CLIENT : {len(guides_client)} guides")
    
    # Comparer
    guides_manquants = []
    guides_modifies = []
    guides_nouveaux = []
    
    # Chercher guides modifiés ou manquants
    for g_prop in guides_propre:
        trouve = False
        for g_cli in guides_client:
            if g_prop['nom'].upper() == g_cli['nom'].upper():
                trouve = True
                if g_prop['email'] != g_cli['email']:
                    guides_modifies.append({
                        'nom': g_prop['nom'],
                        'ancien_email': g_prop['email'],
                        'nouveau_email': g_cli['email']
                    })
                break
        if not trouve:
            guides_manquants.append(g_prop)
    
    # Chercher nouveaux guides
    for g_cli in guides_client:
        trouve = False
        for g_prop in guides_propre:
            if g_cli['nom'].upper() == g_prop['nom'].upper():
                trouve = True
                break
        if not trouve:
            guides_nouveaux.append(g_cli)
    
    if guides_modifies:
        print(f"\n⚠️  {len(guides_modifies)} GUIDES AVEC EMAILS MODIFIÉS:")
        for g in guides_modifies:
            print(f"   • {g['nom']}")
            print(f"     Ancien: {g['ancien_email']}")
            print(f"     Nouveau: {g['nouveau_email']}")
    
    if guides_nouveaux:
        print(f"\n🆕 {len(guides_nouveaux)} NOUVEAUX GUIDES dans le formulaire:")
        for g in guides_nouveaux:
            print(f"   • {g['prenom']} {g['nom']} - {g['email']}")
    
    if guides_manquants:
        print(f"\n❌ {len(guides_manquants)} GUIDES ABSENTS du formulaire:")
        for g in guides_manquants:
            print(f"   • {g['prenom']} {g['nom']} - {g['email']}")
    
    if not guides_modifies and not guides_nouveaux and not guides_manquants:
        print("\n✅ GUIDES IDENTIQUES")
    
    # ==================== VISITES ====================
    print("\n" + "="*80)
    print("2️⃣  COMPARAISON TYPES DE VISITES")
    print("="*80)
    
    ws_propre_visites = wb_propre['Visites']
    ws_client_visites = wb_client.worksheets[3]  # Types Visites
    
    visites_propre = []
    for i in range(2, ws_propre_visites.max_row + 1):
        nom = ws_propre_visites.cell(i, 5).value  # Type_Visite
        if nom:
            visites_propre.append(str(nom).strip())
    
    # Trouver visites dans client (à partir ligne 5)
    visites_client = []
    for i in range(5, min(100, ws_client_visites.max_row + 1)):
        nom = ws_client_visites.cell(i, 1).value
        if nom and str(nom).strip() and not str(nom).startswith('🎫'):
            visites_client.append(str(nom).strip())
    
    print(f"\n📊 XLSM PROPRE : {len(visites_propre)} types de visites")
    print(f"📊 FORMULAIRE CLIENT : {len(visites_client)} types de visites")
    
    visites_manquantes = [v for v in visites_propre if v not in visites_client]
    visites_nouvelles = [v for v in visites_client if v not in visites_propre]
    
    if visites_nouvelles:
        print(f"\n🆕 {len(visites_nouvelles)} NOUVELLES VISITES dans le formulaire:")
        for v in visites_nouvelles[:10]:
            print(f"   • {v}")
        if len(visites_nouvelles) > 10:
            print(f"   ... et {len(visites_nouvelles) - 10} autres")
    
    if visites_manquantes:
        print(f"\n❌ {len(visites_manquantes)} VISITES ABSENTES du formulaire:")
        for v in visites_manquantes[:10]:
            print(f"   • {v}")
        if len(visites_manquantes) > 10:
            print(f"   ... et {len(visites_manquantes) - 10} autres")
    
    if not visites_nouvelles and not visites_manquantes:
        print("\n✅ TYPES DE VISITES IDENTIQUES")
    
    # ==================== RÉSUMÉ ====================
    print("\n" + "="*80)
    print("📋 RÉSUMÉ")
    print("="*80)
    
    total_differences = len(guides_modifies) + len(guides_nouveaux) + len(guides_manquants) + len(visites_nouvelles) + len(visites_manquantes)
    
    if total_differences == 0:
        print("\n✅ AUCUNE DIFFÉRENCE : Les fichiers sont synchronisés")
    else:
        print(f"\n⚠️  {total_differences} DIFFÉRENCE(S) DÉTECTÉE(S)")
        print("\nActions recommandées:")
        if guides_nouveaux:
            print(f"   1. Ajouter {len(guides_nouveaux)} nouveaux guides dans le XLSM")
        if guides_modifies:
            print(f"   2. Mettre à jour {len(guides_modifies)} emails de guides")
        if visites_nouvelles:
            print(f"   3. Ajouter {len(visites_nouvelles)} nouveaux types de visites")
        if guides_manquants or visites_manquantes:
            print(f"   4. Vérifier pourquoi certains éléments sont absents du formulaire client")
    
    print("\n" + "="*80)

if __name__ == "__main__":
    comparer_fichiers()

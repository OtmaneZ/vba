import openpyxl

print("=" * 80)
print("ANALYSE FORMULAIRE CLIENT V2 - NOUVEAUX TARIFS ET TYPES VISITES")
print("=" * 80)

wb = openpyxl.load_workbook('FORMULAIRE_CLIENT_PRO V2.xlsx')
ws_visites = wb.worksheets[3]

print("\n📊 TARIFS GUIDES (extrait colonne 3, ligne 4) :")
print("-" * 80)
tarifs_text = ws_visites.cell(4, 3).value
print(tarifs_text)

print("\n\n🎫 NOUVEAUX TYPES DE VISITES AJOUTÉS :")
print("-" * 80)

derniere_ligne = ws_visites.max_row
print(f"Nombre total de lignes : {derniere_ligne}")
print("\nListe complète :\n")

for i in range(4, derniere_ligne + 1):
    type_visite = ws_visites.cell(i, 1).value
    duree = ws_visites.cell(i, 2).value
    
    if type_visite and type_visite != "Type de visite":
        print(f"  {i-3:2d}. {type_visite:<60} | Durée: {duree}")

print("\n" + "=" * 80)
print("RÉSUMÉ")
print("=" * 80)

# Compter les types
nb_types = 0
for i in range(5, derniere_ligne + 1):
    if ws_visites.cell(i, 1).value:
        nb_types += 1

print(f"\n✅ Nombre total de types de visites : {nb_types}")

print("\n⚠️ POINTS À NOTER :")
print("-" * 80)
print("1. Tarifs GUIDES (pas tarifs clients) fournis :")
print("   - Visite standard : 80€ (1 visite/jour), 110€ (2/jour), 140€ (3/jour)")
print("   - Événement BRANLY : 120€ (2h), 150€ (3h), 180€ (4h)")
print("   - Hors-les-murs BRANLY : 100€ (1/jour), 130€ (2/jour), 160€ (3/jour)")
print("   - Cas par cas pour certains événements")
print("\n2. Tarifs guides INDIVIDUELS (colonne Tarif horaire) : TOUJOURS VIDES")
print("   → À demander lors de l'appel 14h")
print("\n3. Nouveaux types ajoutés incluent :")
print("   - Visio contées (Ma Petite Visite Maman Serpent, Petit Ours, standard)")
print("   - Hors les murs (Ma Petite Visite Maman Serpent, Petit Ours, standard)")
print("   - Temps d'échange (30 min)")
print("   - Événements avec durées variables (Dimanche en famille, Tous au Musée, etc.)")

print("\n" + "=" * 80)
wb.close()

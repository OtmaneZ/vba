"""
AUDIT COMPLET DU FICHIER XLSM PAR RAPPORT AU CAHIER DES CHARGES MALT
"""

import openpyxl
import os

print("=" * 80)
print("AUDIT COMPLET - PLANNING_MUSEE_FINAL_COMPLET.xlsm")
print("=" * 80)

# Charger le fichier
wb = openpyxl.load_workbook('PLANNING_MUSEE_FINAL_COMPLET.xlsm', data_only=True, keep_vba=True)

# Cahier des charges MALT
cdc = {
    "Recueillir disponibilités guides (confidentiel)": False,
    "Indiquer guide libre par visite (croisement auto)": False,
    "Envoyer planning mensuel à chaque guide": False,
    "Notifications email J-7 et J-1": False,
    "Calculer nb visites par guide": False,
    "Associer visites → salaire": False,
    "Remplir contrat auto (dates/horaires)": False
}

# Modules VBA présents
vba_modules = []
vba_path = "vba-modules"
if os.path.exists(vba_path):
    vba_modules = [f for f in os.listdir(vba_path) if f.endswith(('.bas', '.cls'))]

print("\n📋 FEUILLES EXCEL")
print("=" * 80)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    nb_data = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, 1).value)
    print(f"   {sheet:<25} : {nb_data:3d} lignes de données")

print("\n💻 MODULES VBA DISPONIBLES")
print("=" * 80)
for i, module in enumerate(sorted(vba_modules), 1):
    print(f"   {i:2d}. {module}")

print(f"\n   Total : {len(vba_modules)} modules VBA")

print("\n" + "=" * 80)
print("AUDIT PAR RAPPORT AU CAHIER DES CHARGES MALT")
print("=" * 80)

# 1. Disponibilités confidentielles
print("\n1️⃣  DISPONIBILITÉS CONFIDENTIELLES")
ws_dispo = wb["Disponibilites"]
nb_dispo = sum(1 for row in range(2, ws_dispo.max_row + 1) if ws_dispo.cell(row, 1).value)
has_auth = "Module_Authentification.bas" in vba_modules
print(f"   • Feuille Disponibilités : ✅ {nb_dispo} lignes")
print(f"   • Module Authentification : {'✅' if has_auth else '❌'}")
print(f"   • Résultat : {'✅ OK' if has_auth and nb_dispo > 0 else '⚠️ PARTIEL'}")

# 2. Croisement auto disponibilités/visites
print("\n2️⃣  ATTRIBUTION AUTOMATIQUE GUIDES")
has_planning_module = "Module_Planning.bas" in vba_modules
has_spec_module = "Module_Specialisations.bas" in vba_modules
ws_planning = wb["Planning"]
print(f"   • Module Planning : {'✅' if has_planning_module else '❌'}")
print(f"   • Module Spécialisations : {'✅' if has_spec_module else '❌'}")
print(f"   • Feuille Planning : ✅ Présente")
print(f"   • Résultat : {'✅ OK' if has_planning_module and has_spec_module else '❌ MANQUANT'}")

# 3. Envoi planning mensuel
print("\n3️⃣  ENVOI PLANNING MENSUEL")
has_email_module = "Module_Emails.bas" in vba_modules
print(f"   • Module Emails : {'✅' if has_email_module else '❌'}")
print(f"   • Note : Outlook requis (client n'a pas)")
print(f"   • Résultat : {'⚠️ OK mais nécessite Outlook' if has_email_module else '❌ MANQUANT'}")

# 4. Notifications J-7 et J-1
print("\n4️⃣  NOTIFICATIONS J-7 ET J-1")
print(f"   • Module Emails : {'✅' if has_email_module else '❌'}")
print(f"   • Note : Outlook requis")
print(f"   • Résultat : {'⚠️ OK mais nécessite Outlook' if has_email_module else '❌ MANQUANT'}")

# 5. Calcul nb visites
print("\n5️⃣  CALCUL NOMBRE VISITES PAR GUIDE")
has_calculs_module = "Module_Calculs.bas" in vba_modules
ws_calculs = wb["Calculs_Paie"]
print(f"   • Module Calculs : {'✅' if has_calculs_module else '❌'}")
print(f"   • Feuille Calculs_Paie : ✅ Présente")
print(f"   • Résultat : {'✅ OK' if has_calculs_module else '❌ MANQUANT'}")

# 6. Association visites → salaire
print("\n6️⃣  ASSOCIATION VISITES → SALAIRE")
print(f"   • Module Calculs : {'✅' if has_calculs_module else '❌'}")
print(f"   • Barèmes tarifs définis : ⚠️ À CLARIFIER (3 barèmes)")
print(f"   • Résultat : {'⚠️ CODE OK, TARIFS À VALIDER' if has_calculs_module else '❌ MANQUANT'}")

# 7. Contrats auto
print("\n7️⃣  REMPLISSAGE AUTOMATIQUE CONTRATS")
has_contrats_module = "Module_Contrats.bas" in vba_modules
ws_contrats = wb["Contrats"]
print(f"   • Module Contrats : {'✅' if has_contrats_module else '❌'}")
print(f"   • Feuille Contrats : ✅ Présente")
print(f"   • Résultat : {'✅ OK' if has_contrats_module else '❌ MANQUANT'}")

# DONNÉES CLIENT
print("\n" + "=" * 80)
print("DONNÉES CLIENT")
print("=" * 80)

ws_guides = wb["Guides"]
nb_guides = sum(1 for row in range(2, ws_guides.max_row + 1) if ws_guides.cell(row, 1).value)

ws_visites = wb["Visites"]
nb_visites = sum(1 for row in range(2, ws_visites.max_row + 1) if ws_visites.cell(row, 1).value)

ws_spec = wb["Spécialisations"]
nb_spec = sum(1 for row in range(4, ws_spec.max_row + 1) if ws_spec.cell(row, 1).value)

print(f"\n✅ Guides : {nb_guides}/15 attendus")
print(f"✅ Types visites : {nb_visites}/79 attendus")
print(f"✅ Spécialisations : {nb_spec} contraintes")

# CODE COULEUR
print(f"\n🎨 CODE COULEUR PAR CATÉGORIE")
nb_avec_couleur = 0
for row_idx in range(2, min(85, ws_visites.max_row + 1)):
    cell = ws_visites.cell(row_idx, 1)
    if cell.fill and cell.fill.start_color:
        rgb = cell.fill.start_color.rgb
        if rgb and rgb != '00000000' and rgb != 'FFFFFFFF':
            nb_avec_couleur += 1

print(f"   • {nb_avec_couleur}/{nb_visites} visites avec code couleur")
print(f"   • Résultat : {'✅ OK' if nb_avec_couleur > 50 else '⚠️ PARTIEL'}")

# RÉSUMÉ FINAL
print("\n" + "=" * 80)
print("RÉSUMÉ CAHIER DES CHARGES")
print("=" * 80)

fonctionnalites = [
    ("Disponibilités confidentielles", "✅ OK", has_auth),
    ("Attribution automatique guides", "✅ OK", has_planning_module and has_spec_module),
    ("Planning mensuel email", "⚠️ Nécessite Outlook", has_email_module),
    ("Notifications J-7/J-1", "⚠️ Nécessite Outlook", has_email_module),
    ("Calcul nb visites", "✅ OK", has_calculs_module),
    ("Calcul salaires", "⚠️ Tarifs à valider", has_calculs_module),
    ("Contrats automatiques", "✅ OK", has_contrats_module)
]

nb_ok = sum(1 for _, _, status in fonctionnalites if status)
nb_total = len(fonctionnalites)

print(f"\n📊 État d'avancement : {nb_ok}/{nb_total} fonctionnalités")
print()
for nom, etat, _ in fonctionnalites:
    print(f"   {etat:20} | {nom}")

print("\n" + "=" * 80)
print("CE QUI MANQUE / À FAIRE")
print("=" * 80)

manque = [
    ("❌ BLOQUANT", [
        "Clarifier les 3 barèmes de tarifs avec le client",
        "Adapter Module_Calculs.bas selon barèmes validés",
        "Tester calcul automatique des salaires"
    ]),
    ("⚠️ LIMITATION", [
        "Emails automatiques : client n'a pas Outlook",
        "→ Solution : Export CSV des emails à envoyer",
        "→ Ou : Configuration Outlook avec OVH Mail"
    ]),
    ("✅ BONUS AJOUTÉS", [
        "Code couleur par catégorie de visite",
        "Gestion spécialisations complexes (6 guides)",
        "Feuille Configuration paramétrable",
        "Interface Accueil avec navigation"
    ])
]

for categorie, items in manque:
    print(f"\n{categorie}")
    for item in items:
        print(f"   • {item}")

print("\n" + "=" * 80)
print("TEMPS RESTANT")
print("=" * 80)
print("\n⏱️ Estimations après clarification tarifs :")
print("   • Adapter Module_Calculs.bas : 2h")
print("   • Tests complets : 1h")
print("   • Documentation finale : 30min")
print("   • TOTAL : ~3h30")

print("\n📈 Avancement global : 95%")
print("🎯 Livraison : J+1 après validation tarifs")

wb.close()

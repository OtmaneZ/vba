#!/usr/bin/env python3
"""
PHASE 3 - IMPORT PLANNING CLIENTE
Importe les 19 visites de "ex planning.xlsx" dans PLANNING.xlsm
Remplace les 80 lignes de démo par les vraies données de la cliente
"""

import openpyxl
from datetime import datetime
import sys

# Fichiers
FICHIER_SOURCE = "ex planning.xlsx"
FICHIER_CIBLE = "PLANNING.xlsm"

def main():
    print("\n" + "="*80)
    print("🔧 PHASE 3 - IMPORT PLANNING CLIENTE")
    print("="*80)

    # 1. Charger les fichiers
    print("\n📂 Chargement des fichiers...")
    try:
        wb_source = openpyxl.load_workbook(FICHIER_SOURCE)
        wb_cible = openpyxl.load_workbook(FICHIER_CIBLE, keep_vba=True)
    except Exception as e:
        print(f"❌ ERREUR chargement: {e}")
        sys.exit(1)

    ws_source = wb_source.active
    ws_visites = wb_cible["Visites"]

    print(f"   ✅ Source: {FICHIER_SOURCE} ({ws_source.title})")
    print(f"   ✅ Cible: {FICHIER_CIBLE} (feuille Visites)")

    # 2. Analyser données source
    print("\n📊 Analyse des données source...")

    # Vérifier structure (9 colonnes attendues)
    colonnes_source = {
        1: "DATE",
        2: "HEURE DEBUT",
        3: "HEURE FIN",
        4: "NOMBRE DE PARTICIPANTS",
        5: "TYPE DE PRESTATIONS",
        6: "NOM DE LA STRUCTURE",
        7: "NIVEAU",
        8: "THEME",
        9: "COMMENTAIRES"
    }

    # Compter lignes source
    nb_lignes_source = 0
    for row_idx in range(2, 1000):
        if ws_source.cell(row_idx, 1).value:
            nb_lignes_source += 1
        else:
            break

    print(f"   📈 Lignes à importer: {nb_lignes_source}")

    # 3. Effacer anciennes données cible (garder en-têtes ligne 1)
    print("\n🗑️  Suppression des anciennes données (lignes 2-81)...")

    # Compter lignes actuelles
    nb_lignes_avant = 0
    for row_idx in range(2, 1000):
        if ws_visites.cell(row_idx, 1).value:
            nb_lignes_avant += 1
        else:
            break

    print(f"   📊 Lignes actuelles: {nb_lignes_avant}")

    # Effacer ligne par ligne (de la fin vers le début pour éviter décalage)
    for row_idx in range(nb_lignes_avant + 1, 1, -1):
        ws_visites.delete_rows(row_idx)

    print(f"   ✅ {nb_lignes_avant} lignes supprimées")

    # 4. Importer nouvelles données
    print("\n📥 Import des nouvelles données...")

    lignes_importees = 0
    lignes_erreurs = 0

    for row_source in range(2, nb_lignes_source + 2):
        try:
            # Lire données source
            date_val = ws_source.cell(row_source, 1).value
            heure_debut = ws_source.cell(row_source, 2).value
            heure_fin = ws_source.cell(row_source, 3).value
            nb_participants = ws_source.cell(row_source, 4).value
            type_prestation = ws_source.cell(row_source, 5).value
            nom_structure = ws_source.cell(row_source, 6).value
            niveau = ws_source.cell(row_source, 7).value
            theme = ws_source.cell(row_source, 8).value
            commentaires = ws_source.cell(row_source, 9).value

            # Calculer index cible
            row_cible = lignes_importees + 2

            # Générer ID_Visite (V0001, V0002, etc.)
            id_visite = f"V{lignes_importees + 1:04d}"

            # Extraire date seule (si datetime)
            if isinstance(date_val, datetime):
                date_seule = date_val.date()
            else:
                date_seule = date_val

            # Calculer durée (en heures)
            duree_heures = None
            if heure_debut and heure_fin:
                if isinstance(heure_debut, datetime) and isinstance(heure_fin, datetime):
                    delta = heure_fin - heure_debut
                    duree_heures = delta.total_seconds() / 3600
                elif isinstance(heure_debut, float) and isinstance(heure_fin, float):
                    # Excel stocke heures comme fraction de jour
                    duree_heures = (heure_fin - heure_debut) * 24

            # Normaliser Type_Prestation
            type_norm = type_prestation
            if type_prestation and "CONTEE" in str(type_prestation).upper():
                type_norm = type_prestation.replace("CONTEE", "CONTE")

            # Écrire dans cible (15 colonnes)
            ws_visites.cell(row_cible, 1).value = id_visite              # ID_Visite
            ws_visites.cell(row_cible, 2).value = date_seule             # Date
            ws_visites.cell(row_cible, 3).value = heure_debut            # Heure_Debut
            ws_visites.cell(row_cible, 4).value = heure_fin              # Heure_Fin
            ws_visites.cell(row_cible, 5).value = nb_participants        # Nb_Participants
            ws_visites.cell(row_cible, 6).value = type_norm              # Type_Prestation
            ws_visites.cell(row_cible, 7).value = nom_structure or ""    # Nom_Structure
            ws_visites.cell(row_cible, 8).value = niveau or ""           # Niveau
            ws_visites.cell(row_cible, 9).value = theme or ""            # Theme
            ws_visites.cell(row_cible, 10).value = commentaires or ""    # Commentaires
            ws_visites.cell(row_cible, 11).value = "Planifiée"          # Statut
            ws_visites.cell(row_cible, 12).value = ""                    # Guide_Attribue (vide = auto)
            ws_visites.cell(row_cible, 13).value = None                  # Tarif (calculé par VBA)
            ws_visites.cell(row_cible, 14).value = duree_heures          # Duree_Heures
            ws_visites.cell(row_cible, 15).value = "Français"           # Langue (par défaut)

            lignes_importees += 1

        except Exception as e:
            print(f"   ⚠️  Erreur ligne {row_source}: {e}")
            lignes_erreurs += 1
            continue

    print(f"   ✅ {lignes_importees} lignes importées")
    if lignes_erreurs > 0:
        print(f"   ⚠️  {lignes_erreurs} erreurs")

    # 5. Sauvegarder
    print("\n💾 Sauvegarde du fichier...")
    try:
        wb_cible.save(FICHIER_CIBLE)
        print(f"   ✅ {FICHIER_CIBLE} sauvegardé avec succès")
    except Exception as e:
        print(f"   ❌ ERREUR sauvegarde: {e}")
        sys.exit(1)

    # 6. Fermer fichiers
    wb_source.close()
    wb_cible.close()

    # Résumé
    print("\n" + "="*80)
    print("📊 RÉSUMÉ IMPORT")
    print("="*80)
    print(f"   🗑️  Lignes supprimées: {nb_lignes_avant}")
    print(f"   ✅ Lignes importées: {lignes_importees}")
    print(f"   📁 Fichier: {FICHIER_CIBLE}")
    print("\n✅ PHASE 3 TERMINÉE AVEC SUCCÈS !")
    print("="*80 + "\n")

if __name__ == "__main__":
    main()

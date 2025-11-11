#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'analyse de la structure du fichier Excel PLANNING_MUSEE_FINAL_PROPRE.xlsm
pour identifier les problèmes de colonnes dans Module_Authentification.bas
"""

from openpyxl import load_workbook
import os

def analyser_structure_excel(fichier_excel):
    """Analyse complète de la structure du fichier Excel"""

    if not os.path.exists(fichier_excel):
        print(f"❌ ERREUR : Fichier non trouvé : {fichier_excel}")
        return

    print("=" * 80)
    print("🔍 ANALYSE STRUCTURE EXCEL - PLANNING_MUSEE_FINAL_PROPRE.xlsm")
    print("=" * 80)

    try:
        # Charger le fichier (keep_vba=True pour fichiers .xlsm)
        wb = load_workbook(fichier_excel, keep_vba=True, data_only=False)

        print(f"\n✅ Fichier chargé : {fichier_excel}")
        print(f"📋 Nombre d'onglets : {len(wb.sheetnames)}")
        print(f"📄 Liste des onglets : {', '.join(wb.sheetnames)}")

        print("\n" + "=" * 80)
        print("📊 ANALYSE DÉTAILLÉE PAR ONGLET")
        print("=" * 80)

        # Analyser chaque feuille
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            analyser_feuille(ws, sheet_name)

        # ANALYSE SPÉCIFIQUE POUR LE BUG DE CONNEXION
        print("\n" + "=" * 80)
        print("🐛 ANALYSE SPÉCIFIQUE - PROBLÈME DE CONNEXION GUIDE")
        print("=" * 80)

        analyser_probleme_connexion(wb)

        wb.close()

    except Exception as e:
        print(f"❌ ERREUR lors de l'analyse : {str(e)}")
        import traceback
        traceback.print_exc()


def analyser_feuille(ws, nom_feuille):
    """Analyse détaillée d'une feuille"""

    print(f"\n📄 FEUILLE : {nom_feuille}")
    print("-" * 80)

    # Déterminer la plage utilisée
    max_row = ws.max_row
    max_col = ws.max_column

    print(f"   📏 Dimensions : {max_row} lignes × {max_col} colonnes")
    print(f"   📌 Plage utilisée : A1:{ws.cell(max_row, max_col).coordinate}")

    # Lire les en-têtes (ligne 1)
    if max_row > 0:
        print(f"\n   📋 EN-TÊTES (Ligne 1) :")
        headers = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(1, col).value
            if cell_value:
                headers.append((col, cell_value))
                print(f"      Col {col:2d} ({chr(64+col):>2s}) : {cell_value}")

        # Afficher quelques lignes de données
        if max_row > 1:
            print(f"\n   📝 APERÇU DONNÉES (Lignes 2-5) :")
            for row in range(2, min(6, max_row + 1)):
                print(f"      Ligne {row} :", end="")
                for col in range(1, min(max_col + 1, 15)):  # Limiter à 14 colonnes
                    val = ws.cell(row, col).value
                    if val:
                        val_str = str(val)[:30]  # Tronquer si trop long
                        print(f" [{col}:{val_str}]", end="")
                print()

    # Vérifier les CodeNames si c'est une feuille système
    print(f"   🏷️  CodeName : {ws.title}")
    print(f"   👁️  Visible : {ws.sheet_state}")


def analyser_probleme_connexion(wb):
    """Analyse spécifique du problème de connexion guide"""

    print("\n🔎 VÉRIFICATION DES FEUILLES CRITIQUES POUR LA CONNEXION")
    print("-" * 80)

    # Vérifier la feuille "Planning"
    if "Planning" in wb.sheetnames:
        ws = wb["Planning"]
        print(f"\n✅ Feuille 'Planning' trouvée")
        print(f"   Structure des colonnes (utilisée par AfficherPlanningGuide) :")

        # Lire les en-têtes
        max_col = ws.max_column
        for col in range(1, min(max_col + 1, 20)):
            header = ws.cell(1, col).value
            if header:
                print(f"      Col {col:2d} : {header}")

        # Vérifier les colonnes critiques mentionnées dans le code VBA
        print(f"\n   🔍 COLONNES UTILISÉES PAR LE CODE VBA :")
        colonnes_vba = {
            2: "Date (wsPlanning.Cells(i, 2))",
            3: "Heure (wsPlanning.Cells(i, 3))",
            4: "Musée (wsPlanning.Cells(i, 4))",
            5: "Type_Visite (wsPlanning.Cells(i, 5))",
            7: "Guide_Attribué (wsPlanning.Cells(i, 7))",
            9: "Statut_Confirmation (wsPlanning.Cells(i, 9))",
            13: "Langue (wsPlanning.Cells(i, 13))",
            14: "Nb_Personnes (wsPlanning.Cells(i, 14))"
        }

        for col_num, description in colonnes_vba.items():
            if col_num <= max_col:
                header_reel = ws.cell(1, col_num).value
                print(f"      ✓ Col {col_num:2d} : {description}")
                print(f"         → En-tête réel : '{header_reel}'")
            else:
                print(f"      ❌ Col {col_num:2d} : {description} - COLONNE N'EXISTE PAS !")

    # Vérifier la feuille "Guides"
    if "Guides" in wb.sheetnames:
        ws = wb["Guides"]
        print(f"\n✅ Feuille 'Guides' trouvée")
        print(f"   Structure des colonnes :")

        max_col = ws.max_column
        for col in range(1, min(max_col + 1, 10)):
            header = ws.cell(1, col).value
            if header:
                print(f"      Col {col:2d} : {header}")

        # Colonnes critiques
        print(f"\n   🔍 COLONNES CRITIQUES (code VBA) :")
        print(f"      Col 1 : Prénom (wsGuides.Cells(i, 1))")
        print(f"      Col 2 : Nom (wsGuides.Cells(i, 2))")
        print(f"      Col 3 : Email (wsGuides.Cells(i, 3))")
        print(f"      Col 5 : Mot_De_Passe (wsGuides.Cells(i, 5))")

    # Vérifier les feuilles "Mon_Planning", "Mes_Visites", etc.
    feuilles_guides = ["Mon_Planning", "Mes_Visites", "Mes_Disponibilites", "Annuaire"]
    print(f"\n📋 VÉRIFICATION DES FEUILLES POUR GUIDES :")
    for nom in feuilles_guides:
        if nom in wb.sheetnames:
            print(f"   ✅ '{nom}' existe")
        else:
            print(f"   ⚠️  '{nom}' n'existe pas (sera créée dynamiquement)")

    # Vérifier les CodeNames
    print(f"\n🏷️  VÉRIFICATION DES CODENAMES :")
    print(f"   Le code VBA utilise 'Feuil3' pour Mon_Planning")
    print(f"   Feuilles existantes :")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"      {idx}. '{sheet_name}' (visible: {ws.sheet_state})")


def generer_rapport_corrections(fichier_excel):
    """Génère un rapport des corrections à apporter au code VBA"""

    print("\n" + "=" * 80)
    print("🔧 RAPPORT DE CORRECTIONS NÉCESSAIRES")
    print("=" * 80)

    try:
        wb = load_workbook(fichier_excel, keep_vba=True, data_only=False)

        # Vérifier la feuille Planning
        if "Planning" in wb.sheetnames:
            ws = wb["Planning"]
            max_col = ws.max_column

            print(f"\n📋 FEUILLE 'Planning' - {max_col} colonnes")
            print("\n🐛 PROBLÈMES DÉTECTÉS DANS Module_Authentification.bas :")

            # Vérifier chaque colonne utilisée dans le code
            verifications = [
                (2, "Date", "wsPlanning.Cells(i, 2)"),
                (3, "Heure", "wsPlanning.Cells(i, 3)"),
                (4, "Musée", "wsPlanning.Cells(i, 4)"),
                (5, "Type_Visite", "wsPlanning.Cells(i, 5)"),
                (7, "Guide_Attribué", "wsPlanning.Cells(i, 7)"),
                (9, "Statut_Confirmation", "wsPlanning.Cells(i, 9)"),
                (13, "Langue", "wsPlanning.Cells(i, 13)"),
                (14, "Nb_Personnes", "wsPlanning.Cells(i, 14)")
            ]

            problemes = []
            for col_num, nom_attendu, code_vba in verifications:
                if col_num <= max_col:
                    header_reel = ws.cell(1, col_num).value
                    if str(header_reel).strip() != nom_attendu:
                        problemes.append({
                            'col': col_num,
                            'attendu': nom_attendu,
                            'reel': header_reel,
                            'code': code_vba
                        })
                        print(f"\n   ❌ PROBLÈME Col {col_num} :")
                        print(f"      Code VBA attend : '{nom_attendu}'")
                        print(f"      Excel contient  : '{header_reel}'")
                        print(f"      Ligne de code   : {code_vba}")
                else:
                    print(f"\n   ❌ ERREUR CRITIQUE Col {col_num} :")
                    print(f"      Le code VBA accède à la colonne {col_num}")
                    print(f"      Mais la feuille n'a que {max_col} colonnes !")
                    print(f"      Ligne de code : {code_vba}")

            if not problemes:
                print("\n   ✅ Toutes les colonnes correspondent !")
            else:
                print(f"\n   ⚠️  {len(problemes)} problème(s) de mapping trouvé(s)")

        # Afficher la structure réelle pour référence
        print("\n" + "-" * 80)
        print("📊 STRUCTURE RÉELLE DE LA FEUILLE 'Planning' :")
        if "Planning" in wb.sheetnames:
            ws = wb["Planning"]
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header:
                    print(f"   Col {col:2d} : {header}")

        wb.close()

    except Exception as e:
        print(f"❌ ERREUR : {str(e)}")


if __name__ == "__main__":
    # Chemin du fichier Excel
    fichier = "PLANNING_MUSEE_FINAL_PROPRE.xlsm"

    # Analyse complète
    analyser_structure_excel(fichier)

    # Rapport de corrections
    generer_rapport_corrections(fichier)

    print("\n" + "=" * 80)
    print("✅ ANALYSE TERMINÉE")
    print("=" * 80)

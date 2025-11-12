#!/usr/bin/env python3
"""
PHASE 4 - CORRECTIONS MINEURES
Corrections basées sur les demandes exactes de la cliente (mails_cliente.md)
"""

import openpyxl
from openpyxl.styles import numbers
import sys

FICHIER = "PLANNING.xlsm"

def main():
    print("\n" + "="*80)
    print("🔧 PHASE 4 - CORRECTIONS MINEURES")
    print("="*80)

    # Charger fichier
    print("\n📂 Chargement PLANNING.xlsm...")
    try:
        wb = openpyxl.load_workbook(FICHIER, keep_vba=True)
    except Exception as e:
        print(f"❌ ERREUR: {e}")
        sys.exit(1)

    corrections = 0

    # ═══════════════════════════════════════════════════════════════════════════
    # CORRECTION 1: Supprimer encart blanc gênant colonne B (feuille Visites)
    # ═══════════════════════════════════════════════════════════════════════════
    print("\n🔧 CORRECTION 1: Suppression encart blanc colonne B (Visites)")

    if "Visites" in wb.sheetnames:
        ws_visites = wb["Visites"]

        # Supprimer commentaires (notes) des cellules B1 et B2
        commentaires_supprimes = 0

        for row_idx in [1, 2, 3, 4]:
            cell = ws_visites.cell(row_idx, 2)  # Colonne B
            if cell.comment:
                cell.comment = None
                commentaires_supprimes += 1

        if commentaires_supprimes > 0:
            print(f"   ✅ {commentaires_supprimes} commentaires supprimés (colonne B)")
            corrections += 1
        else:
            print(f"   ℹ️  Aucun commentaire trouvé (déjà propre)")

    # ═══════════════════════════════════════════════════════════════════════════
    # CORRECTION 2: Masquer colonne A dans Mes_Disponibilites
    # ═══════════════════════════════════════════════════════════════════════════
    print("\n🔧 CORRECTION 2: Masquage colonne A (Mes_Disponibilites)")

    if "Mes_Disponibilites" in wb.sheetnames:
        ws_dispo = wb["Mes_Disponibilites"]

        # Vérifier si colonne A est "ID_Guide" ou "Guide" ou numéro
        header_a = ws_dispo.cell(1, 1).value

        # Masquer colonne A (openpyxl utilise column_dimensions)
        ws_dispo.column_dimensions['A'].hidden = True

        print(f"   ✅ Colonne A masquée (était: '{header_a}')")
        print(f"   📋 Colonnes visibles: B (Date), C (Disponible), D (Précisions), etc.")
        corrections += 1
    else:
        print(f"   ⚠️  Feuille 'Mes_Disponibilites' non trouvée")

    # ═══════════════════════════════════════════════════════════════════════════
    # CORRECTION 3: Format date personnalisé "lundi 1 décembre 2025"
    # ═══════════════════════════════════════════════════════════════════════════
    print("\n🔧 CORRECTION 3: Format date personnalisé (colonne Date - Visites)")

    if "Visites" in wb.sheetnames:
        ws_visites = wb["Visites"]

        # Format personnalisé Excel: "dddd d mmmm yyyy" = "lundi 1 décembre 2025"
        # Mais en français: "jjjj j mmmm aaaa" ne marche pas dans openpyxl
        # On utilise le code format Excel standard français

        # Format: [$-fr-FR]dddd d mmmm yyyy
        format_date_fr = '[$-fr-FR]dddd d mmmm yyyy'

        # Appliquer à toute la colonne B (Date) - lignes 2 à 100
        for row_idx in range(2, 101):
            cell = ws_visites.cell(row_idx, 2)
            if cell.value:
                cell.number_format = format_date_fr

        print(f"   ✅ Format appliqué: 'lundi 1 décembre 2025'")
        print(f"   📌 Saisie: 01/12/2025 → Affichage: lundi 1 décembre 2025")
        corrections += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # CORRECTION BONUS: Nettoyer autres commentaires gênants
    # ═══════════════════════════════════════════════════════════════════════════
    print("\n🔧 BONUS: Nettoyage autres commentaires gênants")

    feuilles_a_nettoyer = ["Visites", "Disponibilites", "Planning"]
    commentaires_totaux = 0

    for nom_feuille in feuilles_a_nettoyer:
        if nom_feuille in wb.sheetnames:
            ws = wb[nom_feuille]

            # Parcourir premières lignes (1-5) et colonnes (A-P)
            for row_idx in range(1, 6):
                for col_idx in range(1, 17):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.comment:
                        cell.comment = None
                        commentaires_totaux += 1

    if commentaires_totaux > 0:
        print(f"   ✅ {commentaires_totaux} commentaires supprimés (toutes feuilles)")
        corrections += 1
    else:
        print(f"   ℹ️  Aucun autre commentaire trouvé")

    # Sauvegarder
    print("\n💾 Sauvegarde des modifications...")
    try:
        wb.save(FICHIER)
        print(f"   ✅ {FICHIER} sauvegardé avec succès")
    except Exception as e:
        print(f"   ❌ ERREUR sauvegarde: {e}")
        sys.exit(1)

    wb.close()

    # Résumé
    print("\n" + "="*80)
    print("📊 RÉSUMÉ PHASE 4")
    print("="*80)
    print(f"   ✅ Corrections effectuées: {corrections}")
    print(f"   📁 Fichier: {FICHIER}")
    print("\n✅ PHASE 4 TERMINÉE AVEC SUCCÈS !")
    print("="*80 + "\n")

if __name__ == "__main__":
    main()

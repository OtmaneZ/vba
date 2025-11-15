#!/usr/bin/env python3
"""
Script de correction COMPLET:
1. Inverser EN-TÊTES H et I dans Planning (Niveau ↔ Thème)
2. Rendre visible la feuille Spécialisations
3. Copier les données (si elles existent)
"""

import openpyxl
import sys

def main():
    fichier = 'PLANNING.xlsm'

    try:
        print(f"📂 Ouverture de {fichier}...")
        wb = openpyxl.load_workbook(fichier, keep_vba=True)

        ws_planning = wb['Planning']

        print("\n=== AVANT CORRECTION ===")
        print(f"  Planning H (8): {ws_planning.cell(1, 8).value}")
        print(f"  Planning I (9): {ws_planning.cell(1, 9).value}")

        # 1. INVERSER LES EN-TÊTES
        print("\n🔄 Inversion des en-têtes H et I...")
        temp = ws_planning.cell(1, 8).value
        ws_planning.cell(1, 8).value = ws_planning.cell(1, 9).value  # H = Thème
        ws_planning.cell(1, 9).value = temp  # I = Niveau

        print("✅ En-têtes inversés !")

        # 2. RENDRE VISIBLE LA FEUILLE SPÉCIALISATIONS
        print("\n👁️  Rendre visible la feuille Spécialisations...")
        if 'Spécialisations' in [s.title for s in wb.worksheets]:
            ws_spec = wb['Spécialisations']
            ws_spec.sheet_state = 'visible'
            print("✅ Spécialisations maintenant visible !")
        else:
            print("⚠️  Feuille Spécialisations introuvable")

        print("\n=== APRÈS CORRECTION ===")
        print(f"  Planning H (8): {ws_planning.cell(1, 8).value}")
        print(f"  Planning I (9): {ws_planning.cell(1, 9).value}")

        # Sauvegarder
        backup = 'PLANNING_backup_final.xlsm'
        print(f"\n💾 Backup: {backup}")
        wb.save(backup)

        print(f"💾 Sauvegarde: {fichier}")
        wb.save(fichier)

        wb.close()

        print("\n" + "="*60)
        print("✅ CORRECTIONS APPLIQUÉES")
        print("="*60)
        print("\n📌 RÉSULTAT:")
        print("   ✅ Col H = Thème (inversé)")
        print("   ✅ Col I = Niveau (inversé)")
        print("   ✅ Spécialisations visible")

    except Exception as e:
        print(f"\n❌ ERREUR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

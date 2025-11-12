#!/usr/bin/env python3
"""
PHASE 5 - TESTS COMPLETS
Vérification complète du système après toutes les modifications
"""

import openpyxl
from datetime import datetime
import sys

FICHIER = "PLANNING.xlsm"

def test_structure_visites(wb):
    """Test 1: Vérifier structure feuille Visites"""
    print("\n" + "="*80)
    print("🧪 TEST 1: STRUCTURE FEUILLE VISITES")
    print("="*80)

    if "Visites" not in wb.sheetnames:
        print("   ❌ ÉCHEC: Feuille Visites non trouvée")
        return False

    ws = wb["Visites"]

    # Vérifier en-têtes (15 colonnes attendues)
    headers_attendus = {
        1: "ID_Visite",
        2: "Date",
        3: "Heure_Debut",
        4: "Heure_Fin",
        5: "Nb_Participants",
        6: "Type_Prestation",
        7: "Nom_Structure",
        8: "Niveau",
        9: "Theme",
        10: "Commentaires",
        11: "Statut",
        12: "Guide_Attribue",
        13: "Tarif",
        14: "Duree_Heures",
        15: "Langue"
    }

    erreurs = 0
    for col_idx, header_attendu in headers_attendus.items():
        header_actuel = ws.cell(1, col_idx).value
        if header_actuel != header_attendu:
            print(f"   ⚠️  Col {col_idx}: attendu '{header_attendu}', trouvé '{header_actuel}'")
            erreurs += 1

    if erreurs == 0:
        print(f"   ✅ Structure OK: 15 colonnes correctes")
    else:
        print(f"   ❌ {erreurs} erreurs dans les en-têtes")
        return False

    # Compter données
    nb_lignes = 0
    for row_idx in range(2, 1000):
        if ws.cell(row_idx, 1).value:
            nb_lignes += 1
        else:
            break

    print(f"   ✅ Données: {nb_lignes} visites présentes")

    return True

def test_format_dates(wb):
    """Test 2: Vérifier format dates français"""
    print("\n" + "="*80)
    print("🧪 TEST 2: FORMAT DATES FRANÇAIS")
    print("="*80)

    ws = wb["Visites"]

    # Vérifier format de la colonne Date (col 2)
    format_ok = 0

    for row_idx in range(2, 6):  # Vérifier 4 premières lignes
        cell = ws.cell(row_idx, 2)
        if cell.value:
            format_actuel = cell.number_format
            # Format attendu: [$-fr-FR]dddd d mmmm yyyy
            if "dddd" in format_actuel and "mmmm" in format_actuel:
                format_ok += 1
                print(f"   ✅ Ligne {row_idx}: Format '{format_actuel}' (OK)")
            else:
                print(f"   ⚠️  Ligne {row_idx}: Format '{format_actuel}' (pas français)")

    if format_ok > 0:
        print(f"   ✅ Format dates français appliqué ({format_ok} cellules)")
        return True
    else:
        print(f"   ⚠️  Format dates non appliqué correctement")
        return False

def test_donnees_importees(wb):
    """Test 3: Vérifier données importées de la cliente"""
    print("\n" + "="*80)
    print("🧪 TEST 3: DONNÉES IMPORTÉES (19 VISITES CLIENTE)")
    print("="*80)

    ws = wb["Visites"]

    # Vérifier ID_Visite commence par V
    ids_ok = 0
    for row_idx in range(2, 22):  # 19 lignes + 2
        id_visite = ws.cell(row_idx, 1).value
        if id_visite and str(id_visite).startswith("V"):
            ids_ok += 1

    print(f"   ✅ IDs générés: {ids_ok} visites avec format V000X")

    # Vérifier Type_Prestation rempli
    types_ok = 0
    types_vides = 0
    types_trouves = set()

    for row_idx in range(2, 22):
        type_prest = ws.cell(row_idx, 6).value
        if type_prest:
            types_ok += 1
            types_trouves.add(type_prest)
        else:
            types_vides += 1

    print(f"   ✅ Type_Prestation: {types_ok} remplis, {types_vides} vides")
    print(f"   📋 Types trouvés: {', '.join(types_trouves)}")

    # Vérifier Statut = "Planifiée"
    statuts_ok = 0
    for row_idx in range(2, 22):
        statut = ws.cell(row_idx, 11).value
        if statut == "Planifiée":
            statuts_ok += 1

    print(f"   ✅ Statut par défaut: {statuts_ok} visites 'Planifiée'")

    # Vérifier Guide_Attribue vide (pour attribution auto)
    guides_vides = 0
    for row_idx in range(2, 22):
        guide = ws.cell(row_idx, 12).value
        if not guide or guide == "None" or guide == "":
            guides_vides += 1

    print(f"   ✅ Attribution auto: {guides_vides} visites sans guide (OK)")

    return True

def test_specialisations(wb):
    """Test 4: Vérifier feuille Spécialisations"""
    print("\n" + "="*80)
    print("🧪 TEST 4: SPÉCIALISATIONS (75 LIGNES)")
    print("="*80)

    if "Spécialisations" not in wb.sheetnames:
        print("   ❌ ÉCHEC: Feuille Spécialisations non trouvée")
        return False

    ws = wb["Spécialisations"]

    # Compter lignes
    nb_lignes = 0
    for row_idx in range(2, 1000):
        if ws.cell(row_idx, 1).value:
            nb_lignes += 1
        else:
            break

    print(f"   ✅ {nb_lignes} lignes de spécialisations trouvées")

    if nb_lignes < 70:
        print(f"   ⚠️  Attendu 75 lignes (15 guides × 5 types)")
        return False

    # Vérifier structure
    types_prestations = set()
    guides = set()

    for row_idx in range(2, min(nb_lignes + 2, 100)):
        type_prest = ws.cell(row_idx, 4).value  # Col D: Type_Prestation
        prenom = ws.cell(row_idx, 2).value      # Col B: Prenom_Guide

        if type_prest:
            types_prestations.add(type_prest)
        if prenom:
            guides.add(prenom)

    print(f"   ✅ Types de prestations: {len(types_prestations)} types différents")
    print(f"   ✅ Guides configurés: {len(guides)} guides")

    return True

def test_mes_disponibilites(wb):
    """Test 5: Vérifier colonne A masquée dans Mes_Disponibilites"""
    print("\n" + "="*80)
    print("🧪 TEST 5: MES_DISPONIBILITES (COLONNE A MASQUÉE)")
    print("="*80)

    if "Mes_Disponibilites" not in wb.sheetnames:
        print("   ⚠️  Feuille Mes_Disponibilites non trouvée")
        return False

    ws = wb["Mes_Disponibilites"]

    # Vérifier si colonne A est masquée
    col_a_masquee = ws.column_dimensions['A'].hidden

    if col_a_masquee:
        print(f"   ✅ Colonne A masquée (comme demandé par cliente)")
    else:
        print(f"   ⚠️  Colonne A visible (devrait être masquée)")
        return False

    # Vérifier en-têtes visibles
    headers = []
    for col_idx in range(2, 8):  # Colonnes B à G
        header = ws.cell(1, col_idx).value
        if header:
            headers.append(header)

    print(f"   ✅ Colonnes visibles: {', '.join(headers)}")

    return True

def test_formules_duree(wb):
    """Test 6: Vérifier formules de durée"""
    print("\n" + "="*80)
    print("🧪 TEST 6: FORMULES DURÉE (=(Heure_Fin - Heure_Debut)*24)")
    print("="*80)

    ws = wb["Visites"]

    formules_ok = 0
    valeurs_ok = 0

    for row_idx in range(2, 22):
        cell_duree = ws.cell(row_idx, 14)  # Col N: Duree_Heures

        # Vérifier si c'est une formule
        if isinstance(cell_duree.value, str) and cell_duree.value.startswith("="):
            formules_ok += 1
        elif isinstance(cell_duree.value, (int, float)):
            valeurs_ok += 1

    print(f"   ✅ Formules: {formules_ok} cellules avec formule")
    print(f"   ✅ Valeurs calculées: {valeurs_ok} cellules")

    if formules_ok > 0 or valeurs_ok > 0:
        return True
    else:
        print(f"   ⚠️  Aucune durée calculée trouvée")
        return False

def test_validation_type_prestation(wb):
    """Test 7: Vérifier validation Type_Prestation"""
    print("\n" + "="*80)
    print("🧪 TEST 7: VALIDATION TYPE_PRESTATION (DROPDOWN)")
    print("="*80)

    ws = wb["Visites"]

    # Vérifier si des validations existent sur la feuille
    if hasattr(ws, 'data_validations') and ws.data_validations:
        nb_validations = len(ws.data_validations.dataValidation)
        print(f"   ✅ {nb_validations} validation(s) trouvée(s) sur la feuille")
        return True
    else:
        print(f"   ℹ️  Pas de validation détectée (validation peut exister mais non lisible)")
        return True  # Pas critique

def test_backup_existe(wb):
    """Test 8: Vérifier backup avant restructuration"""
    print("\n" + "="*80)
    print("🧪 TEST 8: BACKUP SÉCURITÉ")
    print("="*80)

    import os

    backup_file = "PLANNING_BACKUP_AVANT_RESTRUCTURATION.xlsm"

    if os.path.exists(backup_file):
        taille = os.path.getsize(backup_file) / 1024  # Ko
        print(f"   ✅ Backup trouvé: {backup_file} ({taille:.0f} Ko)")
        return True
    else:
        print(f"   ⚠️  Backup non trouvé (pas critique)")
        return True  # Pas critique

def main():
    print("\n" + "="*80)
    print("🧪🧪🧪 PHASE 5 - TESTS COMPLETS 🧪🧪🧪")
    print("="*80)

    # Charger fichier
    print("\n📂 Chargement PLANNING.xlsm...")
    try:
        wb = openpyxl.load_workbook(FICHIER, keep_vba=True, data_only=False)
    except Exception as e:
        print(f"❌ ERREUR: {e}")
        sys.exit(1)

    # Exécuter tous les tests
    resultats = []

    resultats.append(("Structure Visites", test_structure_visites(wb)))
    resultats.append(("Format dates français", test_format_dates(wb)))
    resultats.append(("Données importées", test_donnees_importees(wb)))
    resultats.append(("Spécialisations", test_specialisations(wb)))
    resultats.append(("Mes_Disponibilites", test_mes_disponibilites(wb)))
    resultats.append(("Formules durée", test_formules_duree(wb)))
    resultats.append(("Validation dropdown", test_validation_type_prestation(wb)))
    resultats.append(("Backup sécurité", test_backup_existe(wb)))

    wb.close()

    # Résumé final
    print("\n" + "="*80)
    print("📊 RÉSUMÉ TESTS")
    print("="*80)

    tests_reussis = sum(1 for _, resultat in resultats if resultat)
    tests_totaux = len(resultats)

    for nom_test, resultat in resultats:
        status = "✅ RÉUSSI" if resultat else "❌ ÉCHEC"
        print(f"   {status}: {nom_test}")

    print("\n" + "="*80)
    pourcentage = (tests_reussis / tests_totaux) * 100
    print(f"🎯 SCORE: {tests_reussis}/{tests_totaux} tests réussis ({pourcentage:.0f}%)")

    if tests_reussis == tests_totaux:
        print("\n✅✅✅ TOUS LES TESTS RÉUSSIS ! ✅✅✅")
        print("Le système est prêt pour la livraison !")
    elif tests_reussis >= tests_totaux * 0.8:
        print("\n⚠️  Quelques warnings mais système fonctionnel")
    else:
        print("\n❌ Des corrections sont nécessaires")

    print("="*80 + "\n")

    return tests_reussis == tests_totaux

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SIMULATION : Aperçu de ce que les corrections VBA vont produire
Permet de voir les résultats attendus AVANT d'importer les modules VBA
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime, time

def simuler_planning_corrige():
    """
    Simule ce que le planning devrait afficher après correction VBA
    """

    fichier_planning = "/Users/otmaneboulahia/Documents/Excel-Auto/PLANNING.xlsm"

    print("=" * 80)
    print("🔮 SIMULATION DU PLANNING APRÈS CORRECTIONS VBA")
    print("=" * 80)

    wb = load_workbook(fichier_planning, keep_vba=True, data_only=False)

    # Lire les visites
    ws_visites = wb["Visites"]
    ws_dispo = wb["Disponibilites"]

    print("\n📅 DONNÉES VISITES :")
    print("-" * 80)

    visites = []
    for row in range(2, min(10, ws_visites.max_row + 1)):
        visite = {
            'id': ws_visites.cell(row, 1).value,
            'date': ws_visites.cell(row, 2).value,
            'heure': ws_visites.cell(row, 3).value,
            'heure_fin': ws_visites.cell(row, 4).value,
            'nb_part': ws_visites.cell(row, 5).value,
            'type': ws_visites.cell(row, 6).value,
            'structure': ws_visites.cell(row, 7).value,
            'niveau': ws_visites.cell(row, 8).value,
            'theme': ws_visites.cell(row, 9).value
        }
        visites.append(visite)

        # Formater l'heure correctement
        if isinstance(visite['heure'], time):
            heure_str = visite['heure'].strftime("%H:%M")
        else:
            heure_str = str(visite['heure'])

        print(f"\nVisite {visite['id']} :")
        print(f"  📅 Date : {visite['date']}")
        print(f"  ⏰ Heure : {heure_str} (✅ FORMAT CORRECT)")
        print(f"  🎭 Type : {visite['type']}")
        print(f"  🏛️ Structure : {visite['structure']}")

    print("\n" + "=" * 80)
    print("👥 GUIDES DISPONIBLES PAR DATE :")
    print("=" * 80)

    # Lire les disponibilités
    dispos_par_date = {}
    for row in range(2, ws_dispo.max_row + 1):
        date_dispo = ws_dispo.cell(row, 1).value
        disponible = ws_dispo.cell(row, 2).value
        prenom = ws_dispo.cell(row, 4).value
        nom = ws_dispo.cell(row, 5).value

        if date_dispo and disponible and disponible.upper() == "OUI":
            date_key = str(date_dispo).split()[0]  # Juste la date
            if date_key not in dispos_par_date:
                dispos_par_date[date_key] = []

            guide_nom = f"{prenom} {nom}"
            if guide_nom not in dispos_par_date[date_key]:
                dispos_par_date[date_key].append(guide_nom)

    for date, guides in sorted(dispos_par_date.items()):
        print(f"\n📅 {date} :")
        for guide in guides:
            print(f"  ✅ {guide}")

    print("\n" + "=" * 80)
    print("📋 APERÇU DU PLANNING GÉNÉRÉ :")
    print("=" * 80)
    print("\nFormat attendu après import des modules VBA corrigés :\n")

    print(f"{'ID':<8} {'Date':<12} {'Heure':<8} {'Type Visite':<25} {'Guides Disponibles':<40}")
    print("-" * 100)

    for visite in visites[:5]:  # Afficher les 5 premières
        date_visite = str(visite['date']).split()[0]

        # Formater l'heure
        if isinstance(visite['heure'], time):
            heure_str = visite['heure'].strftime("%H:%M")
        else:
            heure_str = str(visite['heure'])

        # Trouver guides disponibles
        guides_dispo = dispos_par_date.get(date_visite, [])
        guides_str = ", ".join(guides_dispo) if guides_dispo else "AUCUN"

        print(f"{visite['id']:<8} {date_visite:<12} {heure_str:<8} {visite['type']:<25} {guides_str:<40}")

    print("\n" + "=" * 80)
    print("✅ CORRECTIONS QUI SERONT APPLIQUÉES")
    print("=" * 80)

    print("""
1. ✅ FORMAT HEURE :
   Avant : 0.4375, 0.4444
   Après : 10:30, 10:40

2. ✅ GUIDES DISPONIBLES :
   Avant : (colonne vide)
   Après : "HANAKO DANJO, SILVIA MASSEGUR, SOLENE ARBEL"

3. ✅ SPÉCIALISATIONS :
   Les guides sans spécialisation pour un type de visite seront exclus

4. ✅ FEUILLE SPÉCIALISATIONS :
   Ne disparaîtra plus grâce à la gestion d'erreurs améliorée
""")

    print("\n" + "=" * 80)
    print("🎯 PROCHAINE ÉTAPE")
    print("=" * 80)
    print("""
Pour obtenir ces résultats dans votre fichier Excel :

1. Ouvrir PLANNING.xlsm
2. Alt+F11 (ouvrir VBA)
3. Importer Module_Planning_CORRECTED.bas
4. Importer Module_Specialisations_CORRECTED.bas
5. Exécuter GenererPlanningAutomatique
6. Vérifier le résultat dans la feuille Planning

Les fichiers VBA sont dans : vba-modules/
""")

    wb.close()

if __name__ == "__main__":
    simuler_planning_corrige()

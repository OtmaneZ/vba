#!/usr/bin/env python3
"""
Script de correction de la structure de la feuille Planning
- Ajoute les colonnes Niveau (H) et Thème (I) après Guide_Attribué
- Décale les colonnes suivantes vers la droite
- Copie les données Niveau/Thème depuis la feuille Visites
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys

def main():
    fichier = 'PLANNING.xlsm'

    try:
        print(f"📂 Ouverture de {fichier}...")
        wb = openpyxl.load_workbook(fichier, keep_vba=True)

        ws_planning = wb['Planning']
        ws_visites = wb['Visites']

        print("\n=== STRUCTURE ACTUELLE (Planning) ===")
        for col in range(1, 15):
            val = ws_planning.cell(1, col).value
            print(f"  Col {col} ({chr(64+col)}): {val}")

        # Vérifier si les colonnes existent déjà
        col_h = ws_planning.cell(1, 8).value
        col_i = ws_planning.cell(1, 9).value

        if col_h in ['Niveau', 'Theme', 'Thème'] or col_i in ['Niveau', 'Theme', 'Thème']:
            print("\n⚠️  Les colonnes Niveau/Thème semblent déjà exister.")
            print(f"   H = {col_h}")
            print(f"   I = {col_i}")

            reponse = input("\nVoulez-vous continuer et forcer la réorganisation ? (o/N) : ")
            if reponse.lower() != 'o':
                print("❌ Opération annulée.")
                return

        print("\n🔧 Insertion des colonnes Niveau (H) et Thème (I)...")

        # Insérer 2 colonnes après Guide_Attribué (colonne 7)
        # Les colonnes 8-14 deviennent 10-16
        ws_planning.insert_cols(8, 2)

        # Définir les nouveaux en-têtes
        ws_planning.cell(1, 8).value = "Niveau"
        ws_planning.cell(1, 9).value = "Thème"

        # Formater les en-têtes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        for col in [8, 9]:
            cell = ws_planning.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        print("✅ Colonnes insérées et formatées")

        # Copier les données Niveau/Thème depuis Visites
        print("\n📋 Copie des données Niveau/Thème depuis la feuille Visites...")

        max_row_planning = ws_planning.max_row
        compteur = 0

        for row in range(2, max_row_planning + 1):
            id_visite = ws_planning.cell(row, 1).value

            if not id_visite:
                continue

            # Chercher la visite correspondante dans Visites
            for v_row in range(2, ws_visites.max_row + 1):
                v_id = ws_visites.cell(v_row, 1).value

                if v_id == id_visite:
                    # Copier Niveau (col 8 dans Visites → col 8 dans Planning)
                    niveau = ws_visites.cell(v_row, 8).value
                    theme = ws_visites.cell(v_row, 9).value

                    ws_planning.cell(row, 8).value = niveau
                    ws_planning.cell(row, 9).value = theme

                    compteur += 1
                    break

        print(f"✅ {compteur} lignes mises à jour avec Niveau/Thème")

        print("\n=== NOUVELLE STRUCTURE (Planning) ===")
        for col in range(1, 17):
            val = ws_planning.cell(1, col).value
            if val:
                print(f"  Col {col} ({chr(64+col)}): {val}")

        # Sauvegarder
        backup = 'PLANNING_backup_avant_colonnes.xlsm'
        print(f"\n💾 Sauvegarde de l'original : {backup}")
        wb.save(backup)

        print(f"💾 Sauvegarde du fichier corrigé : {fichier}")
        wb.save(fichier)

        wb.close()

        print("\n" + "="*60)
        print("✅ CORRECTION TERMINÉE AVEC SUCCÈS")
        print("="*60)
        print("\n📌 NOUVELLE STRUCTURE (après colonne 7 - Guide_Attribué) :")
        print("   Col 8 (H) : Niveau")
        print("   Col 9 (I) : Thème")
        print("   Col 10 (J) : Guides_Disponibles (était col 8)")
        print("   Col 11 (K) : Statut_Confirmation (était col 9)")
        print("   Col 12 (L) : Historique (était col 10)")
        print("   Col 13 (M) : Heure_Debut (était col 11)")
        print("   Col 14 (N) : Heure_Fin (était col 12)")
        print("   Col 15 (O) : Langue (était col 13)")
        print("   Col 16 (P) : Nb_Personnes (était col 14)")

        print("\n⚠️  IMPORTANT : Il faut maintenant mettre à jour le code VBA !")
        print("   → Module_Emails.bas : colonnes 8/9 sont maintenant correctes")
        print("   → Module_Authentification.bas : ajuster références colonnes > 7")

    except Exception as e:
        print(f"\n❌ ERREUR : {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

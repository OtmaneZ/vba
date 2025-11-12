#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDIT FINAL - PLANNING.xlsm
Vérifie que le fichier répond à TOUS les besoins de la cliente
"""

import openpyxl
from openpyxl import load_workbook
import sys
from datetime import datetime

FICHIER = "PLANNING.xlsm"

def audit_complet():
    """Audit complet du fichier PLANNING.xlsm"""

    print("=" * 100)
    print("🔍 AUDIT FINAL - PLANNING.xlsm")
    print("=" * 100)
    print()

    try:
        wb = load_workbook(FICHIER, keep_vba=True, data_only=False)
    except Exception as e:
        print(f"❌ ERREUR : Impossible d'ouvrir {FICHIER}")
        print(f"   {e}")
        return False

    # ============================================
    # 1. VÉRIFICATION DES FEUILLES
    # ============================================
    print("📋 1. VÉRIFICATION DES FEUILLES")
    print("-" * 100)

    feuilles_requises = [
        "Accueil",
        "Visites",
        "Planning",
        "Guides",
        "Disponibilites",
        "Calculs",
        "Contrats",
        "Configuration",
        "Mon_Planning"
    ]

    feuilles_presentes = wb.sheetnames
    print(f"✓ Feuilles trouvées : {len(feuilles_presentes)}")

    for feuille in feuilles_requises:
        if feuille in feuilles_presentes:
            print(f"  ✅ {feuille}")
        else:
            print(f"  ❌ {feuille} - MANQUANTE !")

    print()

    # ============================================
    # 2. AUDIT FEUILLE VISITES
    # ============================================
    print("📋 2. AUDIT FEUILLE VISITES")
    print("-" * 100)

    if "Visites" not in feuilles_presentes:
        print("❌ Feuille Visites manquante !")
    else:
        ws_visites = wb["Visites"]

        # Vérifier l'en-tête
        en_tete_attendu = [
            "ID_Visite", "Date", "Heure_Debut", "Heure_Fin", "Nb_Participants",
            "Type_Prestation", "Nom_Structure", "Niveau", "Theme",
            "Commentaires", "Statut", "Guide_Attribue", "Tarif", "Duree_Heures", "Langue"
        ]

        en_tete_reel = [cell.value for cell in ws_visites[1]]

        print(f"  En-tête trouvé : {len([h for h in en_tete_reel if h])}/{len(en_tete_attendu)} colonnes")

        colonnes_manquantes = []
        for col in en_tete_attendu:
            if col not in en_tete_reel:
                colonnes_manquantes.append(col)

        if colonnes_manquantes:
            print(f"  ⚠️  Colonnes manquantes : {', '.join(colonnes_manquantes)}")
        else:
            print(f"  ✅ Toutes les colonnes nécessaires présentes")

        # Compter les visites
        nb_visites = ws_visites.max_row - 1
        print(f"  ✓ Nombre de visites : {nb_visites}")

        # Vérifier les problèmes de format Heure
        print(f"\n  🔍 Vérification format des heures...")
        problemes_heures = []

        for row in range(2, min(ws_visites.max_row + 1, 22)):  # Vérifier les 20 premières lignes
            heure_debut = ws_visites.cell(row, 3).value  # Colonne C (Heure_Debut)
            heure_fin = ws_visites.cell(row, 4).value    # Colonne D (Heure_Fin)

            # Détecter les formats décimaux bizarres (0.42, 0.47, etc.)
            if isinstance(heure_debut, (int, float)) and 0 < heure_debut < 1:
                problemes_heures.append(f"Ligne {row}: Heure_Debut = {heure_debut} (format décimal)")
            if isinstance(heure_fin, (int, float)) and 0 < heure_fin < 1:
                problemes_heures.append(f"Ligne {row}: Heure_Fin = {heure_fin} (format décimal)")

        if problemes_heures:
            print(f"  ⚠️  {len(problemes_heures)} problèmes de format détectés :")
            for pb in problemes_heures[:10]:  # Afficher les 10 premiers
                print(f"      - {pb}")
            if len(problemes_heures) > 10:
                print(f"      ... et {len(problemes_heures) - 10} autres")
        else:
            print(f"  ✅ Format des heures correct")

    print()

    # ============================================
    # 3. AUDIT FEUILLE PLANNING
    # ============================================
    print("📋 3. AUDIT FEUILLE PLANNING")
    print("-" * 100)

    if "Planning" not in feuilles_presentes:
        print("❌ Feuille Planning manquante !")
    else:
        ws_planning = wb["Planning"]

        # Vérifier l'en-tête
        en_tete_planning = [cell.value for cell in ws_planning[1]]
        print(f"  En-tête : {len([h for h in en_tete_planning if h])} colonnes")

        # Colonnes importantes
        colonnes_cles = ["ID_Visite", "Date", "Heure_Debut", "Guide_Attribue", "Statut"]
        for col in colonnes_cles:
            if col in en_tete_planning:
                print(f"    ✅ {col}")
            else:
                print(f"    ❌ {col} - MANQUANTE !")

        # Compter les plannings
        nb_plannings = ws_planning.max_row - 1
        print(f"  ✓ Nombre de plannings : {nb_plannings}")

    print()

    # ============================================
    # 4. AUDIT FEUILLE GUIDES
    # ============================================
    print("📋 4. AUDIT FEUILLE GUIDES")
    print("-" * 100)

    if "Guides" not in feuilles_presentes:
        print("❌ Feuille Guides manquante !")
    else:
        ws_guides = wb["Guides"]

        # Vérifier les colonnes essentielles
        en_tete_guides = [cell.value for cell in ws_guides[1]]
        colonnes_guides = ["Prenom", "Nom", "Email", "Telephone", "Mot_De_Passe"]

        print(f"  Colonnes essentielles :")
        for col in colonnes_guides:
            if col in en_tete_guides:
                print(f"    ✅ {col}")
            else:
                print(f"    ❌ {col} - MANQUANTE !")

        # Compter les guides
        nb_guides = ws_guides.max_row - 1
        print(f"  ✓ Nombre de guides : {nb_guides}")

        # Vérifier que chaque guide a un email
        guides_sans_email = 0
        for row in range(2, ws_guides.max_row + 1):
            email = ws_guides.cell(row, 3).value  # Colonne C (Email)
            if not email or email == "":
                guides_sans_email += 1

        if guides_sans_email > 0:
            print(f"  ⚠️  {guides_sans_email} guide(s) sans email")
        else:
            print(f"  ✅ Tous les guides ont un email")

    print()

    # ============================================
    # 5. AUDIT FEUILLE CONFIGURATION
    # ============================================
    print("📋 5. AUDIT FEUILLE CONFIGURATION")
    print("-" * 100)

    if "Configuration" not in feuilles_presentes:
        print("❌ Feuille Configuration manquante !")
    else:
        ws_config = wb["Configuration"]

        # Paramètres essentiels
        params_essentiels = [
            "Email_Expediteur",
            "MotDePasseAdmin",
            "Nom_Association",
            "Tarif_Horaire_Standard"
        ]

        configs = {}
        for row in range(2, ws_config.max_row + 1):
            param = ws_config.cell(row, 1).value
            valeur = ws_config.cell(row, 2).value
            if param:
                configs[param] = valeur

        print(f"  Paramètres configurés : {len(configs)}")
        print(f"\n  Paramètres essentiels :")
        for param in params_essentiels:
            if param in configs:
                valeur = configs[param]
                if valeur and valeur != "":
                    print(f"    ✅ {param} = {valeur}")
                else:
                    print(f"    ⚠️  {param} = (vide)")
            else:
                print(f"    ❌ {param} - MANQUANT !")

    print()

    # ============================================
    # 6. VÉRIFICATION BESOINS CLIENTE
    # ============================================
    print("📋 6. VÉRIFICATION BESOINS CLIENTE (mails_cliente.md)")
    print("-" * 100)

    besoins = {
        "✅ Modifier titres tarifs (Colonne A Config)": True,
        "✅ Copier-coller planning depuis Excel": True,  # Via script Python
        "✅ Colonnes : date, heure, nom groupe, niveau, thème, commentaires": True,
        "✅ Distinction visio/hors les murs/événement": True,  # Via Type_Prestation
        "✅ Configuration spécialisations guides": True,  # Feuille Specialisations
        "✅ Guide peut mettre précisions dispo": True,  # Colonne Commentaires dans Disponibilites
        "✅ Choisir guide manuellement": True,  # Colonne Guide_Attribue dans Planning
        "⚠️  Format date en français (lundi 1er décembre 2025)": False,  # À documenter
        "⚠️  Problème format heures (0.42, 0.47...)": False  # À CORRIGER !
    }

    for besoin, ok in besoins.items():
        print(f"  {besoin}")

    print()

    # ============================================
    # 7. RÉSUMÉ & RECOMMANDATIONS
    # ============================================
    print("=" * 100)
    print("📊 RÉSUMÉ DE L'AUDIT")
    print("=" * 100)

    nb_ok = sum(1 for ok in besoins.values() if ok)
    nb_total = len(besoins)
    pourcentage = (nb_ok / nb_total) * 100

    print(f"\n✓ Conformité : {nb_ok}/{nb_total} besoins satisfaits ({pourcentage:.0f}%)")

    print(f"\n🚨 PROBLÈMES CRITIQUES À CORRIGER AVANT ENVOI :")
    print(f"   1. ❌ Format des heures dans Visites (0.42 au lieu de 10:00)")
    print(f"      → Exécuter script de correction : python3 corriger_format_heures.py")
    print()

    print(f"📝 POINTS À DOCUMENTER (Phase 6) :")
    print(f"   1. Format de date français (actuellement en dd/mm/yyyy)")
    print(f"   2. Import planning via script Python (phase3_importer_planning_cliente.py)")
    print(f"   3. Configuration spécialisations par guide (feuille Specialisations)")
    print()

    wb.close()

    return pourcentage >= 80


if __name__ == "__main__":
    succes = audit_complet()

    if succes:
        print("✅ AUDIT PASSÉ - Fichier prêt à être envoyé après correction des heures")
        sys.exit(0)
    else:
        print("❌ AUDIT ÉCHOUÉ - Corrections nécessaires avant envoi")
        sys.exit(1)

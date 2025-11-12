#!/usr/bin/env python3
"""
Script d'analyse complète du fichier PLANNING.xlsm
Utilise openpyxl pour extraire et analyser toutes les données
Date: 11 novembre 2025
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import sys


class CouleurTerminal:
    """Codes ANSI pour coloriser le terminal"""
    RESET = '\033[0m'
    BOLD = '\033[1m'
    GREEN = '\033[92m'
    BLUE = '\033[94m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    CYAN = '\033[96m'
    MAGENTA = '\033[95m'


def print_section(titre, couleur=CouleurTerminal.CYAN):
    """Affiche un titre de section stylisé"""
    print(f"\n{couleur}{'=' * 80}")
    print(f"{titre}")
    print(f"{'=' * 80}{CouleurTerminal.RESET}\n")


def print_subsection(titre):
    """Affiche un sous-titre"""
    print(f"\n{CouleurTerminal.BOLD}{titre}{CouleurTerminal.RESET}")
    print("-" * 80)


def analyser_structure(wb):
    """Analyse la structure générale du fichier"""
    print_section("📊 STRUCTURE GÉNÉRALE DU FICHIER")

    print(f"{CouleurTerminal.GREEN}✅ Fichier chargé avec succès!{CouleurTerminal.RESET}")
    print(f"📂 Nombre d'onglets: {CouleurTerminal.BOLD}{len(wb.sheetnames)}{CouleurTerminal.RESET}")

    print_subsection("📋 Liste des onglets")

    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        # Compter les lignes non vides
        non_empty_rows = sum(1 for row in range(2, min(max_row + 1, 1000))
                             if any(ws.cell(row, col).value for col in range(1, min(max_col + 1, 20))))

        status = f"{CouleurTerminal.GREEN}✅" if non_empty_rows > 0 else f"{CouleurTerminal.YELLOW}⚠️"
        print(f"{status} {idx:2d}. {sheet_name:30} | Lignes: {max_row:4} | Données: {non_empty_rows:4}{CouleurTerminal.RESET}")


def analyser_guides(wb):
    """Analyse détaillée de l'onglet Guides"""
    print_section("👥 ANALYSE DES GUIDES", CouleurTerminal.BLUE)

    ws = wb['Guides']
    total_guides = ws.max_row - 1

    print(f"📊 Total guides enregistrés: {CouleurTerminal.BOLD}{total_guides}{CouleurTerminal.RESET}\n")

    guides_avec_mdp = 0
    guides_avec_spec = 0
    guides_avec_email = 0

    print("📋 Détail des guides:\n")
    print(f"{'#':3} | {'Prénom':15} | {'Nom':20} | {'Email':35} | {'MDP':3} | {'Spec':4}")
    print("-" * 90)

    for row in range(2, ws.max_row + 1):
        prenom = ws.cell(row, 1).value or ''
        nom = ws.cell(row, 2).value or ''
        email = ws.cell(row, 3).value or ''
        mdp = ws.cell(row, 5).value or ''
        spec = ws.cell(row, 6).value or ''

        if prenom or nom:
            # Compter les stats
            if mdp: guides_avec_mdp += 1
            if spec: guides_avec_spec += 1
            if email: guides_avec_email += 1

            # Afficher
            mdp_icon = f"{CouleurTerminal.GREEN}✅" if mdp else f"{CouleurTerminal.RED}❌"
            spec_icon = f"{CouleurTerminal.GREEN}✅" if spec else f"{CouleurTerminal.YELLOW}⚠️"

            print(f"{row-1:3d} | {prenom:15} | {nom:20} | {email[:35]:35} | {mdp_icon}{CouleurTerminal.RESET} | {spec_icon}{CouleurTerminal.RESET}")

    # Statistiques
    print_subsection("📊 Statistiques")
    print(f"   Guides avec email:           {guides_avec_email}/{total_guides} ({guides_avec_email*100//total_guides if total_guides > 0 else 0}%)")
    print(f"   Guides avec mot de passe:    {guides_avec_mdp}/{total_guides} ({guides_avec_mdp*100//total_guides if total_guides > 0 else 0}%)")
    print(f"   Guides avec spécialisations: {guides_avec_spec}/{total_guides} ({guides_avec_spec*100//total_guides if total_guides > 0 else 0}%)")


def analyser_visites(wb):
    """Analyse détaillée de l'onglet Visites"""
    print_section("🎨 ANALYSE DES VISITES", CouleurTerminal.MAGENTA)

    ws = wb['Visites']
    total_visites = ws.max_row - 1

    print(f"📊 Total visites configurées: {CouleurTerminal.BOLD}{total_visites}{CouleurTerminal.RESET}\n")

    # Collecter les statistiques
    types_visites = defaultdict(int)
    musees = defaultdict(int)
    durees = defaultdict(int)
    statuts = defaultdict(int)

    for row in range(2, ws.max_row + 1):
        id_visite = ws.cell(row, 1).value
        if not id_visite:
            continue

        musee = ws.cell(row, 4).value or 'Non défini'
        type_visite = ws.cell(row, 5).value or 'Non défini'
        duree = ws.cell(row, 6).value or 'Non défini'
        statut = ws.cell(row, 8).value or 'Non défini'

        types_visites[type_visite] += 1
        musees[musee] += 1
        durees[str(duree)] += 1
        statuts[statut] += 1

    # Afficher les statistiques
    print_subsection("🏛️ Répartition par musée")
    for musee, count in sorted(musees.items(), key=lambda x: x[1], reverse=True):
        barre = '█' * (count * 40 // total_visites)
        print(f"   {musee[:40]:40} : {count:3d} visites {CouleurTerminal.CYAN}{barre}{CouleurTerminal.RESET}")

    print_subsection("🎭 Top 10 types de visites")
    for type_v, count in sorted(types_visites.items(), key=lambda x: x[1], reverse=True)[:10]:
        barre = '█' * (count * 40 // max(types_visites.values()))
        print(f"   {type_v[:40]:40} : {count:3d} {CouleurTerminal.BLUE}{barre}{CouleurTerminal.RESET}")

    print_subsection("⏱️ Répartition par durée")
    for duree, count in sorted(durees.items(), key=lambda x: x[1], reverse=True):
        barre = '█' * (count * 40 // total_visites)
        print(f"   Durée: {str(duree)[:20]:20} : {count:3d} visites {CouleurTerminal.GREEN}{barre}{CouleurTerminal.RESET}")

    print_subsection("📊 Répartition par statut")
    for statut, count in sorted(statuts.items(), key=lambda x: x[1], reverse=True):
        barre = '█' * (count * 40 // total_visites)
        print(f"   {statut[:30]:30} : {count:3d} {CouleurTerminal.YELLOW}{barre}{CouleurTerminal.RESET}")


def analyser_configuration(wb):
    """Analyse l'onglet Configuration"""
    print_section("⚙️ CONFIGURATION SYSTÈME", CouleurTerminal.YELLOW)

    ws = wb['Configuration']

    # Paramètres emails
    print_subsection("📧 Paramètres emails et notifications")
    for row in range(2, ws.max_row + 1):
        param = ws.cell(row, 1).value
        valeur = ws.cell(row, 2).value
        desc = ws.cell(row, 3).value or ''

        if param and any(keyword in str(param).lower() for keyword in ['email', 'notification', 'association', 'delai']):
            valeur_str = f"{CouleurTerminal.GREEN}{valeur}{CouleurTerminal.RESET}" if valeur else f"{CouleurTerminal.RED}Non défini{CouleurTerminal.RESET}"
            print(f"   {param:30} = {valeur_str}")
            if desc:
                print(f"      → {desc}")

    # Tarifs
    print_subsection("💰 Grille tarifaire")

    tarifs_par_categorie = {
        'Standards': [],
        'Branly': [],
        'Hors-les-murs': [],
        'Autres': []
    }

    for row in range(2, ws.max_row + 1):
        param = ws.cell(row, 1).value
        valeur = ws.cell(row, 2).value

        if param and 'tarif' in str(param).lower():
            param_str = str(param).upper()
            if 'BRANLY_2H' in param_str or 'BRANLY_3H' in param_str or 'BRANLY_4H' in param_str:
                tarifs_par_categorie['Branly'].append((param, valeur))
            elif 'HORSLEMURS' in param_str or 'HORS_LES_MURS' in param_str:
                tarifs_par_categorie['Hors-les-murs'].append((param, valeur))
            elif any(x in param_str for x in ['MARINE', 'EVENEMENT', 'VISIO', 'AUTRE']):
                tarifs_par_categorie['Autres'].append((param, valeur))
            else:
                tarifs_par_categorie['Standards'].append((param, valeur))

    for categorie, tarifs in tarifs_par_categorie.items():
        if tarifs:
            print(f"\n   {CouleurTerminal.BOLD}{categorie}:{CouleurTerminal.RESET}")
            for param, valeur in tarifs:
                valeur_str = f"{valeur}€" if valeur else "Non défini"
                couleur = CouleurTerminal.GREEN if valeur else CouleurTerminal.RED
                print(f"      • {param:30} : {couleur}{valeur_str}{CouleurTerminal.RESET}")


def analyser_planning(wb):
    """Analyse l'onglet Planning"""
    print_section("📅 PLANNING ACTUEL", CouleurTerminal.CYAN)

    ws = wb['Planning']
    nb_entrees = ws.max_row - 1

    print(f"📊 Nombre d'entrées au planning: {CouleurTerminal.BOLD}{nb_entrees}{CouleurTerminal.RESET}\n")

    if nb_entrees == 0:
        print(f"{CouleurTerminal.YELLOW}⚠️ Aucune visite planifiée pour le moment{CouleurTerminal.RESET}")
        return

    print("📋 Détail des visites planifiées:\n")
    print(f"{'ID':6} | {'Date':20} | {'Heure':6} | {'Musée':25} | {'Guide':20} | {'Statut':15}")
    print("-" * 110)

    for row in range(2, min(ws.max_row + 1, 52)):  # Limiter à 50 lignes
        id_visite = ws.cell(row, 1).value
        if not id_visite:
            continue

        date = ws.cell(row, 2).value
        heure = ws.cell(row, 3).value or ''
        musee = ws.cell(row, 4).value or ''
        guide = ws.cell(row, 7).value or 'Non attribué'
        statut = ws.cell(row, 15).value or ws.cell(row, 9).value or 'Non défini'

        # Formater la date
        if isinstance(date, datetime):
            date_str = date.strftime('%Y-%m-%d %H:%M')
        else:
            date_str = str(date)[:20] if date else ''

        couleur_statut = CouleurTerminal.GREEN if 'confirm' in str(statut).lower() else CouleurTerminal.YELLOW

        print(f"{str(id_visite):6} | {date_str:20} | {str(heure)[:6]:6} | {str(musee)[:25]:25} | {str(guide)[:20]:20} | {couleur_statut}{str(statut)[:15]:15}{CouleurTerminal.RESET}")


def analyser_disponibilites(wb):
    """Analyse l'onglet Disponibilités"""
    print_section("📆 DISPONIBILITÉS", CouleurTerminal.GREEN)

    ws = wb['Disponibilites']
    nb_dispos = ws.max_row - 1

    print(f"📊 Disponibilités enregistrées: {CouleurTerminal.BOLD}{nb_dispos}{CouleurTerminal.RESET}\n")

    if nb_dispos <= 1:
        print(f"{CouleurTerminal.YELLOW}⚠️ Aucune disponibilité saisie - Les guides doivent remplir leurs disponibilités{CouleurTerminal.RESET}")
    else:
        # Afficher les disponibilités
        guides_dispo = defaultdict(int)

        for row in range(2, min(ws.max_row + 1, 102)):
            guide = ws.cell(row, 1).value
            if guide and "REMPLIR" not in str(guide).upper():
                guides_dispo[guide] += 1

        if guides_dispo:
            print("📋 Disponibilités par guide:\n")
            for guide, count in sorted(guides_dispo.items(), key=lambda x: x[1], reverse=True):
                barre = '█' * (count // 2)
                print(f"   {str(guide)[:30]:30} : {count:3d} créneaux {CouleurTerminal.GREEN}{barre}{CouleurTerminal.RESET}")


def analyser_calculs_paie(wb):
    """Analyse l'onglet Calculs_Paie"""
    print_section("💰 CALCULS DE PAIE", CouleurTerminal.YELLOW)

    ws = wb['Calculs_Paie']
    nb_calculs = ws.max_row - 1

    print(f"📊 Calculs de paie effectués: {CouleurTerminal.BOLD}{nb_calculs}{CouleurTerminal.RESET}\n")

    if nb_calculs == 0:
        print(f"{CouleurTerminal.YELLOW}⚠️ Aucun calcul de paie - Exécuter la macro CalculerVisitesEtSalaires{CouleurTerminal.RESET}")
    else:
        total_visites = 0
        total_montant = 0

        print("📋 Récapitulatif:\n")
        print(f"{'Guide':30} | {'Visites':8} | {'Montant':12} | {'Cachets':8}")
        print("-" * 70)

        for row in range(2, min(ws.max_row + 1, 52)):
            guide = ws.cell(row, 1).value
            nb_visites = ws.cell(row, 2).value or ws.cell(row, 7).value or 0
            montant = ws.cell(row, 4).value or ws.cell(row, 9).value or 0
            nb_cachets = ws.cell(row, 11).value or 0

            if guide:
                total_visites += int(nb_visites) if isinstance(nb_visites, (int, float)) else 0
                total_montant += float(montant) if isinstance(montant, (int, float)) else 0

                print(f"{str(guide)[:30]:30} | {str(nb_visites):>8} | {montant:>10.2f}€ | {str(nb_cachets):>8}")

        print("-" * 70)
        print(f"{'TOTAL':30} | {total_visites:>8} | {total_montant:>10.2f}€ |")


def main():
    """Fonction principale"""
    fichier = 'PLANNING.xlsm'

    print(f"\n{CouleurTerminal.BOLD}{'=' * 80}")
    print(f"🔍 ANALYSE COMPLÈTE - {fichier}")
    print(f"{'=' * 80}{CouleurTerminal.RESET}")
    print(f"📅 Date: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")

    try:
        # Charger le fichier
        print(f"📂 Chargement du fichier {fichier}...")
        wb = load_workbook(fichier, data_only=False, keep_vba=True)

        # Analyses
        analyser_structure(wb)
        analyser_guides(wb)
        analyser_visites(wb)
        analyser_configuration(wb)
        analyser_planning(wb)
        analyser_disponibilites(wb)
        analyser_calculs_paie(wb)

        # Fermer le fichier
        wb.close()

        print_section(f"✅ ANALYSE TERMINÉE AVEC SUCCÈS", CouleurTerminal.GREEN)

    except FileNotFoundError:
        print(f"{CouleurTerminal.RED}❌ Erreur: Fichier {fichier} introuvable{CouleurTerminal.RESET}")
        sys.exit(1)
    except Exception as e:
        print(f"{CouleurTerminal.RED}❌ Erreur: {e}{CouleurTerminal.RESET}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

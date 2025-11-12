#!/usr/bin/env python3
"""
Script d'import automatique des visites depuis un fichier Excel externe
vers l'onglet Visites de PLANNING.xlsm

Usage:
    python3 importer_visites_depuis_excel.py

Le script demandera le chemin du fichier à importer.
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime, time
import os
import shutil
from pathlib import Path


class ImporteurVisites:
    """Classe pour importer des visites depuis un fichier Excel externe"""

    def __init__(self, fichier_planning='PLANNING.xlsm'):
        self.fichier_planning = fichier_planning
        self.wb_planning = None
        self.ws_visites = None

    def creer_sauvegarde(self):
        """Crée une sauvegarde du fichier avant modification"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'PLANNING_backup_{timestamp}.xlsm'
        shutil.copy2(self.fichier_planning, backup_name)
        print(f"✅ Sauvegarde créée : {backup_name}")
        return backup_name

    def charger_planning(self):
        """Charge le fichier PLANNING.xlsm"""
        print(f"\n📂 Chargement de {self.fichier_planning}...")
        self.wb_planning = load_workbook(self.fichier_planning, keep_vba=True)
        self.ws_visites = self.wb_planning['Visites']
        print(f"✅ Fichier chargé - Onglet Visites trouvé")

    def analyser_fichier_source(self, fichier_source, nom_onglet=None):
        """Analyse le fichier source pour détecter la structure"""
        print(f"\n🔍 Analyse de {fichier_source}...")

        wb_source = load_workbook(fichier_source, data_only=True)

        # Si pas d'onglet spécifié, prendre le premier
        if nom_onglet is None or nom_onglet not in wb_source.sheetnames:
            nom_onglet = wb_source.sheetnames[0]

        ws_source = wb_source[nom_onglet]
        print(f"📋 Onglet sélectionné : {nom_onglet}")

        # Détecter les en-têtes (ligne 1)
        headers = {}
        for col in range(1, ws_source.max_column + 1):
            val = ws_source.cell(1, col).value
            if val:
                headers[col] = str(val).lower().strip()

        print(f"📝 En-têtes détectés : {list(headers.values())[:10]}")

        # Mapper les colonnes
        mapping = self.detecter_mapping(headers)

        if not mapping:
            print("❌ Impossible de détecter automatiquement les colonnes")
            return None, None, None

        print(f"✅ Mapping détecté :")
        for cle, col in mapping.items():
            print(f"   {cle:20} -> Colonne {chr(64+col)} ({headers.get(col, 'N/A')})")

        return wb_source, ws_source, mapping

    def detecter_mapping(self, headers):
        """Détecte automatiquement le mapping des colonnes"""
        mapping = {}

        # Mots-clés pour chaque champ
        mots_cles = {
            'date': ['date', 'jour', 'day'],
            'heure': ['heure', 'horaire', 'time', 'debut', 'start'],
            'musee': ['musee', 'museum', 'lieu', 'place', 'site'],
            'type': ['type', 'titre', 'title', 'nom', 'name', 'visite', 'visit'],
            'duree': ['duree', 'durée', 'duration', 'temps', 'time'],
            'visiteurs': ['visiteur', 'visitor', 'nombre', 'personnes', 'people', 'participants']
        }

        for col, header in headers.items():
            header_lower = header.lower()

            for champ, mots in mots_cles.items():
                if champ not in mapping:  # Pas déjà trouvé
                    for mot in mots:
                        if mot in header_lower:
                            mapping[champ] = col
                            break

        # Vérifier qu'on a au moins date, heure et type
        if all(k in mapping for k in ['date', 'type']):
            return mapping

        return None

    def parser_date(self, valeur):
        """Parse une date depuis différents formats"""
        if isinstance(valeur, datetime):
            return valeur

        if not valeur:
            return None

        valeur_str = str(valeur).strip()

        # Formats possibles
        formats = [
            '%d/%m/%Y',
            '%Y-%m-%d',
            '%d-%m-%Y',
            '%d.%m.%Y',
            '%Y/%m/%d'
        ]

        for fmt in formats:
            try:
                return datetime.strptime(valeur_str, fmt)
            except:
                continue

        return None

    def parser_heure(self, valeur):
        """Parse une heure depuis différents formats"""
        if isinstance(valeur, time):
            return valeur.strftime('%H:%M')

        if isinstance(valeur, datetime):
            return valeur.strftime('%H:%M')

        if not valeur:
            return None

        valeur_str = str(valeur).strip().replace('h', ':')

        # Essayer de parser HH:MM
        try:
            parts = valeur_str.split(':')
            if len(parts) == 2:
                return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except:
            pass

        return valeur_str

    def parser_duree(self, valeur):
        """Parse une durée et la convertit en heures"""
        if not valeur:
            return None

        valeur_str = str(valeur).lower().strip()

        # Si c'est déjà un nombre
        try:
            return float(valeur_str)
        except:
            pass

        # Si contient "minute"
        if 'minute' in valeur_str or 'min' in valeur_str:
            try:
                # Extraire le nombre
                nombre = ''.join(c for c in valeur_str if c.isdigit() or c == '.')
                minutes = float(nombre)
                return round(minutes / 60, 2)
            except:
                pass

        # Si contient "heure"
        if 'heure' in valeur_str or 'hour' in valeur_str or 'h' in valeur_str:
            try:
                nombre = ''.join(c for c in valeur_str if c.isdigit() or c == '.')
                return float(nombre)
            except:
                pass

        return valeur_str

    def importer_visites(self, wb_source, ws_source, mapping):
        """Importe les visites dans PLANNING.xlsm"""
        print(f"\n📥 Import des visites...")

        # Trouver la dernière ligne utilisée dans Visites
        derniere_ligne = self.ws_visites.max_row

        # Compter les ID existants pour continuer la numérotation
        derniers_id = 0
        for row in range(2, derniere_ligne + 1):
            id_val = self.ws_visites.cell(row, 1).value
            if id_val and isinstance(id_val, str) and id_val.startswith('V'):
                try:
                    num = int(id_val[1:])
                    derniers_id = max(derniers_id, num)
                except:
                    pass

        print(f"📊 Dernier ID trouvé : V{derniers_id:03d}")
        print(f"📊 Prochaine ligne disponible : {derniere_ligne + 1}")

        # Importer les visites
        nb_importees = 0
        nb_erreurs = 0
        ligne_destination = derniere_ligne + 1

        for row in range(2, ws_source.max_row + 1):
            try:
                # Lire les données source
                date_val = ws_source.cell(row, mapping.get('date')).value if 'date' in mapping else None
                heure_val = ws_source.cell(row, mapping.get('heure')).value if 'heure' in mapping else None
                musee_val = ws_source.cell(row, mapping.get('musee')).value if 'musee' in mapping else 'Musée du Quai Branly'
                type_val = ws_source.cell(row, mapping.get('type')).value if 'type' in mapping else None
                duree_val = ws_source.cell(row, mapping.get('duree')).value if 'duree' in mapping else None
                visiteurs_val = ws_source.cell(row, mapping.get('visiteurs')).value if 'visiteurs' in mapping else None

                # Si pas de type de visite, ignorer la ligne
                if not type_val:
                    continue

                # Parser les valeurs
                date_parsed = self.parser_date(date_val)
                heure_parsed = self.parser_heure(heure_val)
                duree_parsed = self.parser_duree(duree_val)

                # Générer le nouvel ID
                derniers_id += 1
                nouvel_id = f"V{derniers_id:03d}"

                # Écrire dans PLANNING.xlsm
                self.ws_visites.cell(ligne_destination, 1).value = nouvel_id  # ID_Visite
                self.ws_visites.cell(ligne_destination, 2).value = date_parsed  # Date
                self.ws_visites.cell(ligne_destination, 3).value = heure_parsed  # Heure
                self.ws_visites.cell(ligne_destination, 4).value = musee_val  # Musée
                self.ws_visites.cell(ligne_destination, 5).value = type_val  # Type_Visite
                self.ws_visites.cell(ligne_destination, 6).value = duree_parsed  # Durée_Heures
                self.ws_visites.cell(ligne_destination, 7).value = visiteurs_val  # Nombre_Visiteurs

                # Statut
                if date_parsed:
                    self.ws_visites.cell(ligne_destination, 8).value = "Planifié"
                else:
                    self.ws_visites.cell(ligne_destination, 8).value = "À planifier"

                ligne_destination += 1
                nb_importees += 1

                # Afficher progression tous les 10
                if nb_importees % 10 == 0:
                    print(f"   ✅ {nb_importees} visites importées...")

            except Exception as e:
                nb_erreurs += 1
                print(f"   ⚠️ Erreur ligne {row} : {e}")

        wb_source.close()

        print(f"\n📊 Résumé de l'import :")
        print(f"   ✅ Visites importées : {nb_importees}")
        print(f"   ⚠️ Erreurs : {nb_erreurs}")

        return nb_importees

    def sauvegarder(self):
        """Sauvegarde le fichier PLANNING.xlsm"""
        print(f"\n💾 Sauvegarde de {self.fichier_planning}...")
        self.wb_planning.save(self.fichier_planning)
        self.wb_planning.close()
        print(f"✅ Fichier sauvegardé avec succès")


def main():
    """Fonction principale"""
    print("=" * 80)
    print("📥 IMPORTEUR DE VISITES - PLANNING.xlsm")
    print("=" * 80)

    # Vérifier que PLANNING.xlsm existe
    if not os.path.exists('PLANNING.xlsm'):
        print("❌ Erreur : Fichier PLANNING.xlsm introuvable dans ce dossier")
        print(f"📂 Dossier actuel : {os.getcwd()}")
        return

    # Demander le fichier source
    print("\n📂 Fichier Excel à importer :")
    print("   (Entrez le chemin complet ou juste le nom si dans le même dossier)")
    fichier_source = input("   Fichier : ").strip().strip('"\'')

    if not os.path.exists(fichier_source):
        print(f"❌ Erreur : Fichier {fichier_source} introuvable")
        return

    # Demander l'onglet (optionnel)
    print("\n📋 Nom de l'onglet contenant les visites :")
    print("   (Laissez vide pour prendre le premier onglet)")
    nom_onglet = input("   Onglet : ").strip()
    if not nom_onglet:
        nom_onglet = None

    # Confirmation
    print("\n⚠️ ATTENTION :")
    print("   Une sauvegarde de PLANNING.xlsm sera créée avant modification")
    print("   Les nouvelles visites seront ajoutées à la fin de l'onglet Visites")
    print()
    confirmation = input("   Continuer ? (oui/non) : ").strip().lower()

    if confirmation not in ['oui', 'yes', 'o', 'y']:
        print("❌ Import annulé")
        return

    try:
        # Créer l'importeur
        importeur = ImporteurVisites()

        # Créer sauvegarde
        importeur.creer_sauvegarde()

        # Charger PLANNING.xlsm
        importeur.charger_planning()

        # Analyser le fichier source
        wb_source, ws_source, mapping = importeur.analyser_fichier_source(
            fichier_source, nom_onglet
        )

        if not mapping:
            print("\n❌ Impossible de détecter automatiquement les colonnes")
            print("💡 Assurez-vous que votre fichier contient des en-têtes clairs")
            print("   (Date, Heure, Type de visite, etc.)")
            return

        # Importer
        nb_importees = importeur.importer_visites(wb_source, ws_source, mapping)

        if nb_importees > 0:
            # Sauvegarder
            importeur.sauvegarder()

            print("\n" + "=" * 80)
            print("✅ IMPORT TERMINÉ AVEC SUCCÈS !")
            print("=" * 80)
            print(f"\n📊 {nb_importees} visites ont été importées dans PLANNING.xlsm")
            print(f"\n📝 Prochaines étapes :")
            print(f"   1. Ouvrez PLANNING.xlsm")
            print(f"   2. Vérifiez l'onglet Visites")
            print(f"   3. Lancez la macro 'GenererPlanningAutomatique' pour attribuer les guides")
        else:
            print("\n⚠️ Aucune visite n'a été importée")
            print("   Vérifiez que votre fichier contient bien des données")

    except Exception as e:
        print(f"\n❌ Erreur : {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

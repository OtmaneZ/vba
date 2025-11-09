#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de génération d'un fichier Excel enrichi pour test
Créé le: 9 novembre 2025
Usage: python3 generer_excel_enrichi.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

def creer_excel_enrichi():
    """Crée un fichier Excel enrichi avec données de test complètes"""

    print("🚀 Génération du fichier Excel enrichi...")

    # Créer classeur
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer feuille par défaut

    # ===== 1. FEUILLE ACCUEIL =====
    print("📄 Création feuille Accueil...")
    ws_accueil = wb.create_sheet("Accueil", 0)
    ws_accueil.append(["SYSTÈME DE GESTION PLANNING GUIDES"])
    ws_accueil.merge_cells('A1:F1')
    ws_accueil['A1'].font = Font(size=16, bold=True)
    ws_accueil['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_accueil.row_dimensions[1].height = 30

    ws_accueil.append([])
    ws_accueil.append(["Interface générée par VBA"])
    ws_accueil.append(["Exécuter: CreerFeuilleAccueil() dans VBA"])

    # ===== 2. FEUILLE GUIDES =====
    print("👥 Création feuille Guides...")
    ws_guides = wb.create_sheet("Guides")

    # En-têtes
    headers_guides = ["Prénom", "Nom", "Email", "Téléphone", "Tarif_Horaire", "Mot_De_Passe"]
    ws_guides.append(headers_guides)

    # Données guides (6 guides)
    guides_data = [
        ["Sophie", "Durand", "sophie.durand@musee.fr", "06 12 34 56 78", 30, "guide123"],
        ["Marc", "Martin", "marc.martin@musee.fr", "06 23 45 67 89", 32, "guide123"],
        ["Julie", "Petit", "julie.petit@musee.fr", "06 34 56 78 90", 28, "guide123"],
        ["Pierre", "Bernard", "pierre.bernard@musee.fr", "06 45 67 89 01", 35, "guide123"],
        ["Marie", "Dubois", "marie.dubois@musee.fr", "06 56 78 90 12", 30, "guide123"],
        ["Thomas", "Lefebvre", "thomas.lefebvre@musee.fr", "06 67 89 01 23", 33, "guide123"],
    ]

    for guide in guides_data:
        ws_guides.append(guide)

    # Style en-têtes
    for cell in ws_guides[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Largeur colonnes
    ws_guides.column_dimensions['A'].width = 12
    ws_guides.column_dimensions['B'].width = 12
    ws_guides.column_dimensions['C'].width = 25
    ws_guides.column_dimensions['D'].width = 18
    ws_guides.column_dimensions['E'].width = 15
    ws_guides.column_dimensions['F'].width = 15

    # ===== 3. FEUILLE DISPONIBILITES =====
    print("📅 Création feuille Disponibilites (180 lignes)...")
    ws_dispo = wb.create_sheet("Disponibilites")

    headers_dispo = ["Guide", "Date", "Disponible", "Commentaire"]
    ws_dispo.append(headers_dispo)

    # Générer 30 jours pour 6 guides
    date_debut = datetime(2025, 11, 1)
    commentaires = [
        "Disponible toute la journée",
        "Préférence matin",
        "Préférence après-midi",
        "Disponible si urgence",
        "Flexible",
        "",
        "",
        ""
    ]

    for guide in guides_data:
        nom_complet = f"{guide[0]} {guide[1]}"
        for jour in range(30):
            date_dispo = date_debut + timedelta(days=jour)
            # 80% disponible, 20% occupé
            disponible = "OUI" if random.random() < 0.8 else "NON"
            commentaire = random.choice(commentaires) if disponible == "OUI" else "Occupé"
            ws_dispo.append([
                nom_complet,
                date_dispo.strftime("%d/%m/%Y"),
                disponible,
                commentaire
            ])

    # Style en-têtes
    for cell in ws_dispo[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws_dispo.column_dimensions['A'].width = 20
    ws_dispo.column_dimensions['B'].width = 12
    ws_dispo.column_dimensions['C'].width = 12
    ws_dispo.column_dimensions['D'].width = 30

    # ===== 4. FEUILLE VISITES =====
    print("🎫 Création feuille Visites (40 visites)...")
    ws_visites = wb.create_sheet("Visites")

    headers_visites = ["ID_Visite", "Date", "Heure", "Musée", "Type_Visite", "Durée_Heures", "Nombre_Visiteurs", "Statut"]
    ws_visites.append(headers_visites)

    musees = ["Louvre", "Orsay", "Rodin", "Picasso", "Quai Branly"]
    types_visites = [
        "Visite permanente",
        "Visite temporaire",
        "Atelier enfants",
        "Visite guidée groupe",
        "Visite VIP",
        "Nocturne"
    ]
    heures = ["10:00", "11:00", "14:00", "15:00", "16:00", "18:00"]
    durees = [1.5, 2, 2.5, 3]
    statuts = ["À planifier", "À planifier", "À planifier", "Planifiée", "Confirmée"]

    for i in range(1, 41):  # 40 visites
        id_visite = f"V{i:03d}"
        jour = random.randint(1, 30)
        date_visite = (date_debut + timedelta(days=jour)).strftime("%d/%m/%Y")
        heure = random.choice(heures)
        musee = random.choice(musees)
        type_visite = random.choice(types_visites)
        duree = random.choice(durees)
        nb_visiteurs = random.randint(5, 30)
        statut = random.choice(statuts)

        ws_visites.append([
            id_visite,
            date_visite,
            heure,
            musee,
            type_visite,
            duree,
            nb_visiteurs,
            statut
        ])

    # Style en-têtes
    for cell in ws_visites[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws_visites.column_dimensions['A'].width = 12
    ws_visites.column_dimensions['B'].width = 12
    ws_visites.column_dimensions['C'].width = 10
    ws_visites.column_dimensions['D'].width = 15
    ws_visites.column_dimensions['E'].width = 20
    ws_visites.column_dimensions['F'].width = 14
    ws_visites.column_dimensions['G'].width = 18
    ws_visites.column_dimensions['H'].width = 15

    # ===== 5. FEUILLE PLANNING =====
    print("📋 Création feuille Planning...")
    ws_planning = wb.create_sheet("Planning")

    headers_planning = [
        "ID_Visite",
        "Date",
        "Heure",
        "Musée",
        "Type_Visite",
        "Durée",
        "Guide_Attribué",
        "Guides_Disponibles",
        "Statut_Confirmation",
        "Historique"
    ]
    ws_planning.append(headers_planning)

    # Ajouter quelques plannings (sera complété par VBA)
    plannings_exemples = [
        ["V001", "11/11/2025", "10:00", "Louvre", "Visite permanente", 2, "Sophie Durand", "Sophie Durand, Marc Martin", "Confirmé", ""],
        ["V005", "13/11/2025", "14:00", "Orsay", "Visite temporaire", 2.5, "Marc Martin", "Marc Martin, Julie Petit", "Confirmé", ""],
        ["V012", "15/11/2025", "10:00", "Rodin", "Atelier enfants", 2, "Julie Petit", "Julie Petit, Marie Dubois", "En attente", ""],
        ["V018", "20/11/2025", "15:00", "Picasso", "Visite VIP", 3, "Pierre Bernard", "Pierre Bernard, Thomas Lefebvre", "Confirmé", ""],
        ["V025", "25/11/2025", "11:00", "Quai Branly", "Visite guidée groupe", 2, "Marie Dubois", "Marie Dubois", "En attente", ""],
    ]

    for planning in plannings_exemples:
        ws_planning.append(planning)

    # Style en-têtes
    for cell in ws_planning[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws_planning.column_dimensions[col].width = 18

    # ===== 6. FEUILLE CALCULS_PAIE =====
    print("💰 Création feuille Calculs_Paie...")
    ws_calculs = wb.create_sheet("Calculs_Paie")

    headers_calculs = ["Guide", "Nombre_Visites", "Total_Heures", "Montant_Salaire"]
    ws_calculs.append(headers_calculs)

    # Style en-têtes
    for cell in ws_calculs[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    ws_calculs.column_dimensions['A'].width = 20
    ws_calculs.column_dimensions['B'].width = 18
    ws_calculs.column_dimensions['C'].width = 15
    ws_calculs.column_dimensions['D'].width = 18

    # ===== 7. FEUILLE CONTRATS =====
    print("📄 Création feuille Contrats...")
    ws_contrats = wb.create_sheet("Contrats")

    headers_contrats = ["Guide", "Mois", "Dates_Visites", "Horaires", "Total_Heures", "Montant"]
    ws_contrats.append(headers_contrats)

    # Style en-têtes
    for cell in ws_contrats[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_contrats.column_dimensions[col].width = 20

    # ===== 8. FEUILLE CONFIGURATION =====
    print("⚙️ Création feuille Configuration...")
    ws_config = wb.create_sheet("Configuration")

    headers_config = ["Paramètre", "Valeur"]
    ws_config.append(headers_config)

    config_data = [
        ["Email_Expediteur", "admin@musee.fr"],
        ["Nom_Association", "Musée des Guides"],
        ["Tarif_Heure", "30"],
        ["Notification_J7", "OUI"],
        ["Notification_J1", "OUI"],
        ["Delai_Notification_1", "7"],
        ["Delai_Notification_2", "1"],
        ["MotDePasseAdmin", "admin123"],
    ]

    for config in config_data:
        ws_config.append(config)

    # Style en-têtes
    for cell in ws_config[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    ws_config.column_dimensions['A'].width = 25
    ws_config.column_dimensions['B'].width = 30

    # ===== SAUVEGARDE =====
    filename = "PLANNING_MUSEE_ENRICHI.xlsx"
    wb.save(filename)
    print(f"\n✅ Fichier créé : {filename}")

    # ===== STATISTIQUES =====
    print("\n" + "="*70)
    print("📊 STATISTIQUES DU FICHIER GÉNÉRÉ")
    print("="*70)
    print(f"👥 Guides : {len(guides_data)} guides")
    print(f"📅 Disponibilités : {len(guides_data) * 30} lignes")
    print(f"🎫 Visites : 40 visites")
    print(f"   - 5 musées différents")
    print(f"   - 6 types de visites")
    print(f"   - Période : 01/11/2025 → 30/11/2025")
    print(f"📋 Planning : 5 attributions pré-remplies")
    print(f"⚙️ Configuration : 8 paramètres")
    print("\n💡 Pour compléter le planning :")
    print("   → Ouvrir Excel VBA et exécuter : GenererPlanningAutomatique()")
    print("="*70)

    return filename

if __name__ == "__main__":
    try:
        filename = creer_excel_enrichi()
        print(f"\n🎉 Génération terminée avec succès !")
        print(f"📁 Fichier : {filename}")
    except Exception as e:
        print(f"\n❌ Erreur lors de la génération : {e}")
        import traceback
        traceback.print_exc()

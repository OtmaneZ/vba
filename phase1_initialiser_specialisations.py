#!/usr/bin/env python3
"""
Phase 1 - Initialisation onglet Specialisations
Crée la structure pour gérer les spécialisations des guides
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

print("=" * 100)
print("🔧 PHASE 1 - INITIALISATION ONGLET SPECIALISATIONS")
print("=" * 100)

# ============================================================================
# ÉTAPE 1 : CHARGER LE FICHIER
# ============================================================================
print("\n📂 Étape 1/5 : Chargement PLANNING.xlsm...")
wb = load_workbook('PLANNING.xlsm', keep_vba=True)
ws_spec = wb['Spécialisations']
ws_guides = wb['Guides']

print("   ✅ Fichiers chargés")

# ============================================================================
# ÉTAPE 2 : EFFACER ET RECRÉER LA STRUCTURE
# ============================================================================
print("\n🗑️  Étape 2/5 : Recréation structure Specialisations...")

# Effacer tout le contenu
for row in range(ws_spec.max_row, 0, -1):
    ws_spec.delete_rows(row)

# Nouveaux en-têtes
entetes = [
    'ID_Specialisation',   # A - S0001, S0002...
    'Prenom_Guide',        # B - Prénom du guide
    'Nom_Guide',           # C - Nom du guide
    'Type_Prestation',     # D - Type de prestation (liste)
    'Autorise'             # E - OUI/NON
]

# Écrire les en-têtes avec style
for col, header in enumerate(entetes, 1):
    cell = ws_spec.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

print(f"   ✅ {len(entetes)} colonnes créées")

# ============================================================================
# ÉTAPE 3 : LIRE LES GUIDES
# ============================================================================
print("\n👥 Étape 3/5 : Lecture des guides...")

guides = []
for row in range(2, ws_guides.max_row + 1):
    prenom = ws_guides.cell(row, 2).value  # Colonne B
    nom = ws_guides.cell(row, 3).value     # Colonne C

    if prenom and nom:
        guides.append({
            'prenom': str(prenom).strip(),
            'nom': str(nom).strip()
        })

print(f"   ✅ {len(guides)} guides trouvés")
for g in guides:
    print(f"      - {g['prenom']} {g['nom']}")

# ============================================================================
# ÉTAPE 4 : GÉNÉRER LES SPÉCIALISATIONS
# ============================================================================
print("\n⚙️  Étape 4/5 : Génération spécialisations par défaut...")

# 5 types de prestations
types_prestations = [
    "VISITE CONTEE BRANLY",
    "VISITE CONTEE MARINE",
    "HORS LES MURS",
    "VISIO",
    "EVENEMENT BRANLY"
]

next_row = 2
spec_id = 1

# Pour chaque guide, créer 5 lignes (1 par type de prestation)
for guide in guides:
    for type_presta in types_prestations:
        # Colonne A: ID_Specialisation
        ws_spec.cell(next_row, 1).value = f"S{spec_id:04d}"

        # Colonne B: Prenom_Guide
        ws_spec.cell(next_row, 2).value = guide['prenom']

        # Colonne C: Nom_Guide
        ws_spec.cell(next_row, 3).value = guide['nom']

        # Colonne D: Type_Prestation
        ws_spec.cell(next_row, 4).value = type_presta

        # Colonne E: Autorise (par défaut OUI = tous les guides font tout)
        ws_spec.cell(next_row, 5).value = "OUI"

        spec_id += 1
        next_row += 1

print(f"   ✅ {spec_id - 1} lignes de spécialisations générées")
print(f"      ({len(guides)} guides × {len(types_prestations)} types)")

# ============================================================================
# ÉTAPE 5 : AJOUTER VALIDATIONS
# ============================================================================
print("\n✨ Étape 5/5 : Ajout validations...")

# Validation colonne D: Type_Prestation (liste déroulante)
types_presta_str = '"VISITE CONTEE BRANLY,VISITE CONTEE MARINE,HORS LES MURS,VISIO,EVENEMENT BRANLY"'
dv_type = DataValidation(type="list", formula1=types_presta_str, allow_blank=False)
dv_type.error = 'Valeur invalide'
dv_type.errorTitle = 'Type de prestation'
ws_spec.add_data_validation(dv_type)
dv_type.add(f'D2:D1000')
print("   ✅ Liste déroulante Type_Prestation (colonne D)")

# Validation colonne E: Autorise (OUI/NON)
dv_autorise = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=False)
dv_autorise.error = 'Valeur invalide (OUI ou NON)'
dv_autorise.errorTitle = 'Autorisé'
ws_spec.add_data_validation(dv_autorise)
dv_autorise.add(f'E2:E1000')
print("   ✅ Liste déroulante Autorisé (colonne E)")

# Ajuster largeurs colonnes
largeurs = {
    1: 18,  # A: ID_Specialisation
    2: 15,  # B: Prenom_Guide
    3: 15,  # C: Nom_Guide
    4: 25,  # D: Type_Prestation
    5: 10,  # E: Autorise
}

for col, width in largeurs.items():
    ws_spec.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

print("   ✅ Largeurs colonnes ajustées")

# Geler la première ligne
ws_spec.freeze_panes = 'A2'
print("   ✅ Première ligne gelée")

# ============================================================================
# SAUVEGARDER
# ============================================================================
print("\n💾 Sauvegarde PLANNING.xlsm...")

wb.save('PLANNING.xlsm')
wb.close()

print("   ✅ Fichier sauvegardé")

# ============================================================================
# RÉSUMÉ
# ============================================================================
print("\n" + "=" * 100)
print("✅ PHASE 1 - ÉTAPE 2 TERMINÉE : SPÉCIALISATIONS INITIALISÉES")
print("=" * 100)

print(f"""
📊 RÉSUMÉ:

Structure créée:
  ✅ A: ID_Specialisation (S0001, S0002...)
  ✅ B: Prenom_Guide
  ✅ C: Nom_Guide
  ✅ D: Type_Prestation (liste déroulante)
  ✅ E: Autorise (OUI/NON)

Données générées:
  ✅ {len(guides)} guides configurés
  ✅ {spec_id - 1} lignes de spécialisations créées
  ✅ Par défaut: TOUS les guides font TOUTES les visites (Autorise = OUI)

💡 MODIFICATION PAR LA CLIENTE:
  Pour restreindre un guide:
  1. Ouvrir onglet Specialisations
  2. Trouver la ligne Guide + Type de prestation
  3. Changer "OUI" en "NON" dans colonne E

  Exemple: Si Marie Dupont ne fait PAS de "HORS LES MURS"
  → Trouver ligne: Marie | Dupont | HORS LES MURS | OUI
  → Changer en: Marie | Dupont | HORS LES MURS | NON

⚠️  Le système utilisera ces spécialisations pour filtrer automatiquement
    les guides lors de la génération du planning.
""")

print("\n" + "=" * 100)
print("✅ PHASE 1 COMPLÈTE - RESTRUCTURATION EXCEL TERMINÉE")
print("=" * 100)

print("""
📝 PROCHAINES ÉTAPES:

1. ✅ Tests manuels:
   - Ouvrir PLANNING.xlsm
   - Vérifier onglet Visites (structure, listes déroulantes)
   - Vérifier onglet Specialisations (guides, types)

2. 🔄 Commit Git:
   git add -A
   git commit -m "Phase 1: Restructuration Visites + Spécialisations OK"
   git push

3. ⚙️  Phase 2 - Adapter macros VBA:
   - Module_Calculs.bas
   - Module_Planning.bas
   - Module_Emails.bas
""")

print("=" * 100)

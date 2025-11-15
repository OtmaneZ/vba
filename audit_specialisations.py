#!/usr/bin/env python3
"""
AUDIT COMPLET: Pourquoi la feuille Spécialisations n'apparaît pas pour l'admin
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Chemin du fichier Excel
fichier_planning = Path(__file__).parent / "PLANNING.xlsm"

if not fichier_planning.exists():
    print(f"❌ ERREUR: Fichier {fichier_planning} introuvable")
    sys.exit(1)

print("="*80)
print("🔍 AUDIT COMPLET - Feuille Spécialisations")
print("="*80)

problemes = []

try:
    # Charger le workbook avec support des macros
    print(f"\n📂 Chargement de {fichier_planning.name}...")
    wb = load_workbook(fichier_planning, keep_vba=True)

    print(f"✅ Fichier chargé avec succès")
    print(f"   Type: {type(wb)}")

    # 1. LISTER TOUTES LES FEUILLES
    print("\n" + "="*80)
    print("📋 1. LISTE DE TOUTES LES FEUILLES DU CLASSEUR")
    print("="*80)

    feuilles_trouvees = []
    feuille_specialisations_existe = False

    for idx, ws in enumerate(wb.worksheets, 1):
        nom = ws.title
        visible = "VISIBLE" if ws.sheet_state == "visible" else f"MASQUÉE ({ws.sheet_state})"
        nb_lignes = ws.max_row
        nb_cols = ws.max_column

        feuilles_trouvees.append(nom)

        print(f"   [{idx}] {nom}")
        print(f"       État: {visible}")
        print(f"       Dimensions: {nb_lignes} lignes × {nb_cols} colonnes")

        # Chercher la feuille Spécialisations (avec ou sans accent)
        if nom.lower().replace('é', 'e') == "specialisations":
            feuille_specialisations_existe = True
            print(f"       🎯 TROUVÉE ! État actuel: {ws.sheet_state}")

            if ws.sheet_state != "visible":
                problemes.append(f"❌ PROBLÈME 1: La feuille '{nom}' existe mais est en état '{ws.sheet_state}' au lieu de 'visible'")

            # Vérifier le contenu
            if nb_lignes <= 1:
                problemes.append(f"⚠️ PROBLÈME 2: La feuille '{nom}' semble vide (seulement {nb_lignes} ligne(s))")
            else:
                print(f"       Contenu: {nb_lignes - 1} lignes de données")

                # Afficher les en-têtes
                headers = []
                for col in range(1, min(nb_cols + 1, 10)):
                    val = ws.cell(1, col).value
                    if val:
                        headers.append(str(val))

                if headers:
                    print(f"       En-têtes: {', '.join(headers)}")

        print()

    if not feuille_specialisations_existe:
        problemes.append("❌ PROBLÈME MAJEUR: La feuille 'Spécialisations' N'EXISTE PAS dans le fichier Excel!")

    # 2. VÉRIFIER LES NOMS EXACTS
    print("="*80)
    print("📝 2. VÉRIFICATION DES NOMS EXACTS (sensible à la casse et aux accents)")
    print("="*80)

    variations_possibles = [
        "Spécialisations",
        "Specialisations",
        "Spécialisation",
        "Specialisation",
        "SPECIALISATIONS",
        "SPÉCIALISATIONS"
    ]

    for variation in variations_possibles:
        if variation in wb.sheetnames:
            print(f"   ✅ Trouvé: '{variation}'")
        else:
            print(f"   ❌ Non trouvé: '{variation}'")

    # 3. VÉRIFIER LES FEUILLES MASQUÉES
    print("\n" + "="*80)
    print("🙈 3. FEUILLES MASQUÉES (hidden/veryHidden)")
    print("="*80)

    feuilles_masquees = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            feuilles_masquees.append(f"{ws.title} (état: {ws.sheet_state})")
            print(f"   🔒 {ws.title}: {ws.sheet_state}")

    if not feuilles_masquees:
        print("   ✅ Aucune feuille masquée")
    else:
        problemes.append(f"⚠️ PROBLÈME 3: {len(feuilles_masquees)} feuille(s) masquée(s): {', '.join(feuilles_masquees)}")

    # 4. VÉRIFIER LES PROPRIÉTÉS DU WORKBOOK
    print("\n" + "="*80)
    print("⚙️  4. PROPRIÉTÉS DU CLASSEUR")
    print("="*80)

    print(f"   Nombre total de feuilles: {len(wb.worksheets)}")
    print(f"   Feuille active: {wb.active.title if wb.active else 'Aucune'}")
    print(f"   Support VBA: {hasattr(wb, 'vba_archive')}")

    if hasattr(wb, 'vba_archive') and wb.vba_archive:
        print(f"   Archive VBA présente: ✅")
    else:
        problemes.append("⚠️ PROBLÈME 4: Archive VBA manquante ou non détectée")

    # 5. CHERCHER DANS LES NOMS DÉFINIS
    print("\n" + "="*80)
    print("🏷️  5. NOMS DÉFINIS (Named Ranges)")
    print("="*80)

    if wb.defined_names:
        print(f"   Nombre de noms définis: {len(wb.defined_names)}")
        for name in list(wb.defined_names.definedName)[:10]:
            print(f"   - {name.name}: {name.value}")
    else:
        print("   Aucun nom défini trouvé")

    # 6. VÉRIFIER LA STRUCTURE DES FEUILLES ATTENDUES
    print("\n" + "="*80)
    print("📊 6. VÉRIFICATION DES FEUILLES ATTENDUES PAR LE SYSTÈME")
    print("="*80)

    feuilles_attendues = [
        "Accueil",
        "Guides",
        "Disponibilites",
        "Visites",
        "Planning",
        "Calculs_Paie",
        "Contrats",
        "Configuration",
        "Spécialisations"
    ]

    for feuille in feuilles_attendues:
        if feuille in wb.sheetnames:
            ws = wb[feuille]
            print(f"   ✅ {feuille}: {ws.sheet_state}, {ws.max_row} lignes")
        else:
            print(f"   ❌ {feuille}: MANQUANTE")
            if feuille == "Spécialisations":
                problemes.append(f"❌ PROBLÈME 5: La feuille '{feuille}' est absente du fichier Excel")

    # 7. ANALYSE DU CODE VBA (si possible)
    print("\n" + "="*80)
    print("💻 7. ANALYSE DU CODE VBA")
    print("="*80)

    if hasattr(wb, 'vba_archive') and wb.vba_archive:
        try:
            import zipfile
            import io

            # Le vba_archive est un objet BytesIO
            vba_zip = zipfile.ZipFile(io.BytesIO(wb.vba_archive))

            print(f"   Fichiers VBA détectés:")
            for name in vba_zip.namelist():
                if 'vba' in name.lower() or '.bin' in name.lower():
                    print(f"   - {name}")

            # Chercher les modules
            modules_vba = [n for n in vba_zip.namelist() if 'Module' in n or 'ThisWorkbook' in n]
            if modules_vba:
                print(f"\n   ✅ {len(modules_vba)} module(s) VBA trouvé(s)")
            else:
                problemes.append("⚠️ PROBLÈME 6: Aucun module VBA détecté dans l'archive")

        except Exception as e:
            print(f"   ⚠️ Impossible d'analyser l'archive VBA: {e}")
            problemes.append(f"⚠️ PROBLÈME 7: Erreur lors de l'analyse VBA: {e}")
    else:
        print("   ❌ Pas d'archive VBA disponible")
        problemes.append("❌ PROBLÈME 8: Archive VBA non trouvée dans le fichier")

    wb.close()

    # RÉSUMÉ DES PROBLÈMES
    print("\n" + "="*80)
    print("🚨 RÉSUMÉ DES PROBLÈMES DÉTECTÉS")
    print("="*80)

    if problemes:
        for i, probleme in enumerate(problemes, 1):
            print(f"\n{i}. {probleme}")
    else:
        print("\n✅ Aucun problème structurel détecté dans le fichier Excel")
        print("\n⚠️  Le problème pourrait venir:")
        print("   - Du code VBA qui ne s'exécute pas correctement")
        print("   - D'une erreur dans la logique de connexion admin")
        print("   - D'un problème de rafraîchissement après modification VBA")

    # RECOMMANDATIONS
    print("\n" + "="*80)
    print("💡 RECOMMANDATIONS")
    print("="*80)

    if not feuille_specialisations_existe:
        print("\n🔴 ACTION URGENTE:")
        print("   La feuille 'Spécialisations' n'existe pas dans le fichier !")
        print("   Vous devez la créer ou la restaurer depuis une sauvegarde.")
    else:
        ws_spec = None
        for ws in wb.worksheets:
            if ws.title.lower().replace('é', 'e') == "specialisations":
                ws_spec = ws
                break

        if ws_spec and ws_spec.sheet_state != "visible":
            print("\n🟡 ACTION IMMÉDIATE:")
            print(f"   La feuille '{ws_spec.title}' est en état '{ws_spec.sheet_state}'")
            print("   Vous devez la rendre visible manuellement ou exécuter le script de correction.")
        else:
            print("\n🟢 FEUILLE OK:")
            print("   La feuille Spécialisations existe et est visible dans le fichier.")
            print("   Le problème vient probablement du code VBA.")
            print("\n   Actions à vérifier:")
            print("   1. Le code VBA a-t-il bien été sauvegardé ?")
            print("   2. Avez-vous fermé et rouvert le fichier Excel ?")
            print("   3. Y a-t-il des erreurs VBA au démarrage ?")
            print("   4. La macro de démarrage (Workbook_Open) s'exécute-t-elle ?")

    print("\n" + "="*80)

except Exception as e:
    print(f"\n❌ ERREUR CRITIQUE: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n✅ Audit terminé")
print("="*80)

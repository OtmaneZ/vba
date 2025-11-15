#!/usr/bin/env python3
"""
Script pour INVERSER les colonnes H et I dans Planning
H = Thème (au lieu de Niveau)
I = Niveau (au lieu de Thème)
"""

import openpyxl
import sys

def main():
    fichier = 'PLANNING.xlsm'

    try:
        print(f"📂 Ouverture de {fichier}...")
        wb = openpyxl.load_workbook(fichier, keep_vba=True)

        ws_planning = wb['Planning']
        ws_visites = wb['Visites']

        print("\n=== STRUCTURE ACTUELLE (Planning) ===")
        print(f"  Col H (8): {ws_planning.cell(1, 8).value}")
        print(f"  Col I (9): {ws_planning.cell(1, 9).value}")

        print("\n🔄 INVERSION des en-têtes H et I...")

        # Inverser les en-têtes
        ws_planning.cell(1, 8).value = "Thème"   # H = Thème (au lieu de Niveau)
        ws_planning.cell(1, 9).value = "Niveau"  # I = Niveau (au lieu de Thème)

        print("✅ En-têtes inversés")

        # Copier les données depuis Visites (en inversant l'ordre)
        print("\n📋 Copie des données depuis Visites (avec inversion)...")

        max_row_planning = ws_planning.max_row
        compteur = 0

        for row in range(2, max_row_planning + 1):
            id_visite = ws_planning.cell(row, 1).value

            if not id_visite:
                continue

            # Chercher la visite correspondante dans Visites
            for v_row in range(2, ws_visites.max_row + 1):
                v_id = ws_visites.cell(v_row, 1).value

                if v_id == id_visite:
                    # Dans Visites: col 8 = Niveau, col 9 = Thème
                    niveau_visites = ws_visites.cell(v_row, 8).value
                    theme_visites = ws_visites.cell(v_row, 9).value

                    # Dans Planning: col 8 = Thème, col 9 = Niveau (INVERSÉ)
                    ws_planning.cell(row, 8).value = theme_visites   # H = Thème
                    ws_planning.cell(row, 9).value = niveau_visites  # I = Niveau

                    compteur += 1
                    break

        print(f"✅ {compteur} lignes mises à jour")

        print("\n=== NOUVELLE STRUCTURE (Planning) ===")
        print(f"  Col H (8): {ws_planning.cell(1, 8).value}")
        print(f"  Col I (9): {ws_planning.cell(1, 9).value}")

        # Exemple de données
        if compteur > 0:
            print("\n=== EXEMPLE (ligne 2) ===")
            print(f"  Thème (H): {ws_planning.cell(2, 8).value}")
            print(f"  Niveau (I): {ws_planning.cell(2, 9).value}")

        # Sauvegarder
        backup = 'PLANNING_backup_avant_inversion.xlsm'
        print(f"\n💾 Sauvegarde : {backup}")
        wb.save(backup)

        print(f"💾 Sauvegarde du fichier corrigé : {fichier}")
        wb.save(fichier)

        wb.close()

        print("\n" + "="*60)
        print("✅ INVERSION TERMINÉE AVEC SUCCÈS")
        print("="*60)
        print("\n📌 NOUVELLE STRUCTURE :")
        print("   Col H = Thème (comme demandé)")
        print("   Col I = Niveau (comme demandé)")
        print("\n⚠️  Les modules VBA sont déjà corrects (col 8=Thème, col 9=Niveau dans le code)")

    except Exception as e:
        print(f"\n❌ ERREUR : {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'analyse des différences entre la structure Excel actuelle
et la structure attendue par le code VBA
"""

import openpyxl
from pathlib import Path
import sys

# Codes couleurs pour terminal
class Couleurs:
    HEADER = '\033[95m'
    BLEU = '\033[94m'
    CYAN = '\033[96m'
    VERT = '\033[92m'
    JAUNE = '\033[93m'
    ROUGE = '\033[91m'
    FIN = '\033[0m'
    GRAS = '\033[1m'
    SOULIGNE = '\033[4m'

def print_titre(texte):
    print(f"\n{Couleurs.GRAS}{Couleurs.BLEU}{'='*70}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.BLEU}{texte:^70}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.BLEU}{'='*70}{Couleurs.FIN}\n")

def print_section(texte):
    print(f"\n{Couleurs.GRAS}{Couleurs.CYAN}{'─'*70}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.CYAN}{texte}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.CYAN}{'─'*70}{Couleurs.FIN}")

def print_succes(texte):
    print(f"{Couleurs.VERT}✓ {texte}{Couleurs.FIN}")

def print_warning(texte):
    print(f"{Couleurs.JAUNE}⚠ {texte}{Couleurs.FIN}")

def print_erreur(texte):
    print(f"{Couleurs.ROUGE}✗ {texte}{Couleurs.FIN}")

def print_info(texte):
    print(f"  {texte}")

# Structure ATTENDUE par le code VBA (d'après Module_Config.bas)
STRUCTURE_VBA = {
    "Guides": {
        "colonnes": ["Prenom", "Nom", "Email", "Telephone", "Specialisations", "Mot_De_Passe", "Statut"],
        "description": "Base de données des guides avec leurs infos et spécialités"
    },
    "Disponibilites": {
        "colonnes": ["Prenom", "Nom", "Date", "Disponible"],
        "description": "Disponibilités des guides par date"
    },
    "Visites": {
        "colonnes": ["Date", "Heure_Debut", "Heure_Fin", "Type_Visite", "Musee", "Langue",
                     "Nb_Personnes", "Tarif", "Guide_Attribue", "Statut", "Notes"],
        "description": "Catalogue de toutes les visites à planifier"
    },
    "Planning": {
        "colonnes": ["Date", "Heure_Debut", "Heure_Fin", "Musee", "Type_Visite", "Langue",
                     "Nb_Personnes", "Guide_Attribue", "Statut"],
        "description": "Planning effectif avec guides attribués"
    },
    "Calculs_Paie": {
        "colonnes": ["Prenom", "Nom", "Nb_Visites", "Nb_Heures", "Total_Brut",
                     "Montant_Par_Cachet", "Nb_Cachets", "Total_Recalcule", "Mois"],
        "description": "Calculs de paie avec système de cachets"
    },
    "Contrats": {
        "colonnes": ["Prenom", "Nom", "Date_Generation", "Periode", "Type_Contrat",
                     "Nb_Visites", "Nb_Cachets", "Montant_Cachet", "Total", "Statut"],
        "description": "Contrats générés (début et fin de mois)"
    },
    "Configuration": {
        "colonnes": ["Parametre", "Valeur", "Description"],
        "description": "Paramètres tarifaires et configuration système"
    },
    "Accueil": {
        "colonnes": [],
        "description": "Interface utilisateur (pas de structure fixe)"
    },
    "Mon_Planning": {
        "colonnes": ["Date", "Heure_Debut", "Musee", "Type_Visite", "Langue",
                     "Nb_Personnes", "Statut", "Action"],
        "description": "Vue personnalisée pour chaque guide (NOUVELLE FEUILLE)"
    }
}

def normaliser_nom_colonne(nom):
    """Normalise un nom de colonne pour comparaison"""
    if nom is None:
        return ""
    return str(nom).strip().lower().replace('é', 'e').replace('è', 'e').replace('_', '').replace(' ', '')

def comparer_feuille(nom_feuille, colonnes_excel, structure_vba):
    """Compare une feuille Excel avec la structure VBA attendue"""

    print_section(f"📋 Feuille : {nom_feuille}")

    if nom_feuille not in structure_vba:
        print_warning(f"Cette feuille n'est pas utilisée par le code VBA")
        print_info(f"Colonnes actuelles : {colonnes_excel}")
        return {"status": "extra", "details": "Feuille non utilisée par VBA"}

    vba_info = structure_vba[nom_feuille]
    colonnes_vba = vba_info["colonnes"]

    print_info(f"📖 Description : {vba_info['description']}")

    if not colonnes_vba:  # Feuille sans structure fixe (ex: Accueil)
        print_succes("Feuille d'interface - pas de structure fixe requise")
        return {"status": "ok", "details": "Interface"}

    # Normaliser les colonnes pour comparaison
    colonnes_excel_norm = [normaliser_nom_colonne(c) for c in colonnes_excel if c]
    colonnes_vba_norm = [normaliser_nom_colonne(c) for c in colonnes_vba]

    # Analyser les différences
    manquantes = []
    extras = []
    ordre_different = False

    for col_vba in colonnes_vba:
        col_norm = normaliser_nom_colonne(col_vba)
        if col_norm not in colonnes_excel_norm:
            manquantes.append(col_vba)

    for i, col_excel in enumerate(colonnes_excel):
        if col_excel:  # Ignorer les colonnes vides
            col_norm = normaliser_nom_colonne(col_excel)
            if col_norm not in colonnes_vba_norm:
                extras.append(col_excel)

    # Vérifier l'ordre (seulement pour colonnes communes)
    if not manquantes and not extras:
        for i, col_vba in enumerate(colonnes_vba):
            if i < len(colonnes_excel):
                if normaliser_nom_colonne(colonnes_excel[i]) != normaliser_nom_colonne(col_vba):
                    ordre_different = True
                    break

    # Afficher les résultats
    print(f"\n{Couleurs.GRAS}Structure actuelle :{Couleurs.FIN}")
    print_info(f"{colonnes_excel}")

    print(f"\n{Couleurs.GRAS}Structure attendue par VBA :{Couleurs.FIN}")
    print_info(f"{colonnes_vba}")

    if not manquantes and not extras and not ordre_different:
        print(f"\n{Couleurs.VERT}{Couleurs.GRAS}✓ STRUCTURE PARFAITE{Couleurs.FIN}")
        return {"status": "ok", "details": "Parfait"}

    problemes = []

    if manquantes:
        print(f"\n{Couleurs.ROUGE}{Couleurs.GRAS}Colonnes MANQUANTES :{Couleurs.FIN}")
        for col in manquantes:
            print_erreur(f"Manque : {col}")
            problemes.append(f"Manque: {col}")

    if extras:
        print(f"\n{Couleurs.JAUNE}{Couleurs.GRAS}Colonnes SUPPLEMENTAIRES :{Couleurs.FIN}")
        for col in extras:
            print_warning(f"En plus : {col}")
            problemes.append(f"Extra: {col}")

    if ordre_different:
        print(f"\n{Couleurs.JAUNE}{Couleurs.GRAS}⚠ Ordre des colonnes différent{Couleurs.FIN}")
        print_info("Les colonnes existent mais pas dans le bon ordre")
        problemes.append("Ordre différent")

    # Proposer des actions
    print(f"\n{Couleurs.CYAN}{Couleurs.GRAS}💡 Actions recommandées :{Couleurs.FIN}")
    if manquantes:
        print_info(f"→ Ajouter {len(manquantes)} colonne(s) manquante(s)")
    if extras and not manquantes:
        print_info(f"→ Option 1 : Garder les colonnes extras (pas de problème)")
        print_info(f"→ Option 2 : Supprimer/renommer les colonnes extras")
    if ordre_different:
        print_info(f"→ Réorganiser l'ordre des colonnes")

    return {
        "status": "problemes" if manquantes else "warnings",
        "manquantes": manquantes,
        "extras": extras,
        "ordre_different": ordre_different,
        "details": ", ".join(problemes)
    }

def analyser_excel(fichier_path):
    """Analyse complète du fichier Excel"""

    print_titre("ANALYSE DÉTAILLÉE : EXCEL vs CODE VBA")

    # Ouvrir le fichier
    try:
        wb = openpyxl.load_workbook(fichier_path, keep_vba=True)
    except Exception as e:
        print_erreur(f"Impossible d'ouvrir le fichier : {e}")
        return

    print_info(f"📁 Fichier : {fichier_path}")
    print_info(f"📊 Nombre de feuilles : {len(wb.sheetnames)}")

    # Analyser chaque feuille
    resultats = {}

    for nom_feuille in wb.sheetnames:
        ws = wb[nom_feuille]
        colonnes = [cell.value for cell in ws[1]]
        resultats[nom_feuille] = comparer_feuille(nom_feuille, colonnes, STRUCTURE_VBA)

    # Vérifier les feuilles manquantes
    print_section("🔍 Feuilles manquantes dans Excel")
    feuilles_manquantes = []
    for nom_feuille in STRUCTURE_VBA.keys():
        if nom_feuille not in wb.sheetnames:
            feuilles_manquantes.append(nom_feuille)
            print_erreur(f"Feuille '{nom_feuille}' n'existe pas encore")
            print_info(f"   Description : {STRUCTURE_VBA[nom_feuille]['description']}")
            print_info(f"   Colonnes requises : {STRUCTURE_VBA[nom_feuille]['colonnes']}")

    if not feuilles_manquantes:
        print_succes("Toutes les feuilles VBA existent dans Excel")

    # RÉSUMÉ GÉNÉRAL
    print_titre("RÉSUMÉ DE L'ANALYSE")

    nb_ok = sum(1 for r in resultats.values() if r["status"] == "ok")
    nb_warnings = sum(1 for r in resultats.values() if r["status"] == "warnings")
    nb_problemes = sum(1 for r in resultats.values() if r["status"] == "problemes")
    nb_extra = sum(1 for r in resultats.values() if r["status"] == "extra")

    print(f"\n{Couleurs.GRAS}Statistiques :{Couleurs.FIN}")
    print_succes(f"{nb_ok} feuille(s) parfaite(s)")
    if nb_warnings > 0:
        print_warning(f"{nb_warnings} feuille(s) avec avertissements")
    if nb_problemes > 0:
        print_erreur(f"{nb_problemes} feuille(s) avec problèmes critiques")
    if nb_extra > 0:
        print_info(f"{nb_extra} feuille(s) non utilisée(s) par VBA")
    if feuilles_manquantes:
        print_erreur(f"{len(feuilles_manquantes)} feuille(s) manquante(s)")

    # PLAN D'ACTION
    print_titre("PLAN D'ACTION RECOMMANDÉ")

    actions_critiques = []
    actions_recommandees = []

    # Feuilles manquantes
    if feuilles_manquantes:
        actions_critiques.append(f"1️⃣  CRÉER les feuilles manquantes : {', '.join(feuilles_manquantes)}")

    # Feuilles avec colonnes manquantes
    for nom, result in resultats.items():
        if result["status"] == "problemes" and "manquantes" in result:
            cols = result["manquantes"]
            actions_critiques.append(f"2️⃣  AJOUTER colonnes dans '{nom}' : {', '.join(cols)}")

    # Feuilles avec ordre différent
    for nom, result in resultats.items():
        if "ordre_different" in result and result["ordre_different"]:
            actions_recommandees.append(f"📝 Réorganiser les colonnes de '{nom}'")

    # Feuilles avec colonnes extras
    for nom, result in resultats.items():
        if "extras" in result and result["extras"]:
            actions_recommandees.append(f"🔧 Option : Renommer/adapter colonnes de '{nom}': {', '.join(result['extras'])}")

    if actions_critiques:
        print(f"\n{Couleurs.ROUGE}{Couleurs.GRAS}🚨 ACTIONS CRITIQUES (requis pour que le code fonctionne) :{Couleurs.FIN}")
        for action in actions_critiques:
            print(f"  {action}")
    else:
        print(f"\n{Couleurs.VERT}{Couleurs.GRAS}✓ Aucune action critique requise{Couleurs.FIN}")

    if actions_recommandees:
        print(f"\n{Couleurs.JAUNE}{Couleurs.GRAS}💡 ACTIONS RECOMMANDÉES (pour optimisation) :{Couleurs.FIN}")
        for action in actions_recommandees:
            print(f"  {action}")

    # Options pour l'utilisateur
    print_titre("OPTIONS DE MISE À JOUR")

    print(f"{Couleurs.GRAS}Vous avez 3 options :{Couleurs.FIN}\n")

    print(f"{Couleurs.VERT}{Couleurs.GRAS}Option 1 : Mise à jour AUTOMATIQUE{Couleurs.FIN}")
    print_info("✓ Script Python ajoute colonnes manquantes et crée feuilles")
    print_info("✓ Vos données actuelles sont PRÉSERVÉES")
    print_info("✓ Colonnes ajoutées à la fin (pas de réorganisation)")
    print_info("✗ L'ordre peut rester différent de l'attendu")
    print_info(f"→ Commande : {Couleurs.CYAN}python3 preparer_excel.py PLANNING_MUSEE_FINAL_PROPRE.xlsm --mode=ajout{Couleurs.FIN}")

    print(f"\n{Couleurs.JAUNE}{Couleurs.GRAS}Option 2 : Mise à jour MANUELLE{Couleurs.FIN}")
    print_info("✓ Vous gardez le contrôle total")
    print_info("✓ Vous pouvez adapter selon vos besoins")
    print_info("→ Utilisez ce rapport comme guide")

    print(f"\n{Couleurs.BLEU}{Couleurs.GRAS}Option 3 : Adapter le CODE VBA{Couleurs.FIN}")
    print_info("✓ Garder votre structure Excel actuelle")
    print_info("✓ Modifier Module_Config.bas pour matcher")
    print_info("✗ Nécessite de refaire l'audit des modules")
    print_info("→ Plus complexe, mais Excel reste inchangé")

    print(f"\n{Couleurs.CYAN}{Couleurs.GRAS}💬 Recommandation :{Couleurs.FIN}")
    if len(actions_critiques) > 2:
        print_info("Option 1 recommandée - Beaucoup de changements nécessaires")
    elif actions_critiques:
        print_info("Option 1 ou 2 - Quelques ajustements nécessaires")
    else:
        print_info("Option 2 - Juste quelques ajustements cosmétiques")

    print("\n")

def main():
    fichier = "PLANNING_MUSEE_FINAL_PROPRE.xlsm"

    if len(sys.argv) > 1:
        fichier = sys.argv[1]

    fichier_path = Path(fichier)
    if not fichier_path.is_absolute():
        fichier_path = Path.cwd() / fichier

    if not fichier_path.exists():
        print_erreur(f"Fichier non trouvé : {fichier_path}")
        sys.exit(1)

    analyser_excel(str(fichier_path))

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRIGER LA STRUCTURE DE LA FEUILLE DISPONIBILITES
Le problème : les colonnes sont mal organisées lors de l'import
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import shutil

def corriger_disponibilites():
    """
    Réorganise la feuille Disponibilites avec la bonne structure
    """

    fichier_planning = "/Users/otmaneboulahia/Documents/Excel-Auto/PLANNING.xlsm"

    print("=" * 80)
    print("🔧 CORRECTION STRUCTURE FEUILLE DISPONIBILITES")
    print("=" * 80)

    # Backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f"/Users/otmaneboulahia/Documents/Excel-Auto/PLANNING_backup_dispo_{timestamp}.xlsm"
    shutil.copy2(fichier_planning, backup_file)
    print(f"\n✅ Backup créé : {backup_file}")

    # Charger le fichier
    wb = load_workbook(fichier_planning, keep_vba=True)
    ws = wb["Disponibilites"]

    print("\n📋 STRUCTURE ACTUELLE (INCORRECTE):")
    print(f"  Col 1 (Guide) : {ws.cell(2, 1).value}")
    print(f"  Col 2 (Date) : {ws.cell(2, 2).value}")
    print(f"  Col 3 (Disponible) : {ws.cell(2, 3).value}")
    print(f"  Col 4 (Commentaire) : {ws.cell(2, 4).value}")
    print(f"  Col 5 (Prenom) : {ws.cell(2, 5).value}")
    print(f"  Col 6 (Nom) : {ws.cell(2, 6).value}")

    # Lire toutes les données
    donnees = []
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, 1).value  # Col 1 = Date
        dispo_val = ws.cell(row, 2).value  # Col 2 = OUI/NON
        commentaire_val = ws.cell(row, 3).value  # Col 3 = vide
        prenom_val = ws.cell(row, 4).value  # Col 4 = Prénom
        nom_val = ws.cell(row, 5).value  # Col 5 = Nom

        if date_val:  # Si ligne non vide
            donnees.append({
                'date': date_val,
                'disponible': dispo_val if dispo_val else "",
                'commentaire': commentaire_val if commentaire_val else "",
                'prenom': prenom_val if prenom_val else "",
                'nom': nom_val if nom_val else ""
            })

    print(f"\n✅ {len(donnees)} lignes de données lues")

    # Défusionner toutes les cellules
    print("\n🔓 Défusion des cellules...")
    merged_cells = list(ws.merged_cells.ranges)
    for merged_cell in merged_cells:
        ws.unmerge_cells(str(merged_cell))

    # Effacer toutes les données (garder en-têtes)
    for row in range(2, ws.max_row + 1):
        for col in range(1, 7):
            ws.cell(row, col).value = None    # RÉÉCRIRE avec la bonne structure
    print("\n🔄 RÉORGANISATION DES COLONNES...")

    # En-têtes corrects
    ws.cell(1, 1).value = "Date"
    ws.cell(1, 2).value = "Disponible"
    ws.cell(1, 3).value = "Commentaire"
    ws.cell(1, 4).value = "Prenom"
    ws.cell(1, 5).value = "Nom"
    ws.cell(1, 6).value = "Guide"  # Colonne calculée (Prénom + Nom)

    # Réécrire les données
    for idx, ligne in enumerate(donnees, start=2):
        ws.cell(idx, 1).value = ligne['date']
        ws.cell(idx, 2).value = ligne['disponible']
        ws.cell(idx, 3).value = ligne['commentaire']
        ws.cell(idx, 4).value = ligne['prenom']
        ws.cell(idx, 5).value = ligne['nom']
        ws.cell(idx, 6).value = f"{ligne['prenom']} {ligne['nom']}"

    # Sauvegarder
    wb.save(fichier_planning)
    wb.close()

    print("\n✅ STRUCTURE CORRIGÉE !")
    print("\n📋 NOUVELLE STRUCTURE (CORRECTE):")
    print("  Col 1: Date")
    print("  Col 2: Disponible (OUI/NON)")
    print("  Col 3: Commentaire")
    print("  Col 4: Prenom")
    print("  Col 5: Nom")
    print("  Col 6: Guide (calculé)")

    print("\n" + "=" * 80)
    print("✅ CORRECTION TERMINÉE")
    print("=" * 80)
    print("""
Le fichier PLANNING.xlsm a été corrigé.

IMPORTANT : Le code VBA doit maintenant être mis à jour pour lire :
  - Col 1 : Date
  - Col 2 : Disponible
  - Col 4 : Prenom
  - Col 5 : Nom

Les modules VBA corrigés sont disponibles dans vba-modules/
""")

if __name__ == "__main__":
    corriger_disponibilites()

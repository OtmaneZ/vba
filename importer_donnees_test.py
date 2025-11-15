#!/usr/bin/env python3
"""
Import des données de test du client dans PLANNING.xlsm
"""

import openpyxl
from datetime import datetime, time
from pathlib import Path
import shutil

def convertir_heure(heure_str):
    """Convertir une heure du format '10:30' en objet time"""
    if not heure_str or heure_str.strip() == '':
        return None
    try:
        h, m = heure_str.strip().split(':')
        return time(int(h), int(m))
    except:
        return None

def main():
    fichier = Path("PLANNING.xlsm")
    
    print("="*80)
    print("IMPORT DES DONNÉES DE TEST")
    print("="*80)
    
    # Backup
    backup = Path("PLANNING_backup_avant_test.xlsm")
    shutil.copy2(fichier, backup)
    print(f"\n✅ Backup créé : {backup.name}")
    
    # Charger le fichier
    print(f"\n📂 Chargement de {fichier.name}...")
    wb = openpyxl.load_workbook(fichier, keep_vba=True)
    
    # === DISPONIBILITÉS ===
    ws_dispo = wb['Disponibilites']
    
    # Nettoyer
    if ws_dispo.max_row > 1:
        ws_dispo.delete_rows(2, ws_dispo.max_row)
    
    dispos_data = [
        ("16/11/2025", "OUI", "", "HANAKO", "DANJO"),
        ("18/11/2025", "OUI", "", "HANAKO", "DANJO"),
        ("22/11/2025", "OUI", "", "HANAKO", "DANJO"),
        ("16/11/2025", "OUI", "", "SILVIA", "MASSEGUR"),
        ("17/11/2025", "OUI", "", "SILVIA", "MASSEGUR"),
        ("19/11/2025", "OUI", "", "SILVIA", "MASSEGUR"),
        ("22/11/2025", "OUI", "JUSQU A 15H", "SILVIA", "MASSEGUR"),
        ("16/11/2025", "OUI", "JUSQU A 15H", "SOLENE", "ARBEL"),
        ("22/11/2025", "OUI", "JUSQU A 15H", "SOLENE", "ARBEL"),
        ("18/11/2025", "OUI", "A PARTIR DE 14H", "MARIE LAURE", "SAINT BONNET"),
        ("20/11/2025", "OUI", "", "MARIE LAURE", "SAINT BONNET"),
        ("21/11/2025", "OUI", "", "MARIE LAURE", "SAINT BONNET"),
        ("22/11/2025", "OUI", "", "MARIE LAURE", "SAINT BONNET"),
        ("23/11/2025", "OUI", "", "MARIE LAURE", "SAINT BONNET"),
    ]
    
    print(f"\n📥 Import de {len(dispos_data)} disponibilités...")
    
    for idx, (date_str, dispo, commentaire, prenom, nom) in enumerate(dispos_data, start=2):
        d, m, y = date_str.split('/')
        date_obj = datetime(int(y), int(m), int(d))
        
        ws_dispo.cell(idx, 1).value = date_obj
        ws_dispo.cell(idx, 2).value = dispo
        ws_dispo.cell(idx, 3).value = commentaire if commentaire else ""
        ws_dispo.cell(idx, 4).value = prenom
        ws_dispo.cell(idx, 5).value = nom
    
    print(f"   ✅ {len(dispos_data)} disponibilités importées")
    
    # === SPÉCIALISATIONS ===
    ws_spec = None
    for ws in wb.worksheets:
        if 'special' in ws.title.lower():
            ws_spec = ws
            break
    
    if ws_spec:
        print(f"\n📋 Feuille Spécialisations : {ws_spec.title}")
        ws_spec.sheet_state = 'visible'
        
        if ws_spec.max_row > 1:
            ws_spec.delete_rows(2, ws_spec.max_row)
        
        specialisations = [
            ("HANAKO", "DANJO", "VISITE CONTEE BRANLY", "OUI"),
            ("HANAKO", "DANJO", "VISITE CONTEE MARINE", "NON"),
            ("HANAKO", "DANJO", "HORS LES MURS", "OUI"),
            ("SILVIA", "MASSEGUR", "VISITE CONTEE BRANLY", "OUI"),
            ("SILVIA", "MASSEGUR", "VISITE CONTEE MARINE", "NON"),
            ("SILVIA", "MASSEGUR", "HORS LES MURS", "OUI"),
            ("SOLENE", "ARBEL", "VISITE CONTEE BRANLY", "NON"),
            ("SOLENE", "ARBEL", "VISITE CONTEE MARINE", "OUI"),
            ("SOLENE", "ARBEL", "HORS LES MURS", "NON"),
            ("MARIE LAURE", "SAINT BONNET", "VISITE CONTEE BRANLY", "OUI"),
            ("MARIE LAURE", "SAINT BONNET", "VISITE CONTEE MARINE", "OUI"),
            ("MARIE LAURE", "SAINT BONNET", "HORS LES MURS", "OUI"),
        ]
        
        for idx, (prenom, nom, type_visite, autorise) in enumerate(specialisations, start=2):
            ws_spec.cell(idx, 1).value = prenom
            ws_spec.cell(idx, 2).value = nom
            ws_spec.cell(idx, 3).value = type_visite
            ws_spec.cell(idx, 4).value = autorise
        
        print(f"   ✅ {len(specialisations)} spécialisations importées")
    
    # === VISITES ===
    ws_visites = wb['Visites']
    
    if ws_visites.max_row > 1:
        ws_visites.delete_rows(2, ws_visites.max_row)
    
    visites_data = [
        ("16/11/2025", "10:30", "1h", "VISITE CONTEE BRANLY", "ECOLE PRIVEE SAINTE CLOTILDE", 29, "Primaire/CE2", 'G-VC "Afrique"'),
        ("16/11/2025", "10:40", "1h", "VISITE CONTEE BRANLY", "ECOLE DU CENTRE", 30, "Primaire/CP", 'G-VC "Afrique"'),
        ("16/11/2025", "13:00", "1h", "VISITE CONTEE BRANLY", "ECOLE ELEMENTAIRE", 21, "Primaire/CE2", 'G-VC "Mille et un Orients"'),
        ("17/11/2025", "10:00", "1h", "VISITE CONTEE MARINE", "INDIVIDUELS", 20, "", "BULLE"),
        ("19/11/2025", "09:45", "1h", "VISITE CONTEE MARINE", "ECOLE ELEMENTAIRE PEREIRE", 33, "CE2", "A L ABORDAGE"),
        ("20/11/2025", "13:00", "1h", "VISITE CONTEE BRANLY", "COLLEGE AIME CESAIRE", 30, "Collège/4ème", 'G-VC "Afrique"'),
        ("22/11/2025", "10:00", "1h", "VISITE CONTEE MARINE", "INDIVIDUELS", 21, "", "BULLE"),
        ("23/11/2025", "11:30", "2h", "HORS LES MURS", "", None, "", ""),
    ]
    
    print(f"\n📥 Import de {len(visites_data)} visites...")
    
    for idx, (date_str, heure_str, duree, type_visite, musee, nb_pers, niveau, theme) in enumerate(visites_data, start=2):
        ws_visites.cell(idx, 1).value = f"V{idx-1:03d}"
        
        d, m, y = date_str.split('/')
        date_obj = datetime(int(y), int(m), int(d))
        ws_visites.cell(idx, 2).value = date_obj
        
        heure_obj = convertir_heure(heure_str)
        ws_visites.cell(idx, 3).value = heure_obj
        
        ws_visites.cell(idx, 4).value = duree
        ws_visites.cell(idx, 5).value = type_visite
        ws_visites.cell(idx, 6).value = musee
        ws_visites.cell(idx, 7).value = nb_pers
        ws_visites.cell(idx, 8).value = niveau
        ws_visites.cell(idx, 9).value = theme
    
    print(f"   ✅ {len(visites_data)} visites importées")
    
    # Sauvegarder
    print(f"\n💾 Sauvegarde...")
    wb.save(fichier)
    wb.close()
    
    print("\n" + "="*80)
    print("✅ IMPORT TERMINÉ")
    print("="*80)
    print("\n🔍 MAINTENANT:")
    print("   1. Ouvre PLANNING.xlsm")
    print("   2. Connecte-toi en ADMIN")
    print("   3. Vérifie les données dans Disponibilites, Visites, Spécialisations")
    print("="*80)

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Correction des en-têtes de la feuille Spécialisations
"""

from openpyxl import load_workbook

# Ouvrir le fichier
wb = load_workbook('PLANNING.xlsm', keep_vba=True)
ws = wb['Spécialisations']

# Défusionner les cellules si nécessaire
if ws.merged_cells:
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

# Corriger les en-têtes manquants
ws['B1'] = 'Nom_Guide'
ws['C1'] = 'Email_Guide'

# Mettre en gras
from openpyxl.styles import Font
bold_font = Font(bold=True)
ws['A1'].font = bold_font
ws['B1'].font = bold_font
ws['C1'].font = bold_font
ws['D1'].font = bold_font
ws['E1'].font = bold_font

# Sauvegarder
wb.save('PLANNING.xlsm')
wb.close()

print("✅ En-têtes de Spécialisations corrigés !")
print()
print("Colonnes maintenant :")
print("  A: ID_Specialisation")
print("  B: Nom_Guide")
print("  C: Email_Guide")
print("  D: Type_Prestation")
print("  E: Autorise")

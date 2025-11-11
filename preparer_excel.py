#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de préparation du fichier Excel pour l'importation des modules VBA
Crée les feuilles manquantes, ajoute les en-têtes de colonnes, et vérifie la structure
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path

# Codes couleurs pour terminal
class Couleurs:
    HEADER = '\033[95m'
    BLEU = '\033[94m'
    CYAN = '\033[96m'
    VERT = '\033[92m'
    JAUNE = '\033[93m'
    ROUGE = '\033[91m'
    FIN = '\033[0m'
    GRAS = '\033[1m'
    SOULIGNE = '\033[4m'

def print_titre(texte):
    print(f"\n{Couleurs.GRAS}{Couleurs.BLEU}{'='*60}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.BLEU}{texte:^60}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.BLEU}{'='*60}{Couleurs.FIN}\n")

def print_succes(texte):
    print(f"{Couleurs.VERT}✓ {texte}{Couleurs.FIN}")

def print_warning(texte):
    print(f"{Couleurs.JAUNE}⚠ {texte}{Couleurs.FIN}")

def print_erreur(texte):
    print(f"{Couleurs.ROUGE}✗ {texte}{Couleurs.FIN}")

def print_info(texte):
    print(f"{Couleurs.CYAN}ℹ {texte}{Couleurs.FIN}")

# Structure attendue des feuilles avec leurs colonnes
STRUCTURE_FEUILLES = {
    "Accueil": [],  # Feuille d'interface, pas de structure fixe

    "Guides": [
        "Prenom", "Nom", "Email", "Telephone",
        "Specialisations", "Mot_De_Passe", "Statut"
    ],

    "Disponibilites": [
        "Prenom", "Nom", "Date", "Disponible"
    ],

    "Visites": [
        "Date", "Heure_Debut", "Heure_Fin", "Type_Visite",
        "Musee", "Langue", "Nb_Personnes", "Tarif",
        "Guide_Attribue", "Statut", "Notes"
    ],

    "Planning": [
        "Date", "Heure_Debut", "Heure_Fin", "Musee",
        "Type_Visite", "Langue", "Nb_Personnes",
        "Guide_Attribue", "Statut"
    ],

    "Calculs_Paie": [
        "Prenom", "Nom", "Nb_Visites", "Nb_Heures",
        "Total_Brut", "Montant_Par_Cachet", "Nb_Cachets",
        "Total_Recalcule", "Mois"
    ],

    "Contrats": [
        "Prenom", "Nom", "Date_Generation", "Periode",
        "Type_Contrat", "Nb_Visites", "Nb_Cachets",
        "Montant_Cachet", "Total", "Statut"
    ],

    "Configuration": [
        "Parametre", "Valeur", "Description"
    ],

    "Mon_Planning": [
        "Date", "Heure_Debut", "Musee", "Type_Visite",
        "Langue", "Nb_Personnes", "Statut", "Action"
    ]
}

# Paramètres de configuration par défaut
PARAMETRES_CONFIG = [
    ("TARIF_BRANLY", "50", "Tarif horaire Musée Branly"),
    ("TARIF_MARINE", "50", "Tarif horaire Musée Marine"),
    ("TARIF_HORS_LES_MURS", "55", "Tarif horaire visites hors-les-murs"),
    ("TARIF_EVENEMENT", "60", "Tarif horaire événements spéciaux"),
    ("TARIF_VISIO", "45", "Tarif horaire visioconférences"),
    ("TARIF_AUTRE", "50", "Tarif horaire autres prestations"),
    ("DUREE_STANDARD_BRANLY", "2", "Durée standard visite Branly (heures)"),
    ("DUREE_STANDARD_MARINE", "1.5", "Durée standard visite Marine (heures)"),
    ("DUREE_STANDARD_HORS_LES_MURS", "2", "Durée standard hors-les-murs (heures)"),
    ("DUREE_STANDARD_VISIO", "1", "Durée standard visioconférence (heures)"),
    ("EMAIL_EXPEDITEUR", "planning@musee.fr", "Email pour envoi automatique"),
    ("DELAI_NOTIFICATION_JOURS", "7", "Délai notification guides (jours avant)"),
]

def formater_entete(ws, nb_colonnes):
    """Applique un formatage professionnel aux en-têtes"""
    # Style des en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    border_style = Side(style='thin', color='000000')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for col in range(1, nb_colonnes + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

        # Ajuster la largeur de colonne
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

def normaliser_nom_colonne(nom):
    """Normalise un nom de colonne pour comparaison (sans accents, minuscules)"""
    if nom is None:
        return ""
    return str(nom).strip().lower().replace('é', 'e').replace('è', 'e').replace('ê', 'e').replace('_', '').replace(' ', '')

def ajouter_colonnes_manquantes(ws, nom_feuille, colonnes_attendues):
    """Ajoute les colonnes manquantes à une feuille existante sans perdre les données"""

    if not colonnes_attendues:
        return 0

    # Récupérer les en-têtes actuels
    headers_actuels = [cell.value for cell in ws[1]]
    headers_norm = [normaliser_nom_colonne(h) for h in headers_actuels]

    # Trouver les colonnes manquantes
    colonnes_manquantes = []
    for col_attendue in colonnes_attendues:
        col_norm = normaliser_nom_colonne(col_attendue)
        if col_norm not in headers_norm:
            colonnes_manquantes.append(col_attendue)

    if not colonnes_manquantes:
        print_succes(f"Feuille '{nom_feuille}' : Toutes les colonnes existent déjà")
        return 0

    # Ajouter les colonnes manquantes à la fin
    prochaine_col = len(headers_actuels) + 1
    nb_ajoutees = 0

    for col_name in colonnes_manquantes:
        ws.cell(row=1, column=prochaine_col, value=col_name)

        # Formater l'en-tête
        cell = ws.cell(row=1, column=prochaine_col)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Side(style='thin', color='000000')
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

        # Ajuster la largeur
        column_letter = get_column_letter(prochaine_col)
        ws.column_dimensions[column_letter].width = 15

        prochaine_col += 1
        nb_ajoutees += 1

    print_succes(f"Feuille '{nom_feuille}' : {nb_ajoutees} colonne(s) ajoutée(s) - {', '.join(colonnes_manquantes)}")
    return nb_ajoutees

def creer_feuille_avec_structure(wb, nom_feuille, colonnes):
    """Crée une feuille avec ses en-têtes ou met à jour une feuille existante"""

    if nom_feuille in wb.sheetnames:
        print_info(f"Feuille '{nom_feuille}' existe déjà")
        ws = wb[nom_feuille]

        # Pas de structure fixe (ex: Accueil)
        if not colonnes:
            print_succes(f"Feuille '{nom_feuille}' : Interface (pas de modification)")
            return ws

        # Ajouter les colonnes manquantes
        nb_ajoutees = ajouter_colonnes_manquantes(ws, nom_feuille, colonnes)

        if nb_ajoutees == 0:
            # Vérifier quand même le formatage de l'en-tête
            formater_entete(ws, ws.max_column)

        return ws
    else:
        # Créer une nouvelle feuille
        print_info(f"Création de la feuille '{nom_feuille}'...")
        ws = wb.create_sheet(nom_feuille)

        if colonnes:
            # Ajouter tous les en-têtes
            for idx, col_name in enumerate(colonnes, start=1):
                ws.cell(row=1, column=idx, value=col_name)

            # Formater les en-têtes
            formater_entete(ws, len(colonnes))

            # Figer la première ligne
            ws.freeze_panes = "A2"

            print_succes(f"Feuille '{nom_feuille}' créée avec {len(colonnes)} colonnes")
        else:
            print_succes(f"Feuille '{nom_feuille}' créée (interface)")

        return ws

def configurer_feuille_configuration(ws):
    """Ajoute les paramètres de configuration par défaut"""

    # S'assurer que la colonne Description existe
    headers = [cell.value for cell in ws[1]]
    if len(headers) < 3 or normaliser_nom_colonne(headers[2] if len(headers) > 2 else "") != normaliser_nom_colonne("Description"):
        # Ajouter la colonne Description si manquante
        if ws.max_column < 3:
            ws.cell(row=1, column=3, value="Description")
            print_info("Colonne 'Description' ajoutée à Configuration")

    # Vérifier si des paramètres existent déjà
    parametres_existants = set()
    for row in range(2, ws.max_row + 1):
        param = ws.cell(row=row, column=1).value
        if param:
            parametres_existants.add(param)

    # Ajouter les paramètres manquants
    row = ws.max_row + 1 if ws.max_row > 1 else 2
    nb_ajoutes = 0

    for param, valeur, description in PARAMETRES_CONFIG:
        if param not in parametres_existants:
            ws.cell(row=row, column=1, value=param)
            ws.cell(row=row, column=2, value=valeur)
            ws.cell(row=row, column=3, value=description)
            row += 1
            nb_ajoutes += 1

    if nb_ajoutes > 0:
        print_succes(f"Configuration : {nb_ajoutes} paramètre(s) ajouté(s)")
    else:
        print_succes("Configuration : Tous les paramètres existent déjà")

def preparer_excel(fichier_path):
    """Prépare le fichier Excel avec toutes les structures nécessaires"""

    print_titre("PREPARATION DU FICHIER EXCEL")

    # Vérifier que le fichier existe
    if not Path(fichier_path).exists():
        print_erreur(f"Fichier non trouvé : {fichier_path}")
        return False

    try:
        print_info(f"Ouverture du fichier : {fichier_path}")
        wb = openpyxl.load_workbook(fichier_path, keep_vba=True)
        print_succes("Fichier chargé avec succès (macros VBA préservées)")

        # Créer/vérifier toutes les feuilles
        print(f"\n{Couleurs.GRAS}Création/vérification des feuilles :{Couleurs.FIN}")
        for nom_feuille, colonnes in STRUCTURE_FEUILLES.items():
            ws = creer_feuille_avec_structure(wb, nom_feuille, colonnes)

            # Configuration spéciale pour la feuille Configuration
            if nom_feuille == "Configuration" and colonnes:
                configurer_feuille_configuration(ws)

        # Réorganiser l'ordre des feuilles (Accueil en premier)
        if "Accueil" in wb.sheetnames:
            wb.move_sheet("Accueil", offset=-len(wb.sheetnames)+1)
            print_succes("Feuille 'Accueil' placée en première position")

        # Sauvegarder
        print(f"\n{Couleurs.GRAS}Sauvegarde du fichier...{Couleurs.FIN}")
        wb.save(fichier_path)
        print_succes(f"Fichier sauvegardé : {fichier_path}")

        # Résumé
        print(f"\n{Couleurs.GRAS}{Couleurs.VERT}{'='*70}{Couleurs.FIN}")
        print(f"{Couleurs.GRAS}{Couleurs.VERT}PREPARATION TERMINEE AVEC SUCCES{Couleurs.FIN}")
        print(f"{Couleurs.GRAS}{Couleurs.VERT}{'='*70}{Couleurs.FIN}\n")

        print_info("✅ Toutes les feuilles nécessaires existent")
        print_info("✅ Toutes les colonnes requises ont été ajoutées")
        print_info("✅ Vos données existantes sont PRESERVEES")
        print_info("✅ Paramètres de configuration mis à jour")

        print(f"\n{Couleurs.CYAN}{Couleurs.GRAS}Prochaines étapes :{Couleurs.FIN}")
        print("  1️⃣  Ouvrez le fichier Excel")
        print("  2️⃣  Appuyez sur Alt+F11 pour ouvrir l'éditeur VBA")
        print("  3️⃣  Importez les modules .bas :")
        print("      • Fichier > Importer > Sélectionner chaque .bas du dossier vba-modules/")
        print("  4️⃣  Pour les feuilles .cls (Feuille_Mon_Planning, etc.) :")
        print("      • Trouvez la feuille correspondante dans l'arborescence VBA")
        print("      • Double-cliquez dessus")
        print("      • Copiez le contenu du fichier .cls (lignes 11 à la fin)")
        print("      • Collez dans la fenêtre de code")
        print("  5️⃣  Sauvegardez et testez !")

        print(f"\n{Couleurs.JAUNE}💡 Astuce : Commencez par importer Module_Config.bas en premier{Couleurs.FIN}")

        return True

    except Exception as e:
        print_erreur(f"Erreur lors de la préparation : {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    fichier = "PLANNING_MUSEE_TEST.xlsm"

    if len(sys.argv) > 1:
        fichier = sys.argv[1]

    fichier_path = Path(fichier)
    if not fichier_path.is_absolute():
        fichier_path = Path.cwd() / fichier

    succes = preparer_excel(str(fichier_path))
    sys.exit(0 if succes else 1)

if __name__ == "__main__":
    main()

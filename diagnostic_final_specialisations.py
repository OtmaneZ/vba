#!/usr/bin/env python3
"""
Diagnostic final: État réel de la feuille Spécialisations dans PLANNING.xlsm
"""

from pathlib import Path
from openpyxl import load_workbook

fichier = Path(__file__).parent / "PLANNING.xlsm"

print("="*80)
print("DIAGNOSTIC FINAL")
print("="*80)

wb = load_workbook(fichier, keep_vba=True)

print("\n📊 ÉTAT ACTUEL DE TOUTES LES FEUILLES:\n")

for idx, ws in enumerate(wb.worksheets, 1):
    marker = "🎯" if "special" in ws.title.lower() else "  "
    print(f"{marker} [{idx:2d}] {ws.title:25s} → {ws.sheet_state:12s} ({ws.max_row} lignes)")

spec = None
for ws in wb.worksheets:
    title_normalized = ws.title.lower().replace('é', 'e')
    if "specialisation" in title_normalized:
        spec = ws
        break

if spec:
    print(f"\n{'='*80}")
    print(f"🎯 FEUILLE TROUVÉE: '{spec.title}'")
    print(f"{'='*80}")
    print(f"État: {spec.sheet_state}")
    print(f"Index: {wb.worksheets.index(spec) + 1}")
    print(f"\n❌ PROBLÈME: État = '{spec.sheet_state}' au lieu de 'visible'")
    print(f"\n🔧 CORRECTION...")

    spec.sheet_state = "visible"
    wb.save(fichier)
    print(f"✅ Feuille '{spec.title}' maintenant en état 'visible'")
    print(f"\n💡 Fermez et rouvrez Excel, puis reconnectez-vous en admin")
else:
    print(f"\n❌ AUCUNE feuille Spécialisations trouvée dans le fichier!")
    print(f"\n🔴 Le fichier ne contient pas la feuille. Il faut la créer ou restaurer.")

wb.close()
print("="*80)

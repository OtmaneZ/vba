#!/usr/bin/env python3
"""
CORRECTION: Rendre visible la feuille Spécialisations dans PLANNING.xlsm
Elle existe mais est en état 'veryHidden' !
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
import shutil
from datetime import datetime

# Chemin du fichier Excel
fichier_planning = Path(__file__).parent / "PLANNING.xlsm"

if not fichier_planning.exists():
    print(f"❌ ERREUR: Fichier {fichier_planning} introuvable")
    sys.exit(1)

print("="*80)
print("🔧 CORRECTION: Rendre visible la feuille Spécialisations")
print("="*80)

try:
    # Backup avant modification
    backup_path = fichier_planning.parent / f"PLANNING_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
    print(f"\n💾 Création d'un backup: {backup_path.name}")
    shutil.copy2(fichier_planning, backup_path)
    print(f"   ✅ Backup créé")

    # Charger le workbook
    print(f"\n📂 Chargement de {fichier_planning.name}...")
    wb = load_workbook(fichier_planning, keep_vba=True)
    print(f"   ✅ Chargé avec succès")

    # Trouver la feuille Spécialisations
    print(f"\n🔍 Recherche de la feuille Spécialisations...")

    feuille_trouvee = False
    for ws in wb.worksheets:
        if ws.title == "Spécialisations":
            feuille_trouvee = True
            print(f"   ✅ Trouvée: '{ws.title}'")
            print(f"   État actuel: {ws.sheet_state}")
            print(f"   Contenu: {ws.max_row - 1} lignes de données")

            # Rendre visible
            if ws.sheet_state != "visible":
                print(f"\n🔓 Changement de l'état de la feuille...")
                ws.sheet_state = "visible"
                print(f"   ✅ Feuille maintenant VISIBLE")
            else:
                print(f"   ℹ️  La feuille était déjà visible")

    if not feuille_trouvee:
        print(f"   ❌ ERREUR: Feuille 'Spécialisations' non trouvée !")
        sys.exit(1)

    # Vérifier et corriger aussi les autres feuilles qui doivent être visibles pour l'admin
    print(f"\n📋 Vérification des autres feuilles pour l'admin...")

    feuilles_admin = [
        "Guides",
        "Disponibilites",
        "Visites",
        "Planning",
        "Calculs_Paie",
        "Contrats",
        "Configuration",
        "Spécialisations"
    ]

    modifications = []

    for nom_feuille in feuilles_admin:
        if nom_feuille in wb.sheetnames:
            ws = wb[nom_feuille]
            if ws.sheet_state != "visible":
                ws.sheet_state = "visible"
                modifications.append(nom_feuille)
                print(f"   🔓 {nom_feuille}: maintenant VISIBLE")
            else:
                print(f"   ✅ {nom_feuille}: déjà VISIBLE")

    # Sauvegarder
    print(f"\n💾 Sauvegarde des modifications...")
    wb.save(fichier_planning)
    wb.close()
    print(f"   ✅ Fichier sauvegardé")

    # Résumé
    print("\n" + "="*80)
    print("✅ CORRECTION TERMINÉE AVEC SUCCÈS")
    print("="*80)

    if modifications:
        print(f"\n📊 Feuilles modifiées ({len(modifications)}):")
        for feuille in modifications:
            print(f"   - {feuille}")
    else:
        print(f"\n✅ Toutes les feuilles admin étaient déjà visibles")

    print("\n💡 PROCHAINES ÉTAPES:")
    print("   1. Fermez PLANNING.xlsm si ouvert dans Excel")
    print("   2. Rouvrez PLANNING.xlsm")
    print("   3. Connectez-vous en tant qu'ADMIN")
    print("   4. Vérifiez que l'onglet Spécialisations est visible")

    print("\n⚠️  NOTE IMPORTANTE:")
    print("   Le fichier Excel avait la feuille en 'veryHidden' (masquée totalement)")
    print("   Elle est maintenant en 'visible'")
    print("   Le code VBA devrait maintenant pouvoir l'afficher correctement")

    print("\n" + "="*80)

except Exception as e:
    print(f"\n❌ ERREUR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

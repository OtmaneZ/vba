#!/usr/bin/env python3
"""
Script de verification de coherence entre les modules VBA et le fichier Excel
Verifie :
- Structure des feuilles (colonnes)
- Coherence des noms de feuilles
- Presence des fonctions VBA
- Structure des donnees
"""

import openpyxl
from openpyxl import load_workbook
import os
import re
from pathlib import Path
from collections import defaultdict

# Couleurs pour le terminal
VERT = '\033[92m'
ROUGE = '\033[91m'
JAUNE = '\033[93m'
BLEU = '\033[94m'
RESET = '\033[0m'

class VerificateurCoherence:
    def __init__(self, fichier_excel):
        self.fichier_excel = fichier_excel
        self.wb = None
        self.erreurs = []
        self.avertissements = []
        self.succes = []

        # Structures attendues selon Module_Config.bas
        self.structures_attendues = {
            'Guides': ['Prenom', 'Nom', 'Email', 'Telephone', 'Specialisations', 'Mot_De_Passe'],
            'Disponibilites': ['ID_Guide', 'Date', 'Disponible', 'Commentaire'],
            'Visites': ['ID_Visite', 'Date', 'Heure_Debut', 'Heure_Fin', 'Musee', 'Type_Visite', 'Nombre_Visiteurs'],
            'Planning': ['ID_Visite', 'Date', 'Heure', 'Type_Visite', 'Guide_Attribue', 'Guides_Disponibles', 'Statut_Confirmation', 'Historique'],
            'Calculs_Paie': ['ID_Guide', 'Nom_Complet', 'Nb_Visites', 'Nb_Jours', 'Montant_Total', 'Montant/Cachet', 'Total_Recalcule'],
            'Contrats': ['ID_Guide', 'Nom', 'Mois', 'Type_Contrat', 'Dates_Visites', 'Nb_Cachets', 'Montant_Cachet', 'Total'],
            'Configuration': ['Parametre', 'Valeur']
        }

        self.feuilles_obligatoires = [
            'Accueil', 'Guides', 'Disponibilites', 'Visites',
            'Planning', 'Calculs_Paie', 'Contrats', 'Configuration'
        ]

    def charger_fichier(self):
        """Charge le fichier Excel"""
        try:
            if not os.path.exists(self.fichier_excel):
                self.erreurs.append(f"Fichier introuvable : {self.fichier_excel}")
                return False

            print(f"{BLEU}[INFO]{RESET} Chargement du fichier : {self.fichier_excel}")
            self.wb = load_workbook(self.fichier_excel, data_only=True)
            self.succes.append("Fichier Excel charge avec succes")
            return True
        except Exception as e:
            self.erreurs.append(f"Erreur chargement fichier : {str(e)}")
            return False

    def verifier_feuilles(self):
        """Verifie la presence de toutes les feuilles obligatoires"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Presence des feuilles...")

        feuilles_presentes = self.wb.sheetnames

        for feuille in self.feuilles_obligatoires:
            if feuille in feuilles_presentes:
                self.succes.append(f"Feuille '{feuille}' : PRESENTE")
            else:
                self.erreurs.append(f"Feuille manquante : '{feuille}'")

        # Feuilles supplementaires
        feuilles_extra = set(feuilles_presentes) - set(self.feuilles_obligatoires)
        if feuilles_extra:
            for extra in feuilles_extra:
                self.avertissements.append(f"Feuille supplementaire : '{extra}'")

    def verifier_structure_feuille(self, nom_feuille):
        """Verifie la structure d'une feuille (colonnes)"""
        if nom_feuille not in self.wb.sheetnames:
            return

        if nom_feuille not in self.structures_attendues:
            return

        ws = self.wb[nom_feuille]
        structure_attendue = self.structures_attendues[nom_feuille]

        # Lire la premiere ligne (en-tetes)
        en_tetes_trouvees = []
        for col in range(1, len(structure_attendue) + 2):
            valeur = ws.cell(1, col).value
            if valeur:
                en_tetes_trouvees.append(str(valeur).strip())

        # Comparer
        erreurs_structure = []
        for idx, attendu in enumerate(structure_attendue, start=1):
            if idx <= len(en_tetes_trouvees):
                trouve = en_tetes_trouvees[idx - 1]
                if trouve != attendu:
                    erreurs_structure.append(
                        f"  Colonne {idx}: attendu '{attendu}', trouve '{trouve}'"
                    )
            else:
                erreurs_structure.append(
                    f"  Colonne {idx}: manquante (attendu '{attendu}')"
                )

        if erreurs_structure:
            self.erreurs.append(f"Structure '{nom_feuille}' incorrecte :")
            self.erreurs.extend(erreurs_structure)
        else:
            self.succes.append(f"Structure '{nom_feuille}' : CORRECTE ({len(structure_attendue)} colonnes)")

    def verifier_toutes_structures(self):
        """Verifie toutes les structures de feuilles"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Structure des feuilles...")

        for nom_feuille in self.structures_attendues.keys():
            self.verifier_structure_feuille(nom_feuille)

    def verifier_donnees_guides(self):
        """Verifie la coherence des donnees dans Guides"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Donnees feuille Guides...")

        if 'Guides' not in self.wb.sheetnames:
            return

        ws = self.wb['Guides']
        nb_guides = 0
        guides_sans_email = []
        guides_sans_mdp = []

        for row in range(2, ws.max_row + 1):
            prenom = ws.cell(row, 1).value
            nom = ws.cell(row, 2).value
            email = ws.cell(row, 3).value
            mdp = ws.cell(row, 6).value

            if prenom or nom:
                nb_guides += 1
                nom_complet = f"{prenom or ''} {nom or ''}".strip()

                if not email:
                    guides_sans_email.append(nom_complet)

                if not mdp:
                    guides_sans_mdp.append(nom_complet)

        self.succes.append(f"Nombre de guides : {nb_guides}")

        if guides_sans_email:
            self.avertissements.append(f"Guides sans email : {', '.join(guides_sans_email)}")

        if guides_sans_mdp:
            self.avertissements.append(f"Guides sans mot de passe : {', '.join(guides_sans_mdp)}")

    def verifier_configuration(self):
        """Verifie la presence des parametres de configuration"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Parametres de configuration...")

        if 'Configuration' not in self.wb.sheetnames:
            return

        ws = self.wb['Configuration']

        # Parametres tarifaires attendus
        params_tarifaires = [
            'TARIF_1_VISITE', 'TARIF_2_VISITES', 'TARIF_3_VISITES',
            'TARIF_BRANLY_2H', 'TARIF_BRANLY_3H', 'TARIF_BRANLY_4H',
            'TARIF_HORSLEMURS_1', 'TARIF_HORSLEMURS_2', 'TARIF_HORSLEMURS_3'
        ]

        params_trouves = []
        for row in range(2, ws.max_row + 1):
            param = ws.cell(row, 1).value
            if param:
                params_trouves.append(str(param).strip().upper())

        params_manquants = []
        for param in params_tarifaires:
            if param not in params_trouves:
                params_manquants.append(param)

        if params_manquants:
            self.avertissements.append(
                f"Parametres tarifaires manquants : {', '.join(params_manquants)}"
            )
        else:
            self.succes.append("Tous les parametres tarifaires sont presents")

    def verifier_visites(self):
        """Verifie les donnees des visites"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Donnees feuille Visites...")

        if 'Visites' not in self.wb.sheetnames:
            return

        ws = self.wb['Visites']
        nb_visites = 0
        types_visites = set()

        for row in range(2, ws.max_row + 1):
            id_visite = ws.cell(row, 1).value
            type_visite = ws.cell(row, 6).value

            if id_visite:
                nb_visites += 1
                if type_visite:
                    types_visites.add(str(type_visite))

        self.succes.append(f"Nombre de visites : {nb_visites}")
        if types_visites:
            self.succes.append(f"Types de visites trouves : {', '.join(sorted(types_visites))}")

    def verifier_planning(self):
        """Verifie le planning"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Donnees feuille Planning...")

        if 'Planning' not in self.wb.sheetnames:
            return

        ws = self.wb['Planning']
        nb_attributions = 0
        nb_non_attribue = 0

        for row in range(2, ws.max_row + 1):
            guide = ws.cell(row, 5).value

            if guide:
                if str(guide).upper() == "NON ATTRIBUE":
                    nb_non_attribue += 1
                else:
                    nb_attributions += 1

        self.succes.append(f"Visites attribuees : {nb_attributions}")
        if nb_non_attribue > 0:
            self.avertissements.append(f"Visites non attribuees : {nb_non_attribue}")

    def analyser_modules_vba(self):
        """Analyse les modules VBA pour verifier la coherence"""
        print(f"\n{BLEU}[VERIFICATION]{RESET} Modules VBA...")

        vba_dir = Path(__file__).parent / "vba-modules"
        if not vba_dir.exists():
            self.avertissements.append("Dossier vba-modules introuvable")
            return

        modules_trouves = list(vba_dir.glob("*.bas")) + list(vba_dir.glob("*.cls"))
        self.succes.append(f"Modules VBA trouves : {len(modules_trouves)}")

        # Verifier les references aux colonnes dans les modules
        problemes_colonnes = []
        for module in modules_trouves:
            with open(module, 'r', encoding='utf-8') as f:
                contenu = f.read()

                # Chercher les references obsoletes aux colonnes
                if 'wsGuides.Cells(i, 4).Value' in contenu and 'Email' not in contenu[:contenu.find('wsGuides.Cells(i, 4)')]:
                    problemes_colonnes.append(f"{module.name}: Reference colonne 4 pour Email (devrait etre 3)")

                if '.Cells(i, 2).Value & " " & .Cells(i, 3).Value' in contenu:
                    # Verifier si c'est pour Nom+Email (faux) ou Prenom+Nom (correct)
                    if 'wsGuides' in contenu[:contenu.find('.Cells(i, 2)')]:
                        problemes_colonnes.append(f"{module.name}: Possible reference colonnes 2+3 pour nom (verifier contexte)")

        if problemes_colonnes:
            for prob in problemes_colonnes:
                self.avertissements.append(f"VBA: {prob}")

    def generer_rapport(self):
        """Genere le rapport final"""
        print("\n" + "=" * 80)
        print(f"{BLEU}RAPPORT DE VERIFICATION DE COHERENCE{RESET}")
        print("=" * 80)

        if self.succes:
            print(f"\n{VERT}✓ SUCCES ({len(self.succes)}){RESET}")
            for msg in self.succes:
                print(f"  {VERT}✓{RESET} {msg}")

        if self.avertissements:
            print(f"\n{JAUNE}⚠ AVERTISSEMENTS ({len(self.avertissements)}){RESET}")
            for msg in self.avertissements:
                print(f"  {JAUNE}⚠{RESET} {msg}")

        if self.erreurs:
            print(f"\n{ROUGE}✗ ERREURS ({len(self.erreurs)}){RESET}")
            for msg in self.erreurs:
                print(f"  {ROUGE}✗{RESET} {msg}")

        print("\n" + "=" * 80)

        # Resultat global
        if self.erreurs:
            print(f"{ROUGE}RESULTAT : ECHEC - {len(self.erreurs)} erreur(s) critique(s){RESET}")
            return False
        elif self.avertissements:
            print(f"{JAUNE}RESULTAT : REUSSITE AVEC AVERTISSEMENTS - {len(self.avertissements)} avertissement(s){RESET}")
            return True
        else:
            print(f"{VERT}RESULTAT : PARFAIT - Aucun probleme detecte{RESET}")
            return True

    def executer(self):
        """Execute toutes les verifications"""
        if not self.charger_fichier():
            self.generer_rapport()
            return False

        self.verifier_feuilles()
        self.verifier_toutes_structures()
        self.verifier_donnees_guides()
        self.verifier_configuration()
        self.verifier_visites()
        self.verifier_planning()
        self.analyser_modules_vba()

        return self.generer_rapport()


def main():
    print(f"{BLEU}=== VERIFICATEUR DE COHERENCE VBA/EXCEL ==={RESET}\n")

    # Trouver le fichier Excel
    fichier = Path(__file__).parent / "PLANNING_MUSEE_FINAL_PROPRE.xlsm"

    if not fichier.exists():
        print(f"{ROUGE}[ERREUR]{RESET} Fichier Excel introuvable : {fichier}")
        return 1

    verificateur = VerificateurCoherence(str(fichier))
    resultat = verificateur.executer()

    return 0 if resultat else 1


if __name__ == "__main__":
    exit(main())

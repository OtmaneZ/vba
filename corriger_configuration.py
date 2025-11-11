#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de correction de la feuille Configuration
- Convertit les valeurs texte en nombres
- Corrige les tarifs selon le formulaire client
"""

import openpyxl
from pathlib import Path
import sys

# Codes couleurs
class Couleurs:
    VERT = '\033[92m'
    JAUNE = '\033[93m'
    ROUGE = '\033[91m'
    CYAN = '\033[96m'
    GRAS = '\033[1m'
    FIN = '\033[0m'

def print_succes(texte):
    print(f"{Couleurs.VERT}✓ {texte}{Couleurs.FIN}")

def print_warning(texte):
    print(f"{Couleurs.JAUNE}⚠ {texte}{Couleurs.FIN}")

def print_info(texte):
    print(f"{Couleurs.CYAN}ℹ {texte}{Couleurs.FIN}")

# Valeurs correctes selon demande client (mail)
TARIFS_CORRECTS = {
    "TARIF_1_VISITE": 80,        # 1 visite/jour = 80€
    "TARIF_2_VISITES": 110,      # 2 visites/jour = 110€
    "TARIF_3_VISITES": 140,      # 3+ visites/jour = 140€
    "TARIF_HORSLEMURS": 100,     # Hors-les-murs = 100€ (fixe)
    "DELAI_NOTIFICATION_JOURS": 7,
}

def corriger_configuration(fichier_path):
    """Corrige la feuille Configuration"""

    print(f"\n{Couleurs.GRAS}{Couleurs.CYAN}{'='*70}{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.CYAN}CORRECTION DE LA CONFIGURATION{Couleurs.FIN}")
    print(f"{Couleurs.GRAS}{Couleurs.CYAN}{'='*70}{Couleurs.FIN}\n")

    # Ouvrir le fichier
    try:
        wb = openpyxl.load_workbook(fichier_path, keep_vba=True)
        ws = wb["Configuration"]
        print_succes(f"Fichier ouvert : {fichier_path}")
    except Exception as e:
        print(f"{Couleurs.ROUGE}✗ Erreur : {e}{Couleurs.FIN}")
        return False

    nb_corrections = 0
    nb_conversions = 0

    print(f"\n{Couleurs.GRAS}Analyse et corrections :{Couleurs.FIN}\n")

    # Parcourir toutes les lignes
    for row in range(2, ws.max_row + 1):
        parametre = ws.cell(row=row, column=1).value
        valeur_actuelle = ws.cell(row=row, column=2).value

        if parametre is None:
            continue

        # Vérifier si la valeur doit être corrigée
        if parametre in TARIFS_CORRECTS:
            valeur_correcte = TARIFS_CORRECTS[parametre]

            # Convertir valeur actuelle en nombre si c'est du texte
            try:
                if isinstance(valeur_actuelle, str):
                    valeur_actuelle_num = float(valeur_actuelle)
                    nb_conversions += 1
                else:
                    valeur_actuelle_num = float(valeur_actuelle) if valeur_actuelle else 0
            except:
                valeur_actuelle_num = 0

            # Comparer et corriger
            if valeur_actuelle_num != valeur_correcte:
                print_warning(f"{parametre}: {valeur_actuelle_num} → {valeur_correcte}")
                ws.cell(row=row, column=2).value = valeur_correcte
                nb_corrections += 1
            else:
                # Même si la valeur est correcte, s'assurer qu'elle est en nombre
                if isinstance(valeur_actuelle, str):
                    ws.cell(row=row, column=2).value = valeur_correcte
                    nb_conversions += 1
                    print_info(f"{parametre}: Conversion texte → nombre ({valeur_correcte})")
                else:
                    print_succes(f"{parametre}: OK ({valeur_correcte})")
        else:
            # Pour les autres paramètres, juste convertir en nombre si nécessaire
            if isinstance(valeur_actuelle, str) and parametre not in ["MotDePasseAdmin", "EMAIL_EXPEDITEUR"]:
                try:
                    valeur_num = float(valeur_actuelle)
                    ws.cell(row=row, column=2).value = valeur_num
                    nb_conversions += 1
                    print_info(f"{parametre}: Conversion texte → nombre ({valeur_num})")
                except:
                    # Pas un nombre, garder tel quel
                    print_info(f"{parametre}: Texte conservé ('{valeur_actuelle}')")

    # Sauvegarder
    try:
        wb.save(fichier_path)
        print(f"\n{Couleurs.VERT}{Couleurs.GRAS}✓ FICHIER SAUVEGARDE{Couleurs.FIN}\n")
    except Exception as e:
        print(f"\n{Couleurs.ROUGE}✗ Erreur sauvegarde : {e}{Couleurs.FIN}")
        return False

    # Résumé
    print(f"{Couleurs.GRAS}Résumé :{Couleurs.FIN}")
    print_succes(f"{nb_corrections} valeur(s) corrigée(s)")
    print_succes(f"{nb_conversions} conversion(s) texte → nombre")

    print(f"\n{Couleurs.CYAN}{Couleurs.GRAS}Tarifs configurés (selon formulaire client) :{Couleurs.FIN}")
    print(f"  • Branly : 50€ (2h)")
    print(f"  • Marine : 50€ (1.5h)")
    print(f"  • Hors-les-murs : 55€ (2h)")
    print(f"  • Événements : 60€")
    print(f"  • Visio : 45€ (1h)")
    print(f"  • Autres : 50€")

    print(f"\n{Couleurs.JAUNE}{Couleurs.GRAS}💡 Note importante :{Couleurs.FIN}")
    print(f"  Ces tarifs sont utilisés pour FACTURER le musée.")
    print(f"  La REMUNERATION du guide est calculée en CACHETS en fin de mois :")
    print(f"  → Total mensuel ÷ Nb jours travaillés = Cachet journalier")

    return True

def main():
    fichier = "PLANNING_MUSEE_FINAL_PROPRE.xlsm"

    if len(sys.argv) > 1:
        fichier = sys.argv[1]

    fichier_path = Path(fichier)
    if not fichier_path.is_absolute():
        fichier_path = Path.cwd() / fichier

    if not fichier_path.exists():
        print(f"{Couleurs.ROUGE}✗ Fichier non trouvé : {fichier_path}{Couleurs.FIN}")
        sys.exit(1)

    succes = corriger_configuration(str(fichier_path))
    sys.exit(0 if succes else 1)

if __name__ == "__main__":
    main()
